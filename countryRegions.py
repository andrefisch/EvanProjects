from importDict import importDict as impDi, number_from_column
from openpyxl.utils import get_column_letter as gcl
import openpyxl
import os
import pygame
import sys
import time

# GETS ALL COLUMNS FROM THIS ROW
# print(sheet[1])

def countryRegions():
    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    fileName = args[0]
    cols = args[1:]

    dicty = impDi('../../regions.txt')

    # Open an existing excel file
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    #################
    # DO STUFF HERE #
    #################
    first = 2
    last = sheet.max_row + 1
    changes = 0
    for col in cols:
        for row in range (first, last):
            if sheet[col + str(row)].value:
                info = str(sheet[col + str(row)].value).strip().title()
                if info in dicty:
                    changes = changes + 1
                    region = dicty[info]
                    sheet[gcl(number_from_column(col) + 1) + str(row)].value = region
                    print(col + str(row) + ":", info, '->', region)

    print("Processed " + str((last - first) * len(cols)) + " rows...")
    print("Changed   " + str(changes) + " values...")

    # add the word 'formatted' and save the new file where the original is
    newName = 'regions'
    index = fileName[::-1].find('/')
    end = fileName[-index - 1:]
    fileName = fileName[:-index - 1] + newName + end[0].capitalize() + end[1:]
    print("Saving " + fileName)
    wb.save(fileName)

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

countryRegions()
