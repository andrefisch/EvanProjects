from openpyxl.utils import get_column_letter as gcl
import openpyxl
import pygame
import re
import string
import sys
import time

printing = True
saving = True

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# Remove all spaces from the last n nonspace characters of the given string
# remove_end_space(string, num)
#{{{
def remove_end_space(string, num):
    string2 = string[::-1]
    count = 0
    index = 0
    for i in range(0, len(string2)):
        if string2[i] != " ":
            count += 1
        if count >= num:
            index = i
            break
    first = string2[index:]
    second = string2[:index]
    second = second.replace(" ", "")
    return (second + first)[::-1]
#}}}

# number_from_column(column_letter)
# {{{
def number_from_column(column_letter):
    return ord(column_letter) - 64
#}}}

# Extract first, last, and middle name from a name with more than 3 parts
# determine_names(listy)
#{{{
def determine_names(listy):
    dicty   = {}
    lasty   = []
    middley = []
    # first spot is always first name at this point
    dicty['first_name'] = listy[0]
    del listy[0]
    # - reverse list 
    # - take first name in reversed list (last name) and add it to last name list, delete it
    # - look at next name and see if it is capitalized
    #   - if not add to last name list, repeat
    #   - otherwise add this and rest to middle name list
    listy = listy[::-1]
    lasty.append(listy[0])
    del listy[0]
    lasts = True
    for i in range(0, len(listy)):
        if (not listy[i].istitle()) and lasts:
            lasty.insert(0, listy[i])
        else:
            lasts = False
            middley.insert(0, listy[i])

    dicty['middle_name'] = ' '.join(middley)
    dicty['last_name'] = ' '.join(lasty)
    return dicty
#}}}

# formatting_name(name)
#{{{
def formatting_name(name):
    original = name
    names = {"first_name": "", "middle_name": "", "last_name": "", "appellation": ""}

    # remove unnecessary suffixes
    regexSuffix = ',? (Jr\.? ?|Sr\.? ?|Ph\.? ?D\.? ?|P\.?Ehj)'
    name = re.sub(regexSuffix, "", name, re.IGNORECASE)

    # extract appellation 
    regexAppellation = '^(Mr\.?|Mrs\.?|Ms\.?|Rev\.?|Hon\.?|Dr\.?|Captain|Capt?\.?|Dcn\.?|Amb\.?|Lt\.?|MIDN\.?|Miss\.?|Fr\.?) ?(.*)'
    # number = re.sub(regexAppellation, "", number)
    # regex0 = '^0+(.*)'
    matchAppellation = re.search(regexAppellation, name, re.IGNORECASE)
    if matchAppellation:
        names['appellation'] = matchAppellation.group(1)
        name = matchAppellation.group(2)

    # split names
    matchList = []
    regexName = "([\w+\.-]+)"
    while True:
        matchName = re.search(regexName, name)
        if matchName:
            matchList.append(matchName.group(1))
            name = name.replace(matchName.group(1), "")
        else:
            break

    # if there are only one, two, or three names in the list it is easy
    nameLen = len(matchList)
    if nameLen == 1:
        names['last_name'] = matchList[0].strip()
    elif nameLen == 2:
        names['first_name'] = matchList[0].strip()
        names['last_name'] = matchList[-1].strip()
    elif nameLen == 3:
        names['first_name'] = matchList[0].strip()
        names['middle_name'] = matchList[1].strip()
        names['last_name'] = matchList[-1].strip()
    else:
        names = {**names, **determine_names(matchList)}

    if printing:
        if original != None:
            print(original + " -> " + names['appellation'] + "-" + names['first_name'] + "-" + names['middle_name'] + "-" + names['last_name'])

    return names
#}}}

# Format all numbers in a column
# format_all_numbers()
#{{{
def format_all_names():
    if printing:
        print("Opening...")

    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    filename = args[0]
    cols = args[1:]

    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]

    first = 2
    last = sheet.max_row + 1

    for col in cols:
        for row in range (first, last):
            name = str(sheet[col + str(row)].value)
            formatted = formatting_name(name)
            if saving:
                sheet[gcl(number_from_column(col) + 1) + str(row)].value = formatted['appellation']
                sheet[gcl(number_from_column(col) + 2) + str(row)].value = formatted['first_name']
                sheet[gcl(number_from_column(col) + 3) + str(row)].value = formatted['middle_name']
                sheet[gcl(number_from_column(col) + 4) + str(row)].value = formatted['last_name']

    if printing:
        print("Processing " + str(last - first) + " rows...")

    if printing and saving:
        print("Saving...")
        wb.save("splitNames.xlsx")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# test stuff
#{{{
name1 = "Mr. Evangelos Efstathiou, Ph.D."
name2 = "Mr. Bradley D.M. Golden"
name3 = "Mr. Jan-Willem Ovind van den Dijssel"

# print(formatting_name(name1))
# print(formatting_name(name2))
# print(formatting_name(name3))
#}}}

format_all_names()
