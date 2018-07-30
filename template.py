import openpyxl
import os
import pygame
import sys
import time

# GETS ALL COLUMNS FROM THIS ROW
# print(sheet[1])

def template():
    # Get all files of a specific type in this directory
    files = [x for x in os.listdir() if x.endswith(".xlsx")]
    for eachfile in files:
        print("----------------------")
        print(eachfile)
        print("----------------------")

    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    fileName = args[0]
    first = int(args[1])
    cols = args[2:]

    # Open a file with sys.argv
    with open(sys.argv[1]) as f:
        print(f)





    # CREATE A NEW EXCEL FILE
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("transactions")

    # if 'sheet' appears randomly we can delete it
    rm = out['Sheet']
    out.remove(rm)

    #################
    # DO STUFF HERE #
    #################
    first = 2
    last = outsheet.max_row + 1
    changes = 0
    for col in cols:
        for row in range (first, last):
            changes = changes + 1
            outsheet['A1'].value = "DATA GOES HERE"


    print("Processed " + str((last - first) * len(cols)) + " rows...")
    print("Changed   " + str(changes) + " values...")

    # add the word 'formatted' and save the new file where the original is
    newName = 'new'
    index = fileName[::-1].find('/')
    end = fileName[-index - 1:]
    fileName = fileName[:-index - 1] + newName + end[0].capitalize() + end[1:]
    print("Saving " + fileName)
    wb.save(fileName)





    # OPEN AN EXISTING EXCEL FILE
    wb = openpyxl.load_workbook(fileName)
    # Usually you just want to use this option
    sheet = wb.worksheets[0]
    # Don't usually want to use this one
    sheet = wb[sheetName]

    #################
    # DO STUFF HERE #
    #################
    first = 2
    last = outsheet.max_row + 1
    changes = 0
    for col in cols:
        for row in range (first, last):
            changes = changes + 1
            sheet['A1'].value = "DATA GOES HERE"


    print("Processed " + str((last - first) * len(cols)) + " rows...")
    print("Changed   " + str(changes) + " values...")

    # add the word 'improved' and save the new file where the original is
    newName = 'improved'
    index = fileName[::-1].find('/')
    end = fileName[-index - 1:]
    fileName = fileName[:-index - 1] + newName + end[0].capitalize() + end[1:]
    print("Saving " + fileName)
    wb.save(fileName)

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

template()
