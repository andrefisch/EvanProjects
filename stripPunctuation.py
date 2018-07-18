from openpyxl.utils import get_column_letter as gcl
import openpyxl
import pygame
import re
import string
import sys
import time

printing = True
saving = True

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        return word.translate(punctuationTable)
    else:
        return ""
#}}}

# number_from_column(column_letter)
# {{{
def number_from_column(column_letter):
    return ord(column_letter) - 64
#}}}

#{{{
def strip_punctuation():
    if printing:
        print("Opening...")

    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    filename = args[0]
    cols = args[1:]

    # if len(args) == 0:
        # print(

    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]

    first = 2
    last = sheet.max_row + 1
    changes = 0

    for col in cols:
        for row in range (first, last):
            word = str(sheet[col + str(row)].value)
            if word != '' and word != 'None' and word != 'Null':
                formatted = standardize_str(word)
                if word != formatted:
                    changes = changes + 1
                    print(col + str(row) + ": ", word, '->', formatted)

                if saving:
                    sheet[col + str(row)].value = formatted.strip()

    if printing:
        print("Processed " + str((last - first) * len(cols)) + " rows...")
        print("Changed   " + str(changes) + " values...")

    if printing and saving:
        print("Saving...")
        wb.save("noPuctuation.xlsx")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

strip_punctuation()
