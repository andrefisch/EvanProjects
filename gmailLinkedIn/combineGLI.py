# which column is the unique identifier (primary key)
# in event of two names having different data in the same column (collision) should I take most recent?
# you have some accounts which only seem to have multiple rows because they are present in more than one industry or segment. do you want to have all of that information represented in the same cell?

# want to eliminate duplicate companies but keep all information
# eliminate duplicate companies and people

# for first part use contact as unique identifier
# for second part use name as unique identifier

from datetime import datetime
from splitNames import determine_names
import openpyxl
import pygame
import re
import string
import sys
import time

saving = True
printing = True
testing = True
low_threshold = 85
high_threshold = 100
min_word_len = 5

'''
FIRST_NAME           = 'A'
MIDDLE_NAME          = 'B'
LAST_NAME            = 'C'
TITLE                = 'D'
SUFFIX               = 'E'
WEB_PAGE             = 'F'
NOTES                = 'G'
EMAIL_ADDRESS        = 'H'
EMAIL_ADDRESS2       = 'I'
EMAIL_ADDRESS3       = 'J'
HOME_PHONE           = 'K'
MOBILE_PHONE         = 'L'
HOME_ADDRESS         = 'M'
HOME_STREET          = 'N'
HOME_CITY            = 'O'
HOME_STATE           = 'P'
HOME_POSTAL_CODE     = 'Q'
HOME_COUNTRY         = 'R'
CONTACT_MAIN_PHONE   = 'S'
BUSINESS_PHONE       = 'T'
BUSINESS_PHONE2      = 'U'
BUSINESS_FAX         = 'V'
COMPANY              = 'W'
JOB_TITLE            = 'X'
DEPARTMENT           = 'Y'
OFFICE_LOCATION      = 'Z'
BUSINESS_ADDRESS     = 'AA'
BUSINESS_STREET      = 'AB'
BUSINESS_CITY        = 'AC'
BUSINESS_STATE       = 'AD'
BUSINESS_POSTAL_CODE = 'AE'
BUSINESS_COUNTRY     = 'AF'
CATEGORIES           = 'AG'
CONNECTED_ON         = 'AH'
LATITUDE             = 'AI'
LONGITUDE            = 'AJ'
'''

NAME                          = "A"
GIVEN_NAME                    = "B"
ADDITIONAL_NAME               = "C"
FAMILY_NAME                   = "D"
YOMI_NAME                     = "E"
GIVEN_NAME_YOMI               = "F"
ADDITIONAL_NAME_YOMI          = "G"
FAMILY_NAME_YOMI              = "H"
NAME_PREFIX                   = "I"
NAME_SUFFIX                   = "J"
INITIALS                      = "K"
NICKNAME                      = "L"
SHORT_NAME                    = "M"
MAIDEN_NAME                   = "N"
BIRTHDAY                      = "O"
GENDER                        = "P"
LOCATION                      = "Q"
BILLING_INFORMATION           = "R"
DIRECTORY_SERVER              = "S"
MILEAGE                       = "T"
OCCUPATION                    = "U"
HOBBY                         = "V"
SENSITIVITY                   = "W"
PRIORITY                      = "X"
SUBJECT                       = "Y"
NOTES                         = "Z"
GROUP_MEMBERSHIP              = "AA"
EMAIL1_TYPE                   = "AB"
EMAIL1_VALUE                  = "AC"
EMAIL2_TYPE                   = "AD"
EMAIL2_VALUE                  = "AE"
EMAIL3_TYPE                   = "AF"
EMAIL3_VALUE                  = "AG"
EMAIL4_TYPE                   = "AH"
EMAIL4_VALUE                  = "AI"
EMAIL5_TYPE                   = "AJ"
EMAIL5_VALUE                  = "AK"
IM1_TYPE                      = "AL"
IM1_SERVICE                   = "AM"
IM1_VALUE                     = "AN"
PHONE1_TYPE                   = "AO"
PHONE1_VALUE                  = "AP"
PHONE2_TYPE                   = "AQ"
PHONE2_VALUE                  = "AR"
PHONE3_TYPE                   = "AS"
PHONE3_VALUE                  = "AT"
PHONE4_TYPE                   = "AU"
PHONE4_VALUE                  = "AV"
PHONE5_TYPE                   = "AW"
PHONE5_VALUE                  = "AX"
ADDRESS1_TYPE                 = "AY"
ADDRESS1_FORMATED             = "AZ"
ADDRESS1_STREET               = "BA"
ADDRESS1_CITY                 = "BB"
ADDRESS1_POBOX                = "BC"
ADDRESS1_REGION               = "BD"
ADDRESS1_POSTAL_CODE          = "BE"
ADDRESS1_COUNTRY              = "BF"
ADDRESS1_EXTENDED_ADDRESS     = "BG"
ADDRESS2_TYPE                 = "BH"
ADDRESS2_FORMATED             = "BI"
ADDRESS2_STREET               = "BJ"
ADDRESS2_CITY                 = "BK"
ADDRESS2_POBOX                = "BL"
ADDRESS2_REGION               = "BM"
ADDRESS2_POSTAL_CODE          = "BN"
ADDRESS2_COUNTRY              = "BO"
ADDRESS2_EXTENDED_ADDRESS     = "BP"
ADDRESS3_TYPE                 = "BQ"
ADDRESS3_FORMATED             = "BR"
ADDRESS3_STREET               = "BS"
ADDRESS3_CITY                 = "BT"
ADDRESS3_POBOX                = "BU"
ADDRESS3_REGION               = "BV"
ADDRESS3_POSTAL_CODE          = "BW"
ADDRESS3_COUNTRY              = "BX"
ADDRESS3_EXTENDED_ADDRESS     = "BY"
ORGANIZATION1_TYPE            = "BZ"
ORGANIZATION1_NAME            = "CA"
ORGANIZATION1_YOMI_NAME       = "CB"
ORGANIZATION1_TITLE           = "CC"
ORGANIZATION1_DEPARTMENT      = "CD"
ORGANIZATION1_SYMBOL          = "CE"
ORGANIZATION1_LOCATION        = "CF"
ORGANIZATION1_JOB_DESCRIPTION = "CG"
RELATION1_TYPE                = "CH"
RELATION1_VALUE               = "CI"
EXTERNAL_ID1_TYPE             = "CJ"
EXTERNAl_ID1_VALUE            = "CK"
WEBSITE1_TYPE                 = "CL"
WEBSITE1_VALUE                = "CM"
CALENDAR_LINK1_TYPE           = "CN"
CALENDAR_LINK1_VALUE          = "CO"
JOT1_TYPE                     = "CP"
JOT1_VALUE                    = "CQ"


##################
# HELPER METHODS #
##################

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# Combine with slash
# special_combine(word1, word2)
# {{{
def special_combine(word1, word2):
    if word1 == "" or word1 == None:
        return word2
    elif word2 == "" or word2 == None:
        return word1
    elif word1 in word2:
        return word2
    elif word2 in word1:
        return word1
    else:
        return word1 + "/" + word2
#}}}

# Add to notes
# add_to_notes(category, info)
#{{{
def add_to_notes(contact, category, info, oldInfo):
    if info != "" and info != None and oldInfo != "" and oldInfo != None and info != oldInfo:
        return str(contact.notes) + "; " + str(category) + " used to be " + str(info)
    else:
        return str(contact.notes)
#}}}

# Keep the non-none value
# keep_non_none(var1, var2)
#{{{
def keep_non_none(var1, var2):
    if var1 == None or var1 == "":
        return (var2, var1)
    elif var2 == None or var2 == "":
        return (var1, var2)
    else:
        return (var2, var1)
#}}}

# Edit Distance Function
# edit_distance(word1, word2, low_threshold, high_threshold)
#{{{
# modified edit distance algorithm:
# - see if string one is a substring of the next string
#   - if it is there is a match
#   - it not look at edit distance for substring and substring of equal length
#     - if it is above a certain percentage or below a certain edit number we can safely assume they are a match. keep that string and try with next one

def edit_distance(word1, word2, low_threshold, high_threshold):
    if printing:
        print("Looking at '" + word1 + "' and '" + word2 + "'")
    word1 = word1.lower()
    word2 = word2.lower()
    len_1 = len(word1)
    len_2 = len(word2)
    edit_distance, percent_match = 0, 0

    # make sure shorter word is first
    if len_1 > len_2:
        word1, word2 = word2, word1
        len_1 = len_2

    if (len_1 >= min_word_len):
        if (word1 in word2):
            if printing:
                print ("Edit distance: 0")
                print ("Percent Match: 100")
            edit_distance, percent_match = 0, 100
        else:
            # shorten longer word to length of first word
            word2 = word2[0:len_1]
            # the matrix whose last element -> edit distance
            x = [[0] * (len_1 + 1) for _ in range(len_1 + 1)]

            # initialization of base case values
            for i in range(0, len_1 + 1): 
                x[i][0] = i
            for j in range(0, len_1 + 1):
                x[0][j] = j
            for i in range (1, len_1 + 1):
                for j in range(1, len_1 + 1):
                    if word1[i - 1] == word2[j - 1]:
                        x[i][j] = x[i - 1][j - 1] 
                    else:
                        x[i][j]= min(x[i][j - 1], x[i - 1][j], x[i - 1][j - 1]) + 1
            edit_distance = x[i][j]
            percent_match = ((len_1 - edit_distance) / len_1) * 100
            if printing:
                print ("Edit distance " + str(x[i][j]))
                print ("Percent match: " + "%.2f" % percent_match)
    if percent_match > low_threshold and percent_match <= high_threshold:
        if printing:
            print("MATCH!")
        return True
    else:
        return False
#}}}

# Contact object with all relevant fields
# Contact
#{{{
class Contact(object):
    name                          = ''
    given_name                    = ''
    additional_name               = ''
    family_name                   = ''
    yomi_name                     = ''
    given_name_yomi               = ''
    additional_name_yomi          = ''
    family_name_yomi              = ''
    name_prefix                   = ''
    name_suffix                   = ''
    initials                      = ''
    nickname                      = ''
    short_name                    = ''
    maiden_name                   = ''
    birthday                      = ''
    gender                        = ''
    location                      = ''
    billing_information           = ''
    directory_server              = ''
    mileage                       = ''
    occupation                    = ''
    hobby                         = ''
    sensitivity                   = ''
    priority                      = ''
    subject                       = ''
    notes                         = ''
    group_membership              = ''
    email1_type                   = ''
    email1_value                  = ''
    email2_type                   = ''
    email2_value                  = ''
    email3_type                   = ''
    email3_value                  = ''
    email4_type                   = ''
    email4_value                  = ''
    email5_type                   = ''
    email5_value                  = ''
    im1_type                      = ''
    im1_service                   = ''
    im1_value                     = ''
    phone1_type                   = ''
    phone1_value                  = ''
    phone2_type                   = ''
    phone2_value                  = ''
    phone3_type                   = ''
    phone3_value                  = ''
    phone4_type                   = ''
    phone4_value                  = ''
    phone5_type                   = ''
    phone5_value                  = ''
    address1_type                 = ''
    address1_formated             = ''
    address1_street               = ''
    address1_city                 = ''
    address1_pobox                = ''
    address1_region               = ''
    address1_postal_code          = ''
    address1_country              = ''
    address1_extended_address     = ''
    address2_type                 = ''
    address2_formated             = ''
    address2_street               = ''
    address2_city                 = ''
    address2_pobox                = ''
    address2_region               = ''
    address2_postal_code          = ''
    address2_country              = ''
    address2_extended_address     = ''
    address3_type                 = ''
    address3_formated             = ''
    address3_street               = ''
    address3_city                 = ''
    address3_pobox                = ''
    address3_region               = ''
    address3_postal_code          = ''
    address3_country              = ''
    address3_extended_address     = ''
    organization1_type            = ''
    organization1_name            = ''
    organization1_yomi_name       = ''
    organization1_title           = ''
    organization1_department      = ''
    organization1_symbol          = ''
    organization1_location        = ''
    organization1_job_description = ''
    relation1_type                = ''
    relation1_value               = ''
    external_id1_type             = ''
    external_id1_value            = ''
    website1_type                 = ''
    website1_value                = ''
    calendar_link1_type           = ''
    calendar_link1_value          = ''
    jot1_type                     = ''
    jot1_value                    = ''

    def __init__(self):
        name      = "-"
        notes     = ""

#}}}

# Create a new contact from sheet information
# new_contact_from_sheet(sheet, row)
#{{{
def new_contact_from_sheet(sheet, row):
    contact = Contact()

    if sheet[NAME + str(row)].value != None:
        contact.name = sheet[NAME + str(row)].value
    if sheet[GIVEN_NAME + str(row)].value != None:
        contact.given_name = sheet[GIVEN_NAME + str(row)].value
    if sheet[ADDITIONAL_NAME + str(row)].value != None:
        contact.additional_name = sheet[ADDITIONAL_NAME + str(row)].value
    if sheet[FAMILY_NAME + str(row)].value != None:
        contact.family_name = sheet[FAMILY_NAME + str(row)].value
    if sheet[YOMI_NAME + str(row)].value != None:
        contact.yomi_name = sheet[YOMI_NAME + str(row)].value
    if sheet[GIVEN_NAME_YOMI + str(row)].value != None:
        contact.given_name_yomi = sheet[GIVEN_NAME_YOMI + str(row)].value
    if sheet[ADDITIONAL_NAME_YOMI + str(row)].value != None:
        contact.additional_name_yomi = sheet[ADDITIONAL_NAME_YOMI + str(row)].value
    if sheet[FAMILY_NAME_YOMI + str(row)].value != None:
        contact.family_name_yomi = sheet[FAMILY_NAME_YOMI + str(row)].value
    if sheet[NAME_PREFIX + str(row)].value != None:
        contact.name_prefix = sheet[NAME_PREFIX + str(row)].value
    if sheet[NAME_SUFFIX + str(row)].value != None:
        contact.name_suffix = sheet[NAME_SUFFIX + str(row)].value
    if sheet[INITIALS + str(row)].value != None:
        contact.initials = sheet[INITIALS + str(row)].value
    if sheet[NICKNAME + str(row)].value != None:
        contact.nickname = sheet[NICKNAME + str(row)].value
    if sheet[SHORT_NAME + str(row)].value != None:
        contact.short_name = sheet[SHORT_NAME + str(row)].value
    if sheet[MAIDEN_NAME + str(row)].value != None:
        contact.maiden_name = sheet[MAIDEN_NAME + str(row)].value
    if sheet[BIRTHDAY + str(row)].value != None:
        contact.birthday = sheet[BIRTHDAY + str(row)].value
    if sheet[GENDER + str(row)].value != None:
        contact.gender = sheet[GENDER + str(row)].value
    if sheet[LOCATION + str(row)].value != None:
        contact.location = sheet[LOCATION + str(row)].value
    if sheet[BILLING_INFORMATION + str(row)].value != None:
        contact.billing_information = sheet[BILLING_INFORMATION + str(row)].value
    if sheet[DIRECTORY_SERVER + str(row)].value != None:
        contact.directory_server = sheet[DIRECTORY_SERVER + str(row)].value
    if sheet[MILEAGE + str(row)].value != None:
        contact.mileage = sheet[MILEAGE + str(row)].value
    if sheet[OCCUPATION + str(row)].value != None:
        contact.occupation = sheet[OCCUPATION + str(row)].value
    if sheet[HOBBY + str(row)].value != None:
        contact.hobby = sheet[HOBBY + str(row)].value
    if sheet[SENSITIVITY + str(row)].value != None:
        contact.sensitivity = sheet[SENSITIVITY + str(row)].value
    if sheet[PRIORITY + str(row)].value != None:
        contact.priority = sheet[PRIORITY + str(row)].value
    if sheet[SUBJECT + str(row)].value != None:
        contact.subject = sheet[SUBJECT + str(row)].value
    if sheet[NOTES + str(row)].value != None:
        contact.notes = sheet[NOTES + str(row)].value
    if sheet[GROUP_MEMBERSHIP + str(row)].value != None:
        contact.group_membership = sheet[GROUP_MEMBERSHIP + str(row)].value
    if sheet[EMAIL1_TYPE + str(row)].value != None:
        contact.email1_type = sheet[EMAIL1_TYPE + str(row)].value
    if sheet[EMAIL1_VALUE + str(row)].value != None:
        contact.email1_value = sheet[EMAIL1_VALUE + str(row)].value
    if sheet[EMAIL2_TYPE + str(row)].value != None:
        contact.email2_type = sheet[EMAIL2_TYPE + str(row)].value
    if sheet[EMAIL2_VALUE + str(row)].value != None:
        contact.email2_value = sheet[EMAIL2_VALUE + str(row)].value
    if sheet[EMAIL3_TYPE + str(row)].value != None:
        contact.email3_type = sheet[EMAIL3_TYPE + str(row)].value
    if sheet[EMAIL3_VALUE + str(row)].value != None:
        contact.email3_value = sheet[EMAIL3_VALUE + str(row)].value
    if sheet[EMAIL4_TYPE + str(row)].value != None:
        contact.email4_type = sheet[EMAIL4_TYPE + str(row)].value
    if sheet[EMAIL4_VALUE + str(row)].value != None:
        contact.email4_value = sheet[EMAIL4_VALUE + str(row)].value
    if sheet[EMAIL5_TYPE + str(row)].value != None:
        contact.email5_type = sheet[EMAIL5_TYPE + str(row)].value
    if sheet[EMAIL5_VALUE + str(row)].value != None:
        contact.email5_value = sheet[EMAIL5_VALUE + str(row)].value
    if sheet[IM1_TYPE + str(row)].value != None:
        contact.im1_type = sheet[IM1_TYPE + str(row)].value
    if sheet[IM1_SERVICE + str(row)].value != None:
        contact.im1_service = sheet[IM1_SERVICE + str(row)].value
    if sheet[IM1_VALUE + str(row)].value != None:
        contact.im1_value = sheet[IM1_VALUE + str(row)].value
    if sheet[PHONE1_TYPE + str(row)].value != None:
        contact.phone1_type = sheet[PHONE1_TYPE + str(row)].value
    if sheet[PHONE1_VALUE + str(row)].value != None:
        contact.phone1_value = sheet[PHONE1_VALUE + str(row)].value
    if sheet[PHONE2_TYPE + str(row)].value != None:
        contact.phone2_type = sheet[PHONE2_TYPE + str(row)].value
    if sheet[PHONE2_VALUE + str(row)].value != None:
        contact.phone2_value = sheet[PHONE2_VALUE + str(row)].value
    if sheet[PHONE3_TYPE + str(row)].value != None:
        contact.phone3_type = sheet[PHONE3_TYPE + str(row)].value
    if sheet[PHONE3_VALUE + str(row)].value != None:
        contact.phone3_value = sheet[PHONE3_VALUE + str(row)].value
    if sheet[PHONE4_TYPE + str(row)].value != None:
        contact.phone4_type = sheet[PHONE4_TYPE + str(row)].value
    if sheet[PHONE4_VALUE + str(row)].value != None:
        contact.phone4_value = sheet[PHONE4_VALUE + str(row)].value
    if sheet[PHONE5_TYPE + str(row)].value != None:
        contact.phone5_type = sheet[PHONE5_TYPE + str(row)].value
    if sheet[PHONE5_VALUE + str(row)].value != None:
        contact.phone5_value = sheet[PHONE5_VALUE + str(row)].value
    if sheet[ADDRESS1_TYPE + str(row)].value != None:
        contact.address1_type = sheet[ADDRESS1_TYPE + str(row)].value
    if sheet[ADDRESS1_FORMATED + str(row)].value != None:
        contact.address1_formated = sheet[ADDRESS1_FORMATED + str(row)].value
    if sheet[ADDRESS1_STREET + str(row)].value != None:
        contact.address1_street = sheet[ADDRESS1_STREET + str(row)].value
    if sheet[ADDRESS1_CITY + str(row)].value != None:
        contact.address1_city = sheet[ADDRESS1_CITY + str(row)].value
    if sheet[ADDRESS1_POBOX + str(row)].value != None:
        contact.address1_pobox = sheet[ADDRESS1_POBOX + str(row)].value
    if sheet[ADDRESS1_REGION + str(row)].value != None:
        contact.address1_region = sheet[ADDRESS1_REGION + str(row)].value
    if sheet[ADDRESS1_POSTAL_CODE + str(row)].value != None:
        contact.address1_postal_code = sheet[ADDRESS1_POSTAL_CODE + str(row)].value
    if sheet[ADDRESS1_COUNTRY + str(row)].value != None:
        contact.address1_country = sheet[ADDRESS1_COUNTRY + str(row)].value
    if sheet[ADDRESS1_EXTENDED_ADDRESS + str(row)].value != None:
        contact.address1_extended_address = sheet[ADDRESS1_EXTENDED_ADDRESS + str(row)].value
    if sheet[ADDRESS2_TYPE + str(row)].value != None:
        contact.address2_type = sheet[ADDRESS2_TYPE + str(row)].value
    if sheet[ADDRESS2_FORMATED + str(row)].value != None:
        contact.address2_formated = sheet[ADDRESS2_FORMATED + str(row)].value
    if sheet[ADDRESS2_STREET + str(row)].value != None:
        contact.address2_street = sheet[ADDRESS2_STREET + str(row)].value
    if sheet[ADDRESS2_CITY + str(row)].value != None:
        contact.address2_city = sheet[ADDRESS2_CITY + str(row)].value
    if sheet[ADDRESS2_POBOX + str(row)].value != None:
        contact.address2_pobox = sheet[ADDRESS2_POBOX + str(row)].value
    if sheet[ADDRESS2_REGION + str(row)].value != None:
        contact.address2_region = sheet[ADDRESS2_REGION + str(row)].value
    if sheet[ADDRESS2_POSTAL_CODE + str(row)].value != None:
        contact.address2_postal_code = sheet[ADDRESS2_POSTAL_CODE + str(row)].value
    if sheet[ADDRESS2_COUNTRY + str(row)].value != None:
        contact.address2_country = sheet[ADDRESS2_COUNTRY + str(row)].value
    if sheet[ADDRESS2_EXTENDED_ADDRESS + str(row)].value != None:
        contact.address2_extended_address = sheet[ADDRESS2_EXTENDED_ADDRESS + str(row)].value
    if sheet[ADDRESS3_TYPE + str(row)].value != None:
        contact.address3_type = sheet[ADDRESS3_TYPE + str(row)].value
    if sheet[ADDRESS3_FORMATED + str(row)].value != None:
        contact.address3_formated = sheet[ADDRESS3_FORMATED + str(row)].value
    if sheet[ADDRESS3_STREET + str(row)].value != None:
        contact.address3_street = sheet[ADDRESS3_STREET + str(row)].value
    if sheet[ADDRESS3_CITY + str(row)].value != None:
        contact.address3_city = sheet[ADDRESS3_CITY + str(row)].value
    if sheet[ADDRESS3_POBOX + str(row)].value != None:
        contact.address3_pobox = sheet[ADDRESS3_POBOX + str(row)].value
    if sheet[ADDRESS3_REGION + str(row)].value != None:
        contact.address3_region = sheet[ADDRESS3_REGION + str(row)].value
    if sheet[ADDRESS3_POSTAL_CODE + str(row)].value != None:
        contact.address3_postal_code = sheet[ADDRESS3_POSTAL_CODE + str(row)].value
    if sheet[ADDRESS3_COUNTRY + str(row)].value != None:
        contact.address3_country = sheet[ADDRESS3_COUNTRY + str(row)].value
    if sheet[ADDRESS3_EXTENDED_ADDRESS + str(row)].value != None:
        contact.address3_extended_address = sheet[ADDRESS3_EXTENDED_ADDRESS + str(row)].value
    if sheet[ORGANIZATION1_TYPE + str(row)].value != None:
        contact.organization1_type = sheet[ORGANIZATION1_TYPE + str(row)].value
    if sheet[ORGANIZATION1_NAME + str(row)].value != None:
        contact.organization1_name = sheet[ORGANIZATION1_NAME + str(row)].value
    if sheet[ORGANIZATION1_YOMI_NAME + str(row)].value != None:
        contact.organization1_yomi_name = sheet[ORGANIZATION1_YOMI_NAME + str(row)].value
    if sheet[ORGANIZATION1_TITLE + str(row)].value != None:
        contact.organization1_title = sheet[ORGANIZATION1_TITLE + str(row)].value
    if sheet[ORGANIZATION1_DEPARTMENT + str(row)].value != None:
        contact.organization1_department = sheet[ORGANIZATION1_DEPARTMENT + str(row)].value
    if sheet[ORGANIZATION1_SYMBOL + str(row)].value != None:
        contact.organization1_symbol = sheet[ORGANIZATION1_SYMBOL + str(row)].value
    if sheet[ORGANIZATION1_LOCATION + str(row)].value != None:
        contact.organization1_location = sheet[ORGANIZATION1_LOCATION + str(row)].value
    if sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value != None:
        contact.organization1_job_description = sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value
    if sheet[RELATION1_TYPE + str(row)].value != None:
        contact.relation1_type = sheet[RELATION1_TYPE + str(row)].value
    if sheet[RELATION1_VALUE + str(row)].value != None:
        contact.relation1_value = sheet[RELATION1_VALUE + str(row)].value
    if sheet[EXTERNAL_ID1_TYPE + str(row)].value != None:
        contact.external_id1_type = sheet[EXTERNAL_ID1_TYPE + str(row)].value
    if sheet[EXTERNAl_ID1_VALUE + str(row)].value != None:
        contact.external_id1_value = sheet[EXTERNAl_ID1_VALUE + str(row)].value
    if sheet[WEBSITE1_TYPE + str(row)].value != None:
        contact.website1_type = sheet[WEBSITE1_TYPE + str(row)].value
    if sheet[WEBSITE1_VALUE + str(row)].value != None:
        contact.website1_value = sheet[WEBSITE1_VALUE + str(row)].value
    if sheet[CALENDAR_LINK1_TYPE + str(row)].value != None:
        contact.calendar_link1_type = sheet[CALENDAR_LINK1_TYPE + str(row)].value
    if sheet[CALENDAR_LINK1_VALUE + str(row)].value != None:
        contact.calendar_link1_value = sheet[CALENDAR_LINK1_VALUE + str(row)].value
    if sheet[JOT1_TYPE + str(row)].value != None:
        contact.jot1_type = sheet[JOT1_TYPE + str(row)].value
    if sheet[JOT1_VALUE + str(row)].value != None:
        contact.jot1_value = sheet[JOT1_VALUE + str(row)].value

    return contact
#}}}

# Update Contact from sheet information
# update_contact_from_sheet(sheet, row, contact)
#{{{
def update_contact_from_sheet(sheet, row, contact):
    # update these fields by adding a slash
    # INCLUDE CHANGE IN COMPANY AND CHANGE IN POSITION IN THE NOTES
    contact.notes       = special_combine(contact.notes,       sheet[NOTES       + str(row)].value)
    # update all 
    if sheet[NAME + str(row)].value != None or contact.name != None:
        info = keep_non_none(contact.name, sheet[NAME + str(row)].value)
        contact.name = info[0]
        contact.notes = add_to_notes(contact, "Name", info[1], info[0])
    if sheet[GIVEN_NAME + str(row)].value != None or contact.given_name != None:
        info = keep_non_none(contact.given_name, sheet[GIVEN_NAME + str(row)].value)
        contact.NAME = info[0]
        contact.notes = add_to_notes(contact, "Given Name", info[1], info[0])
    if sheet[ADDITIONAL_NAME + str(row)].value != None or contact.ADDITIONAL_NAME != None:
        info = keep_non_none(contact.given_name, sheet[GIVEN_NAME + str(row)].value)
        contact.NAME = info[0]
        contact.notes = add_to_notes(contact, "Additional Name", info[1], info[0])
    if sheet[FAMILY_NAME + str(row)].value != None or contact.family_name != None:
        info = keep_non_none(contact.family_name, sheet[FAMILY_NAME + str(row)].value)
        contact.GIVEN_NAME = info[0]
        contact.notes = add_to_notes(contact, "Family Name", info[1], info[0])
    if sheet[YOMI_NAME + str(row)].value != None or contact.YOMI_NAME != None:
        info = keep_non_none(contact.family_name, sheet[FAMILY_NAME + str(row)].value)
        contact.GIVEN_NAME = info[0]
        contact.notes = add_to_notes(contact, "Yomi Name", info[1], info[0])
    if sheet[GIVEN_NAME_YOMI + str(row)].value != None or contact.GIVEN_NAME_YOMI  != None:
        info = keep_non_none(contact.given_name_yomi, sheet[GIVEN_NAME_YOMI + str(row)].value)
        contact.GIVEN_NAME_YOMI = info[0]
        contact.notes = add_to_notes(contact, "GIVEN_NAME_YOMI", info[1], info[0])
    if sheet[ADDITIONAL_NAME_YOMI + str(row)].value != None or contact.ADDITIONAL_NAME_YOMI  != None:
        info = keep_non_none(contact.additional_name_yomi, sheet[ADDITIONAL_NAME_YOMI + str(row)].value)
        contact.ADDITIONAL_NAME_YOMI = info[0]
        contact.notes = add_to_notes(contact, "ADDITIONAL_NAME_YOMI", info[1], info[0])
    if sheet[FAMILY_NAME_YOMI + str(row)].value != None or contact.FAMILY_NAME_YOMI  != None:
        info = keep_non_none(contact.family_name_yomi, sheet[FAMILY_NAME_YOMI + str(row)].value)
        contact.FAMILY_NAME_YOMI = info[0]
        contact.notes = add_to_notes(contact, "FAMILY_NAME_YOMI", info[1], info[0])
    if sheet[NAME_PREFIX + str(row)].value != None or contact.NAME_PREFIX  != None:
        info = keep_non_none(contact.name_prefix, sheet[NAME_PREFIX + str(row)].value)
        contact.NAME_PREFIX = info[0]
        contact.notes = add_to_notes(contact, "NAME_PREFIX", info[1], info[0])
    if sheet[NAME_SUFFIX + str(row)].value != None or contact.NAME_SUFFIX  != None:
        info = keep_non_none(contact.name_suffix, sheet[NAME_SUFFIX + str(row)].value)
        contact.NAME_SUFFIX = info[0]
        contact.notes = add_to_notes(contact, "NAME_SUFFIX", info[1], info[0])
    if sheet[INITIALS + str(row)].value != None or contact.INITIALS  != None:
        info = keep_non_none(contact.initials, sheet[INITIALS + str(row)].value)
        contact.INITIALS = info[0]
        contact.notes = add_to_notes(contact, "INITIALS", info[1], info[0])
    if sheet[NICKNAME + str(row)].value != None or contact.NICKNAME  != None:
        info = keep_non_none(contact.nickname, sheet[NICKNAME + str(row)].value)
        contact.NICKNAME = info[0]
        contact.notes = add_to_notes(contact, "NICKNAME", info[1], info[0])
    if sheet[SHORT_NAME + str(row)].value != None or contact.SHORT_NAME  != None:
        info = keep_non_none(contact.short_name, sheet[SHORT_NAME + str(row)].value)
        contact.SHORT_NAME = info[0]
        contact.notes = add_to_notes(contact, "SHORT_NAME", info[1], info[0])
    if sheet[MAIDEN_NAME + str(row)].value != None or contact.MAIDEN_NAME  != None:
        info = keep_non_none(contact.maiden_name, sheet[MAIDEN_NAME + str(row)].value)
        contact.MAIDEN_NAME = info[0]
        contact.notes = add_to_notes(contact, "MAIDEN_NAME", info[1], info[0])
    if sheet[BIRTHDAY + str(row)].value != None or contact.BIRTHDAY  != None:
        info = keep_non_none(contact.birthday, sheet[BIRTHDAY + str(row)].value)
        contact.BIRTHDAY = info[0]
        contact.notes = add_to_notes(contact, "BIRTHDAY", info[1], info[0])
    if sheet[GENDER + str(row)].value != None or contact.GENDER  != None:
        info = keep_non_none(contact.gender, sheet[GENDER + str(row)].value)
        contact.GENDER = info[0]
        contact.notes = add_to_notes(contact, "GENDER", info[1], info[0])
    if sheet[LOCATION + str(row)].value != None or contact.LOCATION  != None:
        info = keep_non_none(contact.location, sheet[LOCATION + str(row)].value)
        contact.LOCATION = info[0]
        contact.notes = add_to_notes(contact, "LOCATION", info[1], info[0])
    if sheet[BILLING_INFORMATION + str(row)].value != None or contact.BILLING_INFORMATION  != None:
        info = keep_non_none(contact.billing_information, sheet[BILLING_INFORMATION + str(row)].value)
        contact.BILLING_INFORMATION = info[0]
        contact.notes = add_to_notes(contact, "BILLING_INFORMATION", info[1], info[0])
    if sheet[DIRECTORY_SERVER + str(row)].value != None or contact.DIRECTORY_SERVER  != None:
        info = keep_non_none(contact.directory_server, sheet[DIRECTORY_SERVER + str(row)].value)
        contact.DIRECTORY_SERVER = info[0]
        contact.notes = add_to_notes(contact, "DIRECTORY_SERVER", info[1], info[0])
    if sheet[MILEAGE + str(row)].value != None or contact.MILEAGE  != None:
        info = keep_non_none(contact.mileage, sheet[MILEAGE + str(row)].value)
        contact.MILEAGE = info[0]
        contact.notes = add_to_notes(contact, "MILEAGE", info[1], info[0])
    if sheet[OCCUPATION + str(row)].value != None or contact.OCCUPATION  != None:
        info = keep_non_none(contact.occupation, sheet[OCCUPATION + str(row)].value)
        contact.OCCUPATION = info[0]
        contact.notes = add_to_notes(contact, "OCCUPATION", info[1], info[0])
    if sheet[HOBBY + str(row)].value != None or contact.HOBBY  != None:
        info = keep_non_none(contact.hobby, sheet[HOBBY + str(row)].value)
        contact.HOBBY = info[0]
        contact.notes = add_to_notes(contact, "HOBBY", info[1], info[0])
    if sheet[SENSITIVITY + str(row)].value != None or contact.SENSITIVITY  != None:
        info = keep_non_none(contact.sensitivity, sheet[SENSITIVITY + str(row)].value)
        contact.SENSITIVITY = info[0]
        contact.notes = add_to_notes(contact, "SENSITIVITY", info[1], info[0])
    if sheet[PRIORITY + str(row)].value != None or contact.PRIORITY  != None:
        info = keep_non_none(contact.priority, sheet[PRIORITY + str(row)].value)
        contact.PRIORITY = info[0]
        contact.notes = add_to_notes(contact, "PRIORITY", info[1], info[0])
    if sheet[SUBJECT + str(row)].value != None or contact.SUBJECT  != None:
        info = keep_non_none(contact.subject, sheet[SUBJECT + str(row)].value)
        contact.SUBJECT = info[0]
        contact.notes = add_to_notes(contact, "SUBJECT", info[1], info[0])
    if sheet[GROUP_MEMBERSHIP + str(row)].value != None or contact.GROUP_MEMBERSHIP  != None:
        info = keep_non_none(contact.group_membership, sheet[GROUP_MEMBERSHIP + str(row)].value)
        contact.GROUP_MEMBERSHIP = info[0]
        contact.notes = add_to_notes(contact, "GROUP_MEMBERSHIP", info[1], info[0])
    if sheet[EMAIL1_TYPE + str(row)].value != None or contact.EMAIL1_TYPE  != None:
        info = keep_non_none(contact.email1_type, sheet[EMAIL1_TYPE + str(row)].value)
        contact.EMAIL1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL1_TYPE", info[1], info[0])
    if sheet[EMAIL1_VALUE + str(row)].value != None or contact.EMAIL1_VALUE  != None:
        info = keep_non_none(contact.email1_value, sheet[EMAIL1_VALUE + str(row)].value)
        contact.EMAIL1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL1_VALUE", info[1], info[0])
    if sheet[EMAIL2_TYPE + str(row)].value != None or contact.EMAIL2_TYPE  != None:
        info = keep_non_none(contact.email2_type, sheet[EMAIL2_TYPE + str(row)].value)
        contact.EMAIL2_TYPE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL2_TYPE", info[1], info[0])
    if sheet[EMAIL2_VALUE + str(row)].value != None or contact.EMAIL2_VALUE  != None:
        info = keep_non_none(contact.email2_value, sheet[EMAIL2_VALUE + str(row)].value)
        contact.EMAIL2_VALUE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL2_VALUE", info[1], info[0])
    if sheet[EMAIL3_TYPE + str(row)].value != None or contact.EMAIL3_TYPE  != None:
        info = keep_non_none(contact.email3_type, sheet[EMAIL3_TYPE + str(row)].value)
        contact.EMAIL3_TYPE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL3_TYPE", info[1], info[0])
    if sheet[EMAIL3_VALUE + str(row)].value != None or contact.EMAIL3_VALUE  != None:
        info = keep_non_none(contact.email3_value, sheet[EMAIL3_VALUE + str(row)].value)
        contact.EMAIL3_VALUE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL3_VALUE", info[1], info[0])
    if sheet[EMAIL4_TYPE + str(row)].value != None or contact.EMAIL4_TYPE  != None:
        info = keep_non_none(contact.email4_type, sheet[EMAIL4_TYPE + str(row)].value)
        contact.EMAIL4_TYPE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL4_TYPE", info[1], info[0])
    if sheet[EMAIL4_VALUE + str(row)].value != None or contact.EMAIL4_VALUE  != None:
        info = keep_non_none(contact.email4_value, sheet[EMAIL4_VALUE + str(row)].value)
        contact.EMAIL4_VALUE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL4_VALUE", info[1], info[0])
    if sheet[EMAIL5_TYPE + str(row)].value != None or contact.EMAIL5_TYPE  != None:
        info = keep_non_none(contact.email5_type, sheet[EMAIL5_TYPE + str(row)].value)
        contact.EMAIL5_TYPE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL5_TYPE", info[1], info[0])
    if sheet[EMAIL5_VALUE + str(row)].value != None or contact.EMAIL5_VALUE  != None:
        info = keep_non_none(contact.email5_value, sheet[EMAIL5_VALUE + str(row)].value)
        contact.EMAIL5_VALUE = info[0]
        contact.notes = add_to_notes(contact, "EMAIL5_VALUE", info[1], info[0])
    if sheet[IM1_TYPE + str(row)].value != None or contact.IM1_TYPE  != None:
        info = keep_non_none(contact.im1_type, sheet[IM1_TYPE + str(row)].value)
        contact.IM1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "IM1_TYPE", info[1], info[0])
    if sheet[IM1_SERVICE + str(row)].value != None or contact.IM1_SERVICE  != None:
        info = keep_non_none(contact.im1_service, sheet[IM1_SERVICE + str(row)].value)
        contact.IM1_SERVICE = info[0]
        contact.notes = add_to_notes(contact, "IM1_SERVICE", info[1], info[0])
    if sheet[IM1_VALUE + str(row)].value != None or contact.IM1_VALUE  != None:
        info = keep_non_none(contact.im1_value, sheet[IM1_VALUE + str(row)].value)
        contact.IM1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "IM1_VALUE", info[1], info[0])
    if sheet[PHONE1_TYPE + str(row)].value != None or contact.PHONE1_TYPE  != None:
        info = keep_non_none(contact.phone1_type, sheet[PHONE1_TYPE + str(row)].value)
        contact.PHONE1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "PHONE1_TYPE", info[1], info[0])
    if sheet[PHONE1_VALUE + str(row)].value != None or contact.PHONE1_VALUE  != None:
        info = keep_non_none(contact.phone1_value, sheet[PHONE1_VALUE + str(row)].value)
        contact.PHONE1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "PHONE1_VALUE", info[1], info[0])
    if sheet[PHONE2_TYPE + str(row)].value != None or contact.PHONE2_TYPE  != None:
        info = keep_non_none(contact.phone2_type, sheet[PHONE2_TYPE + str(row)].value)
        contact.PHONE2_TYPE = info[0]
        contact.notes = add_to_notes(contact, "PHONE2_TYPE", info[1], info[0])
    if sheet[PHONE2_VALUE + str(row)].value != None or contact.PHONE2_VALUE  != None:
        info = keep_non_none(contact.phone2_value, sheet[PHONE2_VALUE + str(row)].value)
        contact.PHONE2_VALUE = info[0]
        contact.notes = add_to_notes(contact, "PHONE2_VALUE", info[1], info[0])
    if sheet[PHONE3_TYPE + str(row)].value != None or contact.PHONE3_TYPE  != None:
        info = keep_non_none(contact.phone3_type, sheet[PHONE3_TYPE + str(row)].value)
        contact.PHONE3_TYPE = info[0]
        contact.notes = add_to_notes(contact, "PHONE3_TYPE", info[1], info[0])
    if sheet[PHONE3_VALUE + str(row)].value != None or contact.PHONE3_VALUE  != None:
        info = keep_non_none(contact.phone3_value, sheet[PHONE3_VALUE + str(row)].value)
        contact.PHONE3_VALUE = info[0]
        contact.notes = add_to_notes(contact, "PHONE3_VALUE", info[1], info[0])
    if sheet[PHONE4_TYPE + str(row)].value != None or contact.PHONE4_TYPE  != None:
        info = keep_non_none(contact.phone4_type, sheet[PHONE4_TYPE + str(row)].value)
        contact.PHONE4_TYPE = info[0]
        contact.notes = add_to_notes(contact, "PHONE4_TYPE", info[1], info[0])
    if sheet[PHONE4_VALUE + str(row)].value != None or contact.PHONE4_VALUE  != None:
        info = keep_non_none(contact.phone4_value, sheet[PHONE4_VALUE + str(row)].value)
        contact.PHONE4_VALUE = info[0]
        contact.notes = add_to_notes(contact, "PHONE4_VALUE", info[1], info[0])
    if sheet[PHONE5_TYPE + str(row)].value != None or contact.PHONE5_TYPE  != None:
        info = keep_non_none(contact.phone5_type, sheet[PHONE5_TYPE + str(row)].value)
        contact.PHONE5_TYPE = info[0]
        contact.notes = add_to_notes(contact, "PHONE5_TYPE", info[1], info[0])
    if sheet[PHONE5_VALUE + str(row)].value != None or contact.PHONE5_VALUE  != None:
        info = keep_non_none(contact.phone5_value, sheet[PHONE5_VALUE + str(row)].value)
        contact.PHONE5_VALUE = info[0]
        contact.notes = add_to_notes(contact, "PHONE5_VALUE", info[1], info[0])
    if sheet[ADDRESS1_TYPE + str(row)].value != None or contact.ADDRESS1_TYPE  != None:
        info = keep_non_none(contact.address1_type, sheet[ADDRESS1_TYPE + str(row)].value)
        contact.ADDRESS1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_TYPE", info[1], info[0])
    if sheet[ADDRESS1_FORMATED + str(row)].value != None or contact.ADDRESS1_FORMATED  != None:
        info = keep_non_none(contact.address1_formated, sheet[ADDRESS1_FORMATED + str(row)].value)
        contact.ADDRESS1_FORMATED = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_FORMATED", info[1], info[0])
    if sheet[ADDRESS1_STREET + str(row)].value != None or contact.ADDRESS1_STREET  != None:
        info = keep_non_none(contact.address1_street, sheet[ADDRESS1_STREET + str(row)].value)
        contact.ADDRESS1_STREET = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_STREET", info[1], info[0])
    if sheet[ADDRESS1_CITY + str(row)].value != None or contact.ADDRESS1_CITY  != None:
        info = keep_non_none(contact.address1_city, sheet[ADDRESS1_CITY + str(row)].value)
        contact.ADDRESS1_CITY = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_CITY", info[1], info[0])
    if sheet[ADDRESS1_POBOX + str(row)].value != None or contact.ADDRESS1_POBOX  != None:
        info = keep_non_none(contact.address1_pobox, sheet[ADDRESS1_POBOX + str(row)].value)
        contact.ADDRESS1_POBOX = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_POBOX", info[1], info[0])
    if sheet[ADDRESS1_REGION + str(row)].value != None or contact.ADDRESS1_REGION  != None:
        info = keep_non_none(contact.address1_region, sheet[ADDRESS1_REGION + str(row)].value)
        contact.ADDRESS1_REGION = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_REGION", info[1], info[0])
    if sheet[ADDRESS1_POSTAL_CODE + str(row)].value != None or contact.ADDRESS1_POSTAL_CODE  != None:
        info = keep_non_none(contact.address1_postal_code, sheet[ADDRESS1_POSTAL_CODE + str(row)].value)
        contact.ADDRESS1_POSTAL_CODE = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_POSTAL_CODE", info[1], info[0])
    if sheet[ADDRESS1_COUNTRY + str(row)].value != None or contact.ADDRESS1_COUNTRY  != None:
        info = keep_non_none(contact.address1_country, sheet[ADDRESS1_COUNTRY + str(row)].value)
        contact.ADDRESS1_COUNTRY = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_COUNTRY", info[1], info[0])
    if sheet[ADDRESS1_EXTENDED_ADDRESS + str(row)].value != None or contact.ADDRESS1_EXTENDED_ADDRESS  != None:
        info = keep_non_none(contact.address1_extended_address, sheet[address1_extended_address + str(row)].value)
        contact.address1_extended_address = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_EXTENDED_ADDRESS", info[1], info[0])
    if sheet[ADDRESS2_TYPE + str(row)].value != None or contact.ADDRESS2_TYPE  != None:
        info = keep_non_none(contact.address2_type, sheet[ADDRESS2_TYPE + str(row)].value)
        contact.address2_type = info[0]
        contact.notes = add_to_notes(contact, "Address2 Type", info[1], info[0])
    if sheet[ADDRESS2_FORMATED + str(row)].value != None or contact.ADDRESS2_FORMATED  != None:
        info = keep_non_none(contact.address2_formated, sheet[ADDRESS2_FORMATED + str(row)].value)
        contact.ADDRESS2_FORMATED = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_FORMATED", info[1], info[0])
    if sheet[ADDRESS2_STREET + str(row)].value != None or contact.ADDRESS2_STREET  != None:
        info = keep_non_none(contact.address2_street, sheet[ADDRESS2_STREET + str(row)].value)
        contact.ADDRESS2_STREET = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_STREET", info[1], info[0])
    if sheet[ADDRESS2_CITY + str(row)].value != None or contact.ADDRESS2_CITY  != None:
        info = keep_non_none(contact.address2_city, sheet[ADDRESS2_CITY + str(row)].value)
        contact.ADDRESS2_CITY = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_CITY", info[1], info[0])
    if sheet[ADDRESS2_POBOX + str(row)].value != None or contact.ADDRESS2_POBOX  != None:
        info = keep_non_none(contact.address2_pobox, sheet[ADDRESS2_POBOX + str(row)].value)
        contact.ADDRESS2_POBOX = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_POBOX", info[1], info[0])
    if sheet[ADDRESS2_REGION + str(row)].value != None or contact.ADDRESS2_REGION  != None:
        info = keep_non_none(contact.address2_region, sheet[ADDRESS2_REGION + str(row)].value)
        contact.ADDRESS2_REGION = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_REGION", info[1], info[0])
    if sheet[ADDRESS2_POSTAL_CODE + str(row)].value != None or contact.ADDRESS2_POSTAL_CODE  != None:
        info = keep_non_none(contact.address2_postal_code, sheet[ADDRESS2_POSTAL_CODE + str(row)].value)
        contact.ADDRESS2_POSTAL_CODE = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_POSTAL_CODE", info[1], info[0])
    if sheet[ADDRESS2_COUNTRY + str(row)].value != None or contact.ADDRESS2_COUNTRY  != None:
        info = keep_non_none(contact.address2_country, sheet[ADDRESS2_COUNTRY + str(row)].value)
        contact.ADDRESS2_COUNTRY = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_COUNTRY", info[1], info[0])
    if sheet[ADDRESS2_EXTENDED_ADDRESS + str(row)].value != None or contact.ADDRESS2_EXTENDED_ADDRESS  != None:
        info = keep_non_none(contact.address2_extended_address , sheet[ADDRESS2_EXTENDED_ADDRESS + str(row)].value)
        contact.ADDRESS2_EXTENDED_ADDRESS = info[0]
        contact.notes = add_to_notes(contact, "Address 2 Extended Address", info[1], info[0])
    if sheet[ADDRESS3_TYPE + str(row)].value != None or contact.ADDRESS3_TYPE  != None:
        info = keep_non_none(contact.address3_type, sheet[ADDRESS3_TYPE + str(row)].value)
        contact.ADDRESS3_TYPE = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_TYPE", info[1], info[0])
    if sheet[ADDRESS3_FORMATED + str(row)].value != None or contact.ADDRESS3_FORMATED  != None:
        info = keep_non_none(contact.address3_formated, sheet[ADDRESS3_FORMATED + str(row)].value)
        contact.ADDRESS3_FORMATED = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_FORMATED", info[1], info[0])
    if sheet[ADDRESS3_STREET + str(row)].value != None or contact.ADDRESS3_STREET  != None:
        info = keep_non_none(contact.address3_street, sheet[ADDRESS3_STREET + str(row)].value)
        contact.ADDRESS3_STREET = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_STREET", info[1], info[0])
    if sheet[ADDRESS3_CITY + str(row)].value != None or contact.ADDRESS3_CITY  != None:
        info = keep_non_none(contact.address3_city, sheet[ADDRESS3_CITY + str(row)].value)
        contact.ADDRESS3_CITY = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_CITY", info[1], info[0])
    if sheet[ADDRESS3_POBOX + str(row)].value != None or contact.ADDRESS3_POBOX  != None:
        info = keep_non_none(contact.address3_pobox, sheet[ADDRESS3_POBOX + str(row)].value)
        contact.ADDRESS3_POBOX = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_POBOX", info[1], info[0])
    if sheet[ADDRESS3_REGION + str(row)].value != None or contact.ADDRESS3_REGION  != None:
        info = keep_non_none(contact.address3_region, sheet[ADDRESS3_REGION + str(row)].value)
        contact.ADDRESS3_REGION = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_REGION", info[1], info[0])
    if sheet[ADDRESS3_POSTAL_CODE + str(row)].value != None or contact.ADDRESS3_POSTAL_CODE  != None:
        info = keep_non_none(contact.address3_postal_code, sheet[ADDRESS3_POSTAL_CODE + str(row)].value)
        contact.ADDRESS3_POSTAL_CODE = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_POSTAL_CODE", info[1], info[0])
    if sheet[ADDRESS3_COUNTRY + str(row)].value != None or contact.ADDRESS3_COUNTRY  != None:
        info = keep_non_none(contact.address3_country, sheet[ADDRESS3_COUNTRY + str(row)].value)
        contact.ADDRESS3_COUNTRY = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_COUNTRY", info[1], info[0])
    if sheet[ADDRESS3_EXTENDED_ADDRESS + str(row)].value != None or contact.ADDRESS3_EXTENDED_ADDRESS  != None:
        info = keep_non_none(contact.address3_extended_address, sheet[ADDRESS3_EXTENDED_ADDRESS + str(row)].value)
        contact.ADDRESS3_EXTENDED_ADDRESS = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_EXTENDED_ADDRESS", info[1], info[0])
    if sheet[ORGANIZATION1_TYPE + str(row)].value != None or contact.ORGANIZATION1_TYPE  != None:
        info = keep_non_none(contact.organization1_type, sheet[ORGANIZATION1_TYPE + str(row)].value)
        contact.ORGANIZATION1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_TYPE", info[1], info[0])
    if sheet[ORGANIZATION1_NAME + str(row)].value != None or contact.ORGANIZATION1_NAME  != None:
        info = keep_non_none(contact.organization1_name, sheet[ORGANIZATION1_NAME + str(row)].value)
        contact.ORGANIZATION1_NAME = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_NAME", info[1], info[0])
    if sheet[ORGANIZATION1_YOMI_NAME + str(row)].value != None or contact.ORGANIZATION1_YOMI_NAME  != None:
        info = keep_non_none(contact.organization1_yomi_name, sheet[ORGANIZATION1_YOMI_NAME + str(row)].value)
        contact.ORGANIZATION1_YOMI_NAME = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_YOMI_NAME", info[1], info[0])
    if sheet[ORGANIZATION1_TITLE + str(row)].value != None or contact.ORGANIZATION1_TITLE  != None:
        info = keep_non_none(contact.organization1_title, sheet[ORGANIZATION1_TITLE + str(row)].value)
        contact.ORGANIZATION1_TITLE = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_TITLE", info[1], info[0])
    if sheet[ORGANIZATION1_DEPARTMENT + str(row)].value != None or contact.ORGANIZATION1_DEPARTMENT  != None:
        info = keep_non_none(contact.organization1_department, sheet[ORGANIZATION1_DEPARTMENT + str(row)].value)
        contact.ORGANIZATION1_DEPARTMENT = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_DEPARTMENT", info[1], info[0])
    if sheet[ORGANIZATION1_SYMBOL + str(row)].value != None or contact.ORGANIZATION1_SYMBOL  != None:
        info = keep_non_none(contact.organization1_symbol, sheet[ORGANIZATION1_SYMBOL + str(row)].value)
        contact.ORGANIZATION1_SYMBOL = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_SYMBOL", info[1], info[0])
    if sheet[ORGANIZATION1_LOCATION + str(row)].value != None or contact.ORGANIZATION1_LOCATION  != None:
        info = keep_non_none(contact.organization1_location, sheet[ORGANIZATION1_LOCATION + str(row)].value)
        contact.ORGANIZATION1_LOCATION = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_LOCATION", info[1], info[0])
    if sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value != None or contact.ORGANIZATION1_JOB_DESCRIPTION  != None:
        info = keep_non_none(contact.organization1_job_description, sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value)
        contact.ORGANIZATION1_JOB_DESCRIPTION = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_JOB_DESCRIPTION", info[1], info[0])
    if sheet[RELATION1_TYPE + str(row)].value != None or contact.RELATION1_TYPE  != None:
        info = keep_non_none(contact.relation1_type, sheet[RELATION1_TYPE + str(row)].value)
        contact.RELATION1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "RELATION1_TYPE", info[1], info[0])
    if sheet[RELATION1_VALUE + str(row)].value != None or contact.RELATION1_VALUE  != None:
        info = keep_non_none(contact.relation1_value, sheet[RELATION1_VALUE + str(row)].value)
        contact.RELATION1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "RELATION1_VALUE", info[1], info[0])
    if sheet[EXTERNAL_ID1_TYPE + str(row)].value != None or contact.EXTERNAL_ID1_TYPE  != None:
        info = keep_non_none(contact.external_id1_type, sheet[EXTERNAL_ID1_TYPE + str(row)].value)
        contact.EXTERNAL_ID1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "EXTERNAL_ID1_TYPE", info[1], info[0])
    if sheet[EXTERNAl_ID1_VALUE + str(row)].value != None or contact.EXTERNAl_ID1_VALUE  != None:
        info = keep_non_none(contact.external_id1_value, sheet[EXTERNAl_ID1_VALUE + str(row)].value)
        contact.EXTERNAl_ID1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "EXTERNAl_ID1_VALUE", info[1], info[0])
    if sheet[WEBSITE1_TYPE + str(row)].value != None or contact.WEBSITE1_TYPE  != None:
        info = keep_non_none(contact.website1_type, sheet[WEBSITE1_TYPE + str(row)].value)
        contact.WEBSITE1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "WEBSITE1_TYPE", info[1], info[0])
    if sheet[WEBSITE1_VALUE + str(row)].value != None or contact.WEBSITE1_VALUE  != None:
        info = keep_non_none(contact.website1_value, sheet[WEBSITE1_VALUE + str(row)].value)
        contact.WEBSITE1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "WEBSITE1_VALUE", info[1], info[0])
    if sheet[CALENDAR_LINK1_TYPE + str(row)].value != None or contact.CALENDAR_LINK1_TYPE  != None:
        info = keep_non_none(contact.calendar_link1_type, sheet[CALENDAR_LINK1_TYPE + str(row)].value)
        contact.CALENDAR_LINK1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "CALENDAR_LINK1_TYPE", info[1], info[0])
    if sheet[CALENDAR_LINK1_VALUE + str(row)].value != None or contact.CALENDAR_LINK1_VALUE  != None:
        info = keep_non_none(contact.calendar_link1_value, sheet[CALENDAR_LINK1_VALUE + str(row)].value)
        contact.CALENDAR_LINK1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "CALENDAR_LINK1_VALUE", info[1], info[0])
    if sheet[JOT1_TYPE + str(row)].value != None or contact.JOT1_TYPE  != None:
        info = keep_non_none(contact.jot1_type, sheet[JOT1_TYPE + str(row)].value)
        contact.JOT1_TYPE = info[0]
        contact.notes = add_to_notes(contact, "JOT1_TYPE", info[1], info[0])
    if sheet[JOT1_VALUE + str(row)].value != None or contact.JOT1_VALUE  != None:
        info = keep_non_none(contact.jot1_value, sheet[JOT1_VALUE + str(row)].value)
        contact.JOT1_VALUE = info[0]
        contact.notes = add_to_notes(contact, "JOT1_VALUE", info[1], info[0])

    return contact
#}}}

################
# MAIN METHODS #
################
# Change 'US' and 'United States of America' to 'United States'
# standardize_USA(fileName, start, column)
#{{{
# def standardize_USA(fileName, start, column):
def standardize_USA(*args):
    # turn the arguments into variable names
    args = args[0]
    fileName = args[1]
    start = int(args[2])
    cols = args[3:]
    # Open an existing excel file
    if printing:
        print("Opening...")
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    #################
    # DO STUFF HERE #
    #################
    for col in cols:
        for row in range (start, sheet.max_row + 1):
            country = str(sheet[col+ str(row)].value)
            regexUSA = '(U\.?S\.?A?\.?|United ?States ?(of ?America)?)'
            matchUSA = re.search(regexUSA, country, re.IGNORECASE)
            if matchUSA:
                sheet[col+ str(row)].value = "United States"
            regexUK = '(U\.?K\.?)'
            matchUK = re.search(regexUK, country, re.IGNORECASE)
            if matchUK:
                sheet[col+ str(row)].value = "United Kingdom"


    if printing:
        print("Saving...")

    wb.save("betterFile.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

    if printing:
        print()
        print("Done!")
#}}}

# Combine LinkedIn and Google Contacts
# Copies all contacts from first list to second
# combine_contacts(fileName1, sheetName1, fileName2, sheetName2)
#{{{
def combine_contacts(fileName1, sheetName1, fileName2, sheetName2):
    if printing:
        print("Opening...")
    wb1 = openpyxl.load_workbook(fileName1 + ".xlsx")
    sheet1 = wb1[sheetName1]
    wb2 = openpyxl.load_workbook(fileName2 + ".xlsx")
    sheet2 = wb2[sheetName2]

    # if 'sheet' appears randomly we can delete it
    # rm = out.get_sheet_by_name('Sheet')
    # out.remove_sheet(rm)

    first = 2
    last = sheet2.max_row
    start = sheet1.max_row

    '''
    for row in range (first, start + 1):
        sheet1[CATEGORIES + str(row)].value = str(sheet1[CATEGORIES + str(row)].value) + ":" + str(sheet1['AH' + str(row)].value) + ":" + str(sheet1['AI' + str(row)].value) + ":" + str(sheet1['AJ' + str(row)].value)
    '''

    for row in range (first, last + 1):
        sheet1[NAME + str(row + start)].value = sheet2["A" + str(row)].value + " " + sheet2["B" + str(row)].value
        sheet1[GIVEN_NAME + str(row + start)].value = sheet2["A" + str(row)].value
        sheet1[FAMILY_NAME + str(row + start)].value = sheet2["B" + str(row)].value
        sheet1[EMAIL1_VALUE + str(row + start)].value = sheet2["C" + str(row)].value
        sheet1[ORGANIZATION1_NAME + str(row + start)].value = sheet2["D" + str(row)].value
        sheet1[ORGANIZATION1_TITLE + str(row + start)].value = sheet2["E" + str(row)].value
        sheet1[GROUP_MEMBERSHIP + str(row + start)].value = sheet2["F" + str(row)].value

    wb1.save("combined.xlsx")

    if printing:
        print()
        print("Done!")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Combine combined list with a CMAShipping List
# Copies all contacts from sheet 1 to sheet 2
# combine_with_CMA(fileName1, sheetName1, fileName2, sheetName2)
#{{{
'''
- loop through list 1 putting in list 2
  - FORMULA: row + (start + 1) - first
'''
def combine_with_CMA(fileName1, sheetName1, fileName2, sheetName2):
    if printing:
        print("Opening...")
    wb1 = openpyxl.load_workbook(fileName1 + ".xlsx")
    sheet = wb1[sheetName1]
    wb2 = openpyxl.load_workbook(fileName2 + ".xlsx")
    outsheet = wb2[sheetName2]

    first = 3
    last = sheet.max_row
    start = outsheet.max_row

    for row in range (first, last + 1):
        index = row + (start + 1) - first
        if printing:
            print(str(row) + " " + sheet['A' + str(row)].value)
        names = determine_names(sheet['A' + str(row)].value)

        # split name here
        outsheet[FIRST_NAME           + str(index)].value = names['first_name']
        outsheet[MIDDLE_NAME          + str(index)].value = names['middle_name']
        outsheet[LAST_NAME            + str(index)].value = names['last_name']

        # position and company
        outsheet[COMPANY              + str(index)].value = sheet['C' + str(row)].value
        outsheet[JOB_TITLE            + str(index)].value = sheet['B' + str(row)].value

        # put address stuff here
        outsheet[BUSINESS_STREET      + str(index)].value = sheet['D' + str(row)].value
        outsheet[BUSINESS_CITY        + str(index)].value = sheet['G' + str(row)].value
        outsheet[BUSINESS_STATE       + str(index)].value = sheet['H' + str(row)].value
        outsheet[BUSINESS_POSTAL_CODE + str(index)].value = sheet['I' + str(row)].value
        outsheet[BUSINESS_COUNTRY     + str(index)].value = sheet['J' + str(row)].value
        
        # contact info
        outsheet[BUSINESS_PHONE       + str(index)].value = sheet['K' + str(row)].value
        outsheet[EMAIL_ADDRESS        + str(index)].value = sheet['L' + str(row)].value

        # location info
        outsheet[LATITUDE             + str(index)].value = sheet['M' + str(row)].value
        outsheet[LONGITUDE            + str(index)].value = sheet['N' + str(row)].value

        # put CMA SHIPPING
        outsheet[CATEGORIES           + str(index)].value = 'CMA Shipping conference'
        # put March 22, 2018
        outsheet[CONNECTED_ON         + str(index)].value = '03/22/2018'


    wb2.save("combined.xlsx")

    if printing:
        print()
        print("Done!")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Remove Duplicate Contacts
# remove_duplicate_contacts(fileName, sheetName, first, last)
#{{{
# all rows algorithm
# - the first row is automatically its own entity
# - for all other rows take combination of primary contact and current contact we are looking at and compare edit distance to it
#   - if it is a match combine and keep looking through list until we find a bad match
#   - if distance is way off save old contact info. store new contact and keep going
# def remove_duplicate_contacts(fileName, first, match_threshold, *cols):
def remove_duplicate_contacts(*args):
    # turn the arguments into variable names
    args = args[0]
    fileName = args[1]
    first = int(args[2])
    low_threshold = int(args[3])
    cols = args[4:]
    if printing:
        print("Opening...")
    # Open the file for editing
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("contacts")
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    # Create a new file to store duplicate contacts
    dupe = openpyxl.Workbook()
    # Open the worksheet we want to edit
    dupesheet = dupe.create_sheet("contacts")
    # if 'sheet' appears randomly we can delete it
    rm = dupe.get_sheet_by_name('Sheet')
    dupe.remove_sheet(rm)

    # - create an object for a new primary contact and account name pair
    #   - store previous object in a new sheet
    #   - store all information here
    #   - look at next contact and see if it matches (edit distance)
    #     - if it is a match combine and keep going, otherwise repeat
    compare = ""
    current = ""
    last = sheet.max_row
    count = 1
    dupes = 2
    contact = Contact()

    # Create Headers
    #{{{
    if saving:
        outsheet[FIRST_NAME           + '1'].value = "First_name"
        outsheet[MIDDLE_NAME          + '1'].value = "Middle_name"
        outsheet[LAST_NAME            + '1'].value = "Last_name"
        outsheet[TITLE                + '1'].value = "Title"
        outsheet[SUFFIX               + '1'].value = "Suffix"
        outsheet[WEB_PAGE             + '1'].value = "Web_page"
        outsheet[NOTES                + '1'].value = "Notes"
        outsheet[EMAIL_ADDRESS        + '1'].value = "Email_address"
        outsheet[EMAIL_ADDRESS2       + '1'].value = "Email_address2"
        outsheet[EMAIL_ADDRESS3       + '1'].value = "Email_address3"
        outsheet[HOME_PHONE           + '1'].value = "Home_phone"
        outsheet[MOBILE_PHONE         + '1'].value = "Mobile_phone"
        outsheet[HOME_ADDRESS         + '1'].value = "Home_address"
        outsheet[HOME_STREET          + '1'].value = "Home_street"
        outsheet[HOME_CITY            + '1'].value = "Home_city"
        outsheet[HOME_STATE           + '1'].value = "Home_state"
        outsheet[HOME_POSTAL_CODE     + '1'].value = "Home_postal_code"
        outsheet[HOME_COUNTRY         + '1'].value = "Home_country"
        outsheet[CONTACT_MAIN_PHONE   + '1'].value = "Contact_main_phone"
        outsheet[BUSINESS_PHONE       + '1'].value = "Business_phone"
        outsheet[BUSINESS_PHONE2      + '1'].value = "Business_phone2"
        outsheet[BUSINESS_FAX         + '1'].value = "Business_fax"
        outsheet[COMPANY              + '1'].value = "Company"
        outsheet[JOB_TITLE            + '1'].value = "Job_title"
        outsheet[DEPARTMENT           + '1'].value = "Department"
        outsheet[OFFICE_LOCATION      + '1'].value = "Office_location"
        outsheet[BUSINESS_ADDRESS     + '1'].value = "Business_address"
        outsheet[BUSINESS_STREET      + '1'].value = "Business_street"
        outsheet[BUSINESS_CITY        + '1'].value = "Business_city"
        outsheet[BUSINESS_STATE       + '1'].value = "Business_state"
        outsheet[BUSINESS_POSTAL_CODE + '1'].value = "Business_postal_code"
        outsheet[BUSINESS_COUNTRY     + '1'].value = "Business_country"
        outsheet[CATEGORIES           + '1'].value = "Categories"
        outsheet[CONNECTED_ON         + '1'].value = "Connected On"
        outsheet[LATITUDE             + '1'].value = "Latitude"
        outsheet[LONGITUDE            + '1'].value = "Longitude"

        dupesheet[FIRST_NAME           + '1'].value = "First_name"
        dupesheet[MIDDLE_NAME          + '1'].value = "Middle_name"
        dupesheet[LAST_NAME            + '1'].value = "Last_name"
        dupesheet[TITLE                + '1'].value = "Title"
        dupesheet[SUFFIX               + '1'].value = "Suffix"
        dupesheet[WEB_PAGE             + '1'].value = "Web_page"
        dupesheet[NOTES                + '1'].value = "Notes"
        dupesheet[EMAIL_ADDRESS        + '1'].value = "Email_address"
        dupesheet[EMAIL_ADDRESS2       + '1'].value = "Email_address2"
        dupesheet[EMAIL_ADDRESS3       + '1'].value = "Email_address3"
        dupesheet[HOME_PHONE           + '1'].value = "Home_phone"
        dupesheet[MOBILE_PHONE         + '1'].value = "Mobile_phone"
        dupesheet[HOME_ADDRESS         + '1'].value = "Home_address"
        dupesheet[HOME_STREET          + '1'].value = "Home_street"
        dupesheet[HOME_CITY            + '1'].value = "Home_city"
        dupesheet[HOME_STATE           + '1'].value = "Home_state"
        dupesheet[HOME_POSTAL_CODE     + '1'].value = "Home_postal_code"
        dupesheet[HOME_COUNTRY         + '1'].value = "Home_country"
        dupesheet[CONTACT_MAIN_PHONE   + '1'].value = "Contact_main_phone"
        dupesheet[BUSINESS_PHONE       + '1'].value = "Business_phone"
        dupesheet[BUSINESS_PHONE2      + '1'].value = "Business_phone2"
        dupesheet[BUSINESS_FAX         + '1'].value = "Business_fax"
        dupesheet[COMPANY              + '1'].value = "Company"
        dupesheet[JOB_TITLE            + '1'].value = "Job_title"
        dupesheet[DEPARTMENT           + '1'].value = "Department"
        dupesheet[OFFICE_LOCATION      + '1'].value = "Office_location"
        dupesheet[BUSINESS_ADDRESS     + '1'].value = "Business_address"
        dupesheet[BUSINESS_STREET      + '1'].value = "Business_street"
        dupesheet[BUSINESS_CITY        + '1'].value = "Business_city"
        dupesheet[BUSINESS_STATE       + '1'].value = "Business_state"
        dupesheet[BUSINESS_POSTAL_CODE + '1'].value = "Business_postal_code"
        dupesheet[BUSINESS_COUNTRY     + '1'].value = "Business_country"
        dupesheet[CATEGORIES           + '1'].value = "Categories"
        dupesheet[CONNECTED_ON         + '1'].value = "Connected On"
        dupesheet[LATITUDE             + '1'].value = "Latitude"
        dupesheet[LONGITUDE            + '1'].value = "Longitude"
    #}}}

    for row in range (first, last + 1):
        # if the previous value is blank we create the new object and store information in it
        '''
        first_name = str(sheet[first_name_col + str(row)].value)
        if first_name != "":
            standardize_str(first_name)
        '''
        compareCriteria = ""
        for col in cols:
            compareCriteria = compareCriteria + standardize_str(sheet[col + str(row)].value) + " "
        if row == first:
            # compare = first_name + " " + last_name
            compare = compareCriteria
            contact = new_contact_from_sheet(sheet, row)
            count = count + 1
        else:
            # current = first_name + " " + last_name
            current = compareCriteria
            match = edit_distance(compare, current, low_threshold, high_threshold)
            matchingSuffixes = standardize_str(sheet[SUFFIX + str(row)].value) == standardize_str(sheet[SUFFIX + str(row - 1)].value)
            # combine information and move on
            if match and matchingSuffixes:
                contact = update_contact_from_sheet(sheet, row, contact)
                '''
                - if there is a duplicate
                - store previous item in list in duplicate list
                - store the duplicate in next spot in duplicate list
                '''
                # combine information
                # keep original
                dupes = dupes + 1

                dupeContact = new_contact_from_sheet(sheet, row - 1)

                #{{{
                dupesheet[FIRST_NAME           + str(dupes)].value = dupeContact.first_name 
                dupesheet[MIDDLE_NAME          + str(dupes)].value = dupeContact.middle_name
                dupesheet[LAST_NAME            + str(dupes)].value = dupeContact.last_name
                dupesheet[TITLE                + str(dupes)].value = dupeContact.title
                dupesheet[SUFFIX               + str(dupes)].value = dupeContact.suffix
                dupesheet[WEB_PAGE             + str(dupes)].value = dupeContact.web_page
                dupesheet[NOTES                + str(dupes)].value = dupeContact.notes
                dupesheet[EMAIL_ADDRESS        + str(dupes)].value = dupeContact.email_address
                dupesheet[EMAIL_ADDRESS2       + str(dupes)].value = dupeContact.email_address2
                dupesheet[EMAIL_ADDRESS3       + str(dupes)].value = dupeContact.email_address3
                dupesheet[HOME_PHONE           + str(dupes)].value = dupeContact.home_phone
                dupesheet[MOBILE_PHONE         + str(dupes)].value = dupeContact.mobile_phone
                dupesheet[HOME_ADDRESS         + str(dupes)].value = dupeContact.home_address
                dupesheet[HOME_STREET          + str(dupes)].value = dupeContact.home_street
                dupesheet[HOME_CITY            + str(dupes)].value = dupeContact.home_city
                dupesheet[HOME_STATE           + str(dupes)].value = dupeContact.home_state
                dupesheet[HOME_POSTAL_CODE     + str(dupes)].value = dupeContact.home_postal_code
                dupesheet[HOME_COUNTRY         + str(dupes)].value = dupeContact.home_country
                dupesheet[CONTACT_MAIN_PHONE   + str(dupes)].value = dupeContact.contact_main_phone
                dupesheet[BUSINESS_PHONE       + str(dupes)].value = dupeContact.business_phone
                dupesheet[BUSINESS_PHONE2      + str(dupes)].value = dupeContact.business_phone2
                dupesheet[BUSINESS_FAX         + str(dupes)].value = dupeContact.business_fax
                dupesheet[COMPANY              + str(dupes)].value = dupeContact.company
                dupesheet[JOB_TITLE            + str(dupes)].value = dupeContact.job_title
                dupesheet[DEPARTMENT           + str(dupes)].value = dupeContact.department
                dupesheet[OFFICE_LOCATION      + str(dupes)].value = dupeContact.office_location
                dupesheet[BUSINESS_ADDRESS     + str(dupes)].value = dupeContact.business_address
                dupesheet[BUSINESS_STREET      + str(dupes)].value = dupeContact.business_street
                dupesheet[BUSINESS_CITY        + str(dupes)].value = dupeContact.business_city
                dupesheet[BUSINESS_STATE       + str(dupes)].value = dupeContact.business_state
                dupesheet[BUSINESS_POSTAL_CODE + str(dupes)].value = dupeContact.business_postal_code
                dupesheet[BUSINESS_COUNTRY     + str(dupes)].value = dupeContact.business_country
                dupesheet[CATEGORIES           + str(dupes)].value = dupeContact.categories
                dupesheet[CONNECTED_ON         + str(dupes)].value = dupeContact.connected_on
                dupesheet['AI' + str(dupes)].value = row - 1
                #}}}

                # keep duplicate
                dupes = dupes + 1

                dupeContact = new_contact_from_sheet(sheet, row)
                #{{{

                dupesheet[FIRST_NAME           + str(dupes)].value = dupeContact.first_name 
                dupesheet[MIDDLE_NAME          + str(dupes)].value = dupeContact.middle_name
                dupesheet[LAST_NAME            + str(dupes)].value = dupeContact.last_name
                dupesheet[TITLE                + str(dupes)].value = dupeContact.title
                dupesheet[SUFFIX               + str(dupes)].value = dupeContact.suffix
                dupesheet[WEB_PAGE             + str(dupes)].value = dupeContact.web_page
                dupesheet[NOTES                + str(dupes)].value = dupeContact.notes
                dupesheet[EMAIL_ADDRESS        + str(dupes)].value = dupeContact.email_address
                dupesheet[EMAIL_ADDRESS2       + str(dupes)].value = dupeContact.email_address2
                dupesheet[EMAIL_ADDRESS3       + str(dupes)].value = dupeContact.email_address3
                dupesheet[HOME_PHONE           + str(dupes)].value = dupeContact.home_phone
                dupesheet[MOBILE_PHONE         + str(dupes)].value = dupeContact.mobile_phone
                dupesheet[HOME_ADDRESS         + str(dupes)].value = dupeContact.home_address
                dupesheet[HOME_STREET          + str(dupes)].value = dupeContact.home_street
                dupesheet[HOME_CITY            + str(dupes)].value = dupeContact.home_city
                dupesheet[HOME_STATE           + str(dupes)].value = dupeContact.home_state
                dupesheet[HOME_POSTAL_CODE     + str(dupes)].value = dupeContact.home_postal_code
                dupesheet[HOME_COUNTRY         + str(dupes)].value = dupeContact.home_country
                dupesheet[CONTACT_MAIN_PHONE   + str(dupes)].value = dupeContact.contact_main_phone
                dupesheet[BUSINESS_PHONE       + str(dupes)].value = dupeContact.business_phone
                dupesheet[BUSINESS_PHONE2      + str(dupes)].value = dupeContact.business_phone2
                dupesheet[BUSINESS_FAX         + str(dupes)].value = dupeContact.business_fax
                dupesheet[COMPANY              + str(dupes)].value = dupeContact.company
                dupesheet[JOB_TITLE            + str(dupes)].value = dupeContact.job_title
                dupesheet[DEPARTMENT           + str(dupes)].value = dupeContact.department
                dupesheet[OFFICE_LOCATION      + str(dupes)].value = dupeContact.office_location
                dupesheet[BUSINESS_ADDRESS     + str(dupes)].value = dupeContact.business_address
                dupesheet[BUSINESS_STREET      + str(dupes)].value = dupeContact.business_street
                dupesheet[BUSINESS_CITY        + str(dupes)].value = dupeContact.business_city
                dupesheet[BUSINESS_STATE       + str(dupes)].value = dupeContact.business_state
                dupesheet[BUSINESS_POSTAL_CODE + str(dupes)].value = dupeContact.business_postal_code
                dupesheet[BUSINESS_COUNTRY     + str(dupes)].value = dupeContact.business_country
                dupesheet[CATEGORIES           + str(dupes)].value = dupeContact.categories
                dupesheet[CONNECTED_ON         + str(dupes)].value = dupeContact.connected_on
                dupesheet['AI' + str(dupes)].value = row
                #}}}

                dupes = dupes + 1

                # save the combined contact
                #{{{
                dupesheet[FIRST_NAME           + str(dupes)].value = contact.first_name 
                dupesheet[MIDDLE_NAME          + str(dupes)].value = contact.middle_name
                dupesheet[LAST_NAME            + str(dupes)].value = contact.last_name
                dupesheet[TITLE                + str(dupes)].value = contact.title
                dupesheet[SUFFIX               + str(dupes)].value = contact.suffix
                dupesheet[WEB_PAGE             + str(dupes)].value = contact.web_page
                dupesheet[NOTES                + str(dupes)].value = contact.notes
                dupesheet[EMAIL_ADDRESS        + str(dupes)].value = contact.email_address
                dupesheet[EMAIL_ADDRESS2       + str(dupes)].value = contact.email_address2
                dupesheet[EMAIL_ADDRESS3       + str(dupes)].value = contact.email_address3
                dupesheet[HOME_PHONE           + str(dupes)].value = contact.home_phone
                dupesheet[MOBILE_PHONE         + str(dupes)].value = contact.mobile_phone
                dupesheet[HOME_ADDRESS         + str(dupes)].value = contact.home_address
                dupesheet[HOME_STREET          + str(dupes)].value = contact.home_street
                dupesheet[HOME_CITY            + str(dupes)].value = contact.home_city
                dupesheet[HOME_STATE           + str(dupes)].value = contact.home_state
                dupesheet[HOME_POSTAL_CODE     + str(dupes)].value = contact.home_postal_code
                dupesheet[HOME_COUNTRY         + str(dupes)].value = contact.home_country
                dupesheet[CONTACT_MAIN_PHONE   + str(dupes)].value = contact.contact_main_phone
                dupesheet[BUSINESS_PHONE       + str(dupes)].value = contact.business_phone
                dupesheet[BUSINESS_PHONE2      + str(dupes)].value = contact.business_phone2
                dupesheet[BUSINESS_FAX         + str(dupes)].value = contact.business_fax
                dupesheet[COMPANY              + str(dupes)].value = contact.company
                dupesheet[JOB_TITLE            + str(dupes)].value = contact.job_title
                dupesheet[DEPARTMENT           + str(dupes)].value = contact.department
                dupesheet[OFFICE_LOCATION      + str(dupes)].value = contact.office_location
                dupesheet[BUSINESS_ADDRESS     + str(dupes)].value = contact.business_address
                dupesheet[BUSINESS_STREET      + str(dupes)].value = contact.business_street
                dupesheet[BUSINESS_CITY        + str(dupes)].value = contact.business_city
                dupesheet[BUSINESS_STATE       + str(dupes)].value = contact.business_state
                dupesheet[BUSINESS_POSTAL_CODE + str(dupes)].value = contact.business_postal_code
                dupesheet[BUSINESS_COUNTRY     + str(dupes)].value = contact.business_country
                dupesheet[CATEGORIES           + str(dupes)].value = contact.categories
                dupesheet[CONNECTED_ON         + str(dupes)].value = contact.connected_on
                #}}}

                # create a blank space
                dupes = dupes + 1
            # store the information and create a new contact
            else:
                # store information
                #{{{
                outsheet[FIRST_NAME           + str(count)].value = contact.first_name 
                outsheet[MIDDLE_NAME          + str(count)].value = contact.middle_name
                outsheet[LAST_NAME            + str(count)].value = contact.last_name
                outsheet[TITLE                + str(count)].value = contact.title
                outsheet[SUFFIX               + str(count)].value = contact.suffix
                outsheet[WEB_PAGE             + str(count)].value = contact.web_page
                outsheet[NOTES                + str(count)].value = contact.notes
                outsheet[EMAIL_ADDRESS        + str(count)].value = contact.email_address
                outsheet[EMAIL_ADDRESS2       + str(count)].value = contact.email_address2
                outsheet[EMAIL_ADDRESS3       + str(count)].value = contact.email_address3
                outsheet[HOME_PHONE           + str(count)].value = contact.home_phone
                outsheet[MOBILE_PHONE         + str(count)].value = contact.mobile_phone
                outsheet[HOME_ADDRESS         + str(count)].value = contact.home_address
                outsheet[HOME_STREET          + str(count)].value = contact.home_street
                outsheet[HOME_CITY            + str(count)].value = contact.home_city
                outsheet[HOME_STATE           + str(count)].value = contact.home_state
                outsheet[HOME_POSTAL_CODE     + str(count)].value = contact.home_postal_code
                outsheet[HOME_COUNTRY         + str(count)].value = contact.home_country
                outsheet[CONTACT_MAIN_PHONE   + str(count)].value = contact.contact_main_phone
                outsheet[BUSINESS_PHONE       + str(count)].value = contact.business_phone
                outsheet[BUSINESS_PHONE2      + str(count)].value = contact.business_phone2
                outsheet[BUSINESS_FAX         + str(count)].value = contact.business_fax
                outsheet[COMPANY              + str(count)].value = contact.company
                outsheet[JOB_TITLE            + str(count)].value = contact.job_title
                outsheet[DEPARTMENT           + str(count)].value = contact.department
                outsheet[OFFICE_LOCATION      + str(count)].value = contact.office_location
                outsheet[BUSINESS_ADDRESS     + str(count)].value = contact.business_address
                outsheet[BUSINESS_STREET      + str(count)].value = contact.business_street
                outsheet[BUSINESS_CITY        + str(count)].value = contact.business_city
                outsheet[BUSINESS_STATE       + str(count)].value = contact.business_state
                outsheet[BUSINESS_POSTAL_CODE + str(count)].value = contact.business_postal_code
                outsheet[BUSINESS_COUNTRY     + str(count)].value = contact.business_country
                outsheet[CATEGORIES           + str(count)].value = contact.categories
                outsheet[CONNECTED_ON         + str(count)].value = contact.connected_on
                outsheet[LATITUDE             + str(count)].value = contact.latitude
                outsheet[LONGITUDE            + str(count)].value = contact.longitude
                #}}}
                # reset compare value
                # compare = first_name + " " + last_name
                compare = compareCriteria
                # create a new contact
                contact = new_contact_from_sheet(sheet, row)
                count = count + 1

    if printing:
        print()
        print("Out of " + str(1 + last - first) + " companies " + str(count) + " were unique contacts")
        print("Saving...")

    out.save("purged" + str(cols) + ".xlsx")
    dupe.save("duplicates" + str(cols) + ".xlsx")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

    if printing:
        print()
        print("Done!")
#}}}

# Fix Country Names
# fix_country_names(sheet)
# {{{
def fix_country_names(fileName, sheetName):
    if printing:
        print("Opening...")
    wb = openpyxl.load_workbook(fileName + ".xlsx")
    sheet = wb[sheetName]

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')

    compare = ""
    current = ""
    count = 1
    first = 2
    last = sheet.max_row

    for row in range (first, last + 1):
        # if the previous value is blank we create the new object and store information in it
        country = str(sheet[BUSINESS_COUNTRY + str(row)].value)
        if country == 'None':
            break
        if row == first:
            compare = country
            count = count + 1
        else:
            current = country
            match = edit_distance(compare, current, low_threshold, high_threshold)
            # combine information and move on
            if match:
                if printing:
                    print("Now changing row " + str(row))
                sheet[BUSINESS_COUNTRY + str(row)].value = compare
            else:
                # reset compare value
                compare = country
                count = count + 1

    if printing:
        print("Saving...")
    wb.save("newCombined.xlsx")

    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# combine_contacts(2, 100)
combine_contacts("google", "contacts", 'linkedIn', 'contacts')
# remove_duplicate_contacts("combined", "contacts", 2, 6903)
# combine_with_CMA("CMAShipping", "Attendees", "gli", "contacts")
# standardize_USA(sys.argv)
# remove_duplicate_contacts(sys.argv)
