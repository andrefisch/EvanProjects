import openpyxl
import os
import pygame
import re
import sys
import time

SHIPMENT_ID            = "A"
DEPARTURE_DATE         = "B"
EXPORT_LICENSE_HOLDERS = "D"
SHIPMENT_TYPE          = "E"
LOADING_PORTS          = "F"
LOADING_PORT_TYPE      = "G"
DESTINATION_PORTS      = "H"
DESTINATION_PORT_TYPE  = "I"
DURATION               = "J"
DISCHARGE_DATE         = "K"
CATTLE_LOAD            = "L"
CATTLE_LOSS            = "M"
CATTLE_PCT             = "N"
SHEEP_LOAD             = "O"
SHEEP_LOSS             = "P"
SHEEP_PCT              = "Q"
BUFFALO_LOAD           = "R"
BUFFALO_LOSS           = "S"
BUFFALO_PCT            = "T"
GOATS_LOAD             = "U"
GOATS_LOSS             = "V"
GOATS_PCT              = "W"
ALPACAS_LOAD           = "X"
ALPACAS_LOSS           = "Y"
ALPACAS_PCT            = "Z"
CAMELS_LOAD            = "AA"
CAMELS_LOSS            = "AB"
CAMELS_PCT             = "AC"
LLAMAS_LOAD            = "AD"
LLAMAS_LOSS            = "AE"
LLAMAS_PCT             = "AF"
DEER_LOAD              = "AG"
DEER_LOSS              = "AH"
DEER_PCT               = "AI"

def reformCol(string):
    #{{{
    arr = string.split('/')
    for i in range(0, len(arr)):
        arr[i] = arr[i].strip()
    return "/".join(arr)
    #}}}

def copyRow(fromSheet, toSheet, fromRow, toRow):
#{{{
    toSheet[SHIPMENT_ID            + str(toRow)].value = fromSheet[SHIPMENT_ID            + str(fromRow)].value 
    toSheet[DEPARTURE_DATE         + str(toRow)].value = fromSheet[DEPARTURE_DATE         + str(fromRow)].value 
    toSheet[EXPORT_LICENSE_HOLDERS + str(toRow)].value = fromSheet[EXPORT_LICENSE_HOLDERS + str(fromRow)].value 
    toSheet[SHIPMENT_TYPE          + str(toRow)].value = fromSheet[SHIPMENT_TYPE          + str(fromRow)].value 
    toSheet[LOADING_PORTS          + str(toRow)].value = fromSheet[LOADING_PORTS          + str(fromRow)].value 
    toSheet[LOADING_PORT_TYPE      + str(toRow)].value = fromSheet[LOADING_PORT_TYPE      + str(fromRow)].value 
    toSheet[DESTINATION_PORTS      + str(toRow)].value = fromSheet[DESTINATION_PORTS      + str(fromRow)].value 
    toSheet[DESTINATION_PORT_TYPE  + str(toRow)].value = fromSheet[DESTINATION_PORT_TYPE  + str(fromRow)].value 
    toSheet[DURATION               + str(toRow)].value = fromSheet[DURATION               + str(fromRow)].value 
    toSheet[DISCHARGE_DATE         + str(toRow)].value = fromSheet[DISCHARGE_DATE         + str(fromRow)].value 
    toSheet[CATTLE_LOAD            + str(toRow)].value = fromSheet[CATTLE_LOAD            + str(fromRow)].value 
    toSheet[CATTLE_LOSS            + str(toRow)].value = fromSheet[CATTLE_LOSS            + str(fromRow)].value 
    toSheet[CATTLE_PCT             + str(toRow)].value = fromSheet[CATTLE_PCT             + str(fromRow)].value 
    toSheet[SHEEP_LOAD             + str(toRow)].value = fromSheet[SHEEP_LOAD             + str(fromRow)].value 
    toSheet[SHEEP_LOSS             + str(toRow)].value = fromSheet[SHEEP_LOSS             + str(fromRow)].value 
    toSheet[SHEEP_PCT              + str(toRow)].value = fromSheet[SHEEP_PCT              + str(fromRow)].value 
    toSheet[BUFFALO_LOAD           + str(toRow)].value = fromSheet[BUFFALO_LOAD           + str(fromRow)].value 
    toSheet[BUFFALO_LOSS           + str(toRow)].value = fromSheet[BUFFALO_LOSS           + str(fromRow)].value 
    toSheet[BUFFALO_PCT            + str(toRow)].value = fromSheet[BUFFALO_PCT            + str(fromRow)].value 
    toSheet[GOATS_LOAD             + str(toRow)].value = fromSheet[GOATS_LOAD             + str(fromRow)].value 
    toSheet[GOATS_LOSS             + str(toRow)].value = fromSheet[GOATS_LOSS             + str(fromRow)].value 
    toSheet[GOATS_PCT              + str(toRow)].value = fromSheet[GOATS_PCT              + str(fromRow)].value 
    toSheet[ALPACAS_LOAD           + str(toRow)].value = fromSheet[ALPACAS_LOAD           + str(fromRow)].value 
    toSheet[ALPACAS_LOSS           + str(toRow)].value = fromSheet[ALPACAS_LOSS           + str(fromRow)].value 
    toSheet[ALPACAS_PCT            + str(toRow)].value = fromSheet[ALPACAS_PCT            + str(fromRow)].value 
    toSheet[CAMELS_LOAD            + str(toRow)].value = fromSheet[CAMELS_LOAD            + str(fromRow)].value 
    toSheet[CAMELS_LOSS            + str(toRow)].value = fromSheet[CAMELS_LOSS            + str(fromRow)].value 
    toSheet[CAMELS_PCT             + str(toRow)].value = fromSheet[CAMELS_PCT             + str(fromRow)].value 
    toSheet[LLAMAS_LOAD            + str(toRow)].value = fromSheet[LLAMAS_LOAD            + str(fromRow)].value 
    toSheet[LLAMAS_LOSS            + str(toRow)].value = fromSheet[LLAMAS_LOSS            + str(fromRow)].value 
    toSheet[LLAMAS_PCT             + str(toRow)].value = fromSheet[LLAMAS_PCT             + str(fromRow)].value 
    toSheet[DEER_LOAD              + str(toRow)].value = fromSheet[DEER_LOAD              + str(fromRow)].value 
    toSheet[DEER_LOSS              + str(toRow)].value = fromSheet[DEER_LOSS              + str(fromRow)].value 
    toSheet[DEER_PCT               + str(toRow)].value = fromSheet[DEER_PCT               + str(fromRow)].value 
#}}}

def splitExplortLicenses(args):
    print(args)
    fileName = args[0]
    # Create a new excel file
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("data")

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    # Open an existing excel file
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    #################
    # DO STUFF HERE #
    #################
    '''
    - log the shipment id then incrememnt it
    - split row based on how many export license holders there are
      - if multiple export license holders -> parcel in shipment type col
      - else -> full in shipment type col
    '''

    outsheet[SHIPMENT_ID            + '2'].value = "Shipment ID"
    outsheet[DEPARTURE_DATE         + '2'].value = "Departure Date"
    outsheet[EXPORT_LICENSE_HOLDERS + '2'].value = "Export License Holders"
    outsheet[SHIPMENT_TYPE          + '2'].value = "Shipment Type"
    outsheet[LOADING_PORTS          + '2'].value = "Loading Ports"
    outsheet[LOADING_PORT_TYPE      + '2'].value = "Loading Ports Type"
    outsheet[DESTINATION_PORTS      + '2'].value = "Destination Ports"
    outsheet[DESTINATION_PORT_TYPE  + '2'].value = "Destination Ports Type"
    outsheet[DURATION               + '2'].value = "Duration"
    outsheet[DISCHARGE_DATE         + '2'].value = "Discharge Date"
    outsheet[CATTLE_LOAD            + '2'].value = "Cattle Load"
    outsheet[CATTLE_LOSS            + '2'].value = "Cattle Loss"
    outsheet[CATTLE_PCT             + '2'].value = "Cattle PCT"
    outsheet[SHEEP_LOAD             + '2'].value = "Sheep Load"
    outsheet[SHEEP_LOSS             + '2'].value = "Sheep Loss"
    outsheet[SHEEP_PCT              + '2'].value = "Sheep PCT"
    outsheet[BUFFALO_LOAD           + '2'].value = "Buffalo Load"
    outsheet[BUFFALO_LOSS           + '2'].value = "Buffalo Loss"
    outsheet[BUFFALO_PCT            + '2'].value = "Buffalo PCT"
    outsheet[GOATS_LOAD             + '2'].value = "Goats Load"
    outsheet[GOATS_LOSS             + '2'].value = "Goats Loss"
    outsheet[GOATS_PCT              + '2'].value = "Goats PCT"
    outsheet[ALPACAS_LOAD           + '2'].value = "Alpacas Load"
    outsheet[ALPACAS_LOSS           + '2'].value = "Alpacas Loss"
    outsheet[ALPACAS_PCT            + '2'].value = "Alpacas PCT"
    outsheet[CAMELS_LOAD            + '2'].value = "Camels Load"
    outsheet[CAMELS_LOSS            + '2'].value = "Camels Loss"
    outsheet[CAMELS_PCT             + '2'].value = "Camels PCT"
    outsheet[LLAMAS_LOAD            + '2'].value = "Llamas Load"
    outsheet[LLAMAS_LOSS            + '2'].value = "Llamas Loss"
    outsheet[LLAMAS_PCT             + '2'].value = "Llamas PCT"
    outsheet[DEER_LOAD              + '2'].value = "Deer Load"
    outsheet[DEER_LOSS              + '2'].value = "Deer Loss"
    outsheet[DEER_PCT               + '2'].value = "Deer PCT"

    shipment = 0
    duplicates = 0
    start = 3
    end = sheet.max_row + 1
    for row in range(start, end):
        # log and increment shipment number
        sheet[SHIPMENT_ID + str(row)].value = shipment
        shipment = shipment + 1
        license_holders = sheet[EXPORT_LICENSE_HOLDERS + str(row)].value.split('/')
        # Are there multiple loading ports?
        info = sheet[LOADING_PORTS + str(row)].value
        print(row, info)
        if '/' in info:
            sheet[LOADING_PORTS     + str(row)].value = reformCol(info)
            sheet[LOADING_PORT_TYPE + str(row)].value = "Multiple"
        else:
            sheet[LOADING_PORT_TYPE + str(row)].value = "Singular"
        # Are there multiple destination ports?
        info = sheet[DESTINATION_PORTS + str(row)].value
        print(row, info)
        if '/' in info:
            sheet[DESTINATION_PORTS     + str(row)].value = reformCol(info)
            sheet[DESTINATION_PORT_TYPE + str(row)].value = "Multiple"
        else:
            sheet[DESTINATION_PORT_TYPE + str(row)].value = "Singular"
        # if there is only once license holder copy the row into the new sheet
        if len(license_holders) == 1:
            sheet[SHIPMENT_TYPE + str(row)].value = 'Full'
            copyRow(sheet, outsheet, row, row + duplicates)
            print("Singular: " + str(row + duplicates))
        # if there are multiple rows split them by export license holders
        else:
            sheet[SHIPMENT_TYPE + str(row)].value = 'Parcel'
            for i in range(0, len(license_holders)):
                print("Multiple: " + str(row + duplicates))
                copyRow(sheet, outsheet, row, row + duplicates)
                outsheet[EXPORT_LICENSE_HOLDERS + str(row + duplicates)].value = license_holders[i]
                if i != len(license_holders) - 1:
                    duplicates = duplicates + 1

    # Save the file
    out.save("splitLicenses.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

splitExplortLicenses(sys.argv[1:])
