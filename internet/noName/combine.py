import openpyxl
import os
# import pygame
import re
import sys
# import time

def fix_date(string):
    months = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}
    regexNum = '^(.*?) ([0-9]+), ([0-9]+)'
    matchNum = re.search(regexNum, string)
    if matchNum:
        mm   = matchNum.group(1)
        dd   = matchNum.group(2)
        yyyy = matchNum.group(3)
        if mm in months:
            return yyyy + '-' + months[mm] + '-' + dd
    return string

def fix_formatting(string):
    return string.replace('\\xe2\\x80\\x99', "'").replace('\\xc3\\xa9', 'e').replace('\\xc2\\xb4', "'").replace('\\xe2\\x80\\x9c', '"').replace('\\xe2\\x80\\x9d', '"').replace('\\n', ' ').replace('\\u00e9', 'e')

def combine():
    # Create a new excel file
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("transactions")
    # Open an existing excel file
    wb1 = openpyxl.load_workbook(sys.argv[1])
    sheet1 = wb1.worksheets[0]
    # Open an existing excel file
    wb2 = openpyxl.load_workbook(sys.argv[2])
    sheet2 = wb2.worksheets[0]

    # if 'sheet' appears randomly we can delete it
    rm = out['Sheet']
    out.remove(rm)

    #################
    # DO STUFF HERE #
    #################
    DATE   = "A"
    STARS  = "B"
    REVIEW = "C"
    SOURCE = "D"
    outsheet[DATE   + '1'].value = "DATE"
    outsheet[STARS  + '1'].value = "STARS"
    outsheet[REVIEW + '1'].value = "REVIEW"
    outsheet[SOURCE + '1'].value = "SOURCE"
    
    row = 2
    for i in range(2, sheet1.max_row + 1):
        outsheet[DATE   + str(row)].value = fix_date(sheet1[DATE   + str(i)].value)
        outsheet[STARS  + str(row)].value = sheet1[STARS  + str(i)].value
        outsheet[REVIEW + str(row)].value = fix_formatting(sheet1[REVIEW + str(i)].value)
        outsheet[SOURCE + str(row)].value = sheet1[SOURCE + str(i)].value
        row = row + 1

    for i in range(2, sheet2.max_row + 1):
        outsheet[DATE   + str(row)].value = fix_date(sheet2[DATE   + str(i)].value)
        outsheet[STARS  + str(row)].value = sheet2[STARS  + str(i)].value
        outsheet[REVIEW + str(row)].value = fix_formatting(sheet2[REVIEW + str(i)].value)
        outsheet[SOURCE + str(row)].value = sheet2[SOURCE + str(i)].value
        row = row + 1

    out.save("combined.xlsx")

    '''
    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
    '''

combine()
