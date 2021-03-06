import openpyxl
import pygame
import time

def reorganize():
    # Open an existing excel file
    wb = openpyxl.load_workbook('data' + ".xlsx")
    sheet = wb.worksheets[0]
    # Create a new excel file
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("data")

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    #################
    # DO STUFF HERE #
    #################
    COMPANY               = "A"
    FINANCIAL_INSTITUTION = "B"
    DEAL_TYPE             = "C"
    AMOUNT_USD            = "D"
    COMMENTS              = "E"
    DATE                  = "F"

    outsheet[COMPANY               + '1'].value = "Company"
    outsheet[FINANCIAL_INSTITUTION + '1'].value = "Financial Institution"
    outsheet[DEAL_TYPE             + '1'].value = "Deal Type"
    outsheet[AMOUNT_USD            + '1'].value = "Amount USD"
    outsheet[COMMENTS              + '1'].value = "Comments"
    outsheet[DATE                  + '1'].value = "Date"

    # Data will always be in A column
    deals = []
    deal = []
    for row in range(2, sheet.max_row + 1):
        info = sheet['A' + str(row)].value
        deal.append(info)
        if type(info) is not str:
            deals.append(deal)
            deal = []

    for row in range(0, len(deals)):
        deal = deals[row]
        if len(deals[row]) == 6:
            outsheet[COMPANY               + str(row + 2)].value = deal[0]
            outsheet[FINANCIAL_INSTITUTION + str(row + 2)].value = deal[1]
            outsheet[DEAL_TYPE             + str(row + 2)].value = deal[2]
            outsheet[AMOUNT_USD            + str(row + 2)].value = deal[3]
            outsheet[COMMENTS              + str(row + 2)].value = deal[4]
            outsheet[DATE                  + str(row + 2)].value = deal[5]
        elif len(deals[row]) == 4:
            outsheet[COMPANY               + str(row + 2)].value = deal[0]
            outsheet[DEAL_TYPE             + str(row + 2)].value = deal[1]
            outsheet[COMMENTS              + str(row + 2)].value = deal[2]
            outsheet[DATE                  + str(row + 2)].value = deal[3]
        # HARD ONE
        elif len(deals[row]) == 5:
            # If we are missing the financial transaction fill it out like this
            if deal[2][0] == "$":
                print(str(row + 2) + " is missing TRANSACTION")
                outsheet[COMPANY               + str(row + 2)].value = deal[0]
                outsheet[DEAL_TYPE             + str(row + 2)].value = deal[1]
                outsheet[AMOUNT_USD            + str(row + 2)].value = deal[2]
                outsheet[COMMENTS              + str(row + 2)].value = deal[3]
                outsheet[DATE                  + str(row + 2)].value = deal[4]
            # Otherwise we are missing dollar amount so fill out like this
            else:
                print(str(row + 2) + " is missing AMOUNT")
                outsheet[COMPANY               + str(row + 2)].value = deal[0]
                outsheet[FINANCIAL_INSTITUTION + str(row + 2)].value = deal[1]
                outsheet[DEAL_TYPE             + str(row + 2)].value = deal[2]
                outsheet[COMMENTS              + str(row + 2)].value = deal[3]
                outsheet[DATE                  + str(row + 2)].value = deal[4]

    print("Reorganized " + str(len(deals)) + " deals")

    out.save("organized.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

# Split the banks in column C into separate rows
def splitDeals():
    # Open an existing excel file
    wb = openpyxl.load_workbook('data' + ".xlsx")
    sheet = wb.worksheets[0]
    # Create a new excel file
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("data")

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    start = 2
    end = sheet.max_row + 1

    #################
    # DO STUFF HERE #
    #################
    ID                    = "A"
    COMPANY               = "B"
    FINANCIAL_INSTITUTION = "C"
    DEAL_TYPE             = "D"
    DEAL_SUB_TYPE         = "E"
    ORIGINAL_DEAL_TYPE    = "F"
    AMOUNT_MM_USD         = "G"
    COMMENTS              = "H"
    DATE                  = "I"

    outsheet[ID                    + '1'].value = "Transaction Id"
    outsheet[COMPANY               + '1'].value = "Company"
    outsheet[FINANCIAL_INSTITUTION + '1'].value = "Financial Institution"
    outsheet[DEAL_TYPE             + '1'].value = "Deal Type"
    outsheet[DEAL_SUB_TYPE         + '1'].value = "Deal Sub Type"
    outsheet[ORIGINAL_DEAL_TYPE    + '1'].value = "Original Deal Type"
    outsheet[AMOUNT_MM_USD         + '1'].value = "Amount MM USD"
    outsheet[COMMENTS              + '1'].value = "Comments"
    outsheet[DATE                  + '1'].value = "Date"

    # Only care about column C    
    count = 2
    for row in range(start, end):
        if sheet[FINANCIAL_INSTITUTION + str(row)].value:
            banks = sheet[FINANCIAL_INSTITUTION + str(row)].value.replace(',', '\n').split('\n')
        else:
            banks = ["Undisclosed"]
        for bank in banks:
            outsheet[ID                    + str(count)].value = sheet[ID                 + str(row)].value 
            outsheet[COMPANY               + str(count)].value = sheet[COMPANY            + str(row)].value 
            outsheet[FINANCIAL_INSTITUTION + str(count)].value = bank.strip()
            outsheet[DEAL_TYPE             + str(count)].value = sheet[DEAL_TYPE          + str(row)].value 
            outsheet[DEAL_SUB_TYPE         + str(count)].value = sheet[DEAL_SUB_TYPE      + str(row)].value 
            outsheet[ORIGINAL_DEAL_TYPE    + str(count)].value = sheet[ORIGINAL_DEAL_TYPE + str(row)].value 
            outsheet[AMOUNT_MM_USD         + str(count)].value = sheet[AMOUNT_MM_USD      + str(row)].value 
            outsheet[COMMENTS              + str(count)].value = sheet[COMMENTS           + str(row)].value 
            outsheet[DATE                  + str(count)].value = sheet[DATE               + str(row)].value 
            count = count + 1

    out.save("organized.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

splitDeals()
# reorganize()
