import openpyxl
import os
import pygame
import sys
import time

# GETS ALL COLUMNS FROM THIS ROW
# print(sheet[1])

def find_addresses():
    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    fileName = args[0]
    fileName2 = args[1]

    # Open an existing excel file
    wb = openpyxl.load_workbook(fileName)
    outsheet = wb.worksheets[0]

    wb2 = openpyxl.load_workbook(fileName2)
    insheet = wb2.worksheets[0]

    OUT_DAUGHTER  = "C"
    OUT_PARENT    = "B"
    OUT_ADDRESS   = "D"
    OUT_CONTACT   = "K"
    OUT_ROWSOURCE = "L"

    IN_COMPANY   = "CB"
    IN_NAME      = "A"
    # ADDRESS STUFF
    IN_FORMATTED = "BA"
    IN_STREET    = "BB"
    IN_CITY      = "BC"
    IN_REGION    = "BE"
    IN_ZIPCODE   = "BF"
    IN_COUNTRY   = "BG"
    IN_EXTENDED  = "BH"

    # get all 7 components of address from the insheet and put them together 
    # get_in_address(sheet, row)
    #{{{
    def get_in_address(sheet, row):
        formatted = ""
        street    = ""
        city      = ""
        region    = ""
        zipcode   = ""
        country   = ""
        extended  = ""

        if sheet[IN_FORMATTED + str(row)].value:
            formatted = str(sheet[IN_FORMATTED + str(row)].value)
        if sheet[IN_STREET + str(row)].value:
            street    = str(sheet[IN_STREET    + str(row)].value)
        if sheet[IN_CITY + str(row)].value:
            city      = str(sheet[IN_CITY      + str(row)].value)
        if sheet[IN_REGION + str(row)].value:
            region    = str(sheet[IN_REGION    + str(row)].value)
        if sheet[IN_ZIPCODE + str(row)].value:
            zipcode   = str(sheet[IN_ZIPCODE   + str(row)].value)
        if sheet[IN_COUNTRY + str(row)].value:
            country   = str(sheet[IN_COUNTRY   + str(row)].value)
        if sheet[IN_EXTENDED + str(row)].value:
            extended  = str(sheet[IN_EXTENDED  + str(row)].value)

        print(str(row) + ": " + formatted + "-" + street + "-" + city  + "-" + region + "-" + zipcode + "-" + country + "-" + extended)
        return (formatted + " " + street + " " + city  + " " + region + " " + zipcode + " " + country + " " + extended)
    #}}}

    #################
    # DO STUFF HERE #
    #################
    '''
    - loop through in_sheet and collect all companies and row numbers they were on in dict
    - loop through out_sheet and look for matching companies
      - when you find a company
        - look up address from in_sheet, save name and row information came from
      - if a company is not found look up parent company
        - look up address from in_sheet, save name and row information came from
      - if still not found take note
    '''
    no_address  = 0
    already_had = 0

    first = 2
    last = insheet.max_row + 1
    companies = {}
    for row in range(first, last):
        company = insheet[IN_COMPANY + str(row)].value
        if company:
            companies[company] = row

    last = outsheet.max_row + 1
    for row in range (first, last):
        address  = outsheet[OUT_ADDRESS  + str(row)].value
        daughter = outsheet[OUT_DAUGHTER + str(row)].value
        parent   = outsheet[OUT_PARENT   + str(row)].value
        # if there is no parent company listed then daughter is parent company
        # but dont make parent = daughter because we dont want to search for the same company twice
        if parent == None:
            outsheet[OUT_PARENT    + str(row)].value = daughter
        # we should only be doing this stuff if there is no address already given
        if address == None:
            if daughter and daughter in companies:
                inrow = companies[daughter]
                outsheet[OUT_ADDRESS   + str(row)].value = get_in_address(insheet, inrow)
                outsheet[OUT_CONTACT   + str(row)].value = insheet[IN_NAME + str(inrow)].value
                outsheet[OUT_ROWSOURCE + str(row)].value = inrow
            elif parent and parent in companies:
                inrow = companies[parent]
                outsheet[OUT_ADDRESS   + str(row)].value = get_in_address(insheet, inrow)
                outsheet[OUT_CONTACT   + str(row)].value = insheet[IN_NAME + str(inrow)].value
                outsheet[OUT_ROWSOURCE + str(row)].value = inrow
            else:
                outsheet[OUT_ROWSOURCE + str(row)].value = "COULD NOT FIND"
                no_address = no_address + 1
        else:
            already_had = already_had + 1



    print("Processed " + str(last - first) + " rows...")
    print("Spreadsheet alread had :      ", already_had)
    print("Could not find an address for:", no_address)

    # add the word 'formatted' and save the new file where the original is
    newName = 'addresses'
    index = fileName[::-1].find('/')
    end = fileName[-index - 1:]
    fileName = fileName[:-index - 1] + newName + end[0].capitalize() + end[1:]
    print("Saving " + fileName)
    wb.save(fileName)

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

find_addresses()
