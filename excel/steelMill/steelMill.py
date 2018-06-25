import openpyxl
import os
import pygame
import re
import string
import sys
import time

# GETS ALL COLUMNS FROM THIS ROW
# print(sheet[1])

ICOMPANY = "A"
ISOURCE  = "B"
ITYPE    = "C"
IDATE    = "D"

CCOMPANY_CHINA = "A"
CCOUNTRY       = "B"
CCOMPANY_INTL  = "C"
CCOUNTRY_INTL  = "D"
CCONTACT       = "E"
CMOBILE        = "F"
CEMAIL         = "G"
CSOURCE        = "H"
CDATE          = "I"
CTYPE          = "J"

def remove_question_mark(string):
    if string:
        if "?" in str(string):
            print("Found a question mark", string)
            return ""
        else:
            return string
    else:
        return ""

def format_company(company):
    if company:
        print("Fixing a company", company)
        company = company.title().replace(",", ' ').replace('.', '').strip()
        print("Fixed!", company)
        return company
    else:
        return ""

# Strip leading numbers from file
# strip_leading_numbers(fileName)
#{{{
def strip_leading_numbers():
    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    fileName = args[0]

    # Open an existing excel file
    wb = openpyxl.load_workbook(fileName)
    sheet = wb['Globalore List']


    #################
    # DO STUFF HERE #
    #################
    start = 2
    end = sheet.max_row + 1
    regexNumber = '^\d*\. ?'
    for row in range(start, end):
        company = sheet[ICOMPANY + str(row)].value
        print(row, company)
        sheet[ICOMPANY + str(row)].value = re.sub(regexNumber, '', company)

    wb.save("betterFile.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Split companies into parent and child
# split_companies(fileName)
#{{{
def split_companies():
    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    fileName = args[0]

    # Open an existing excel file
    print(fileName)
    wb = openpyxl.load_workbook(fileName)
    sheet2 = wb['CBMX LIST']

    #################
    # DO STUFF HERE #
    #################
    start = 2
    end = sheet2.max_row + 1
    extra = end
    '''
    - Move parent company to child company
    - Put china in country
    '''
    for row in range(start, end):
        # Make sure company is formatted correctly
        sheet2[CCOMPANY_CHINA + str(row)].value = format_compay(sheet2[CCOMPANY_CHINA + str(row)].value)
        sheet2[CCOMPANY_INTL + str(row)].value = format_compay(sheet2[CCOMPANY_INTL + str(row)].value)

        # If there is no child company
        if not sheet2[CCOMPANY_INTL + str(row)].value: 
            extra = extra + 1
            # Create a new row
            sheet2[CCOMPANY_CHINA + str(extra)].value = sheet2[CCOMPANY_CHINA + str(row)].value
            sheet2[CCOMPANY_INTL  + str(extra)].value = sheet2[CCOMPANY_CHINA + str(row)].value
            sheet2[CCOUNTRY_INTL  + str(extra)].value = sheet2[CCOUNTRY       + str(row)].value

        # Remove all question marks from the row
        sheet2[CCOUNTRY       + str(row)].value = remove_question_mark(sheet2[CCOUNTRY + str(row)].value)
        sheet2[CCONTACT       + str(row)].value = remove_question_mark(sheet2[CCONTACT + str(row)].value)
        sheet2[CMOBILE        + str(row)].value = remove_question_mark(sheet2[CMOBILE  + str(row)].value)
        sheet2[CEMAIL         + str(row)].value = remove_question_mark(sheet2[CEMAIL   + str(row)].value)

    wb.save("splitFile.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Combine Globalore List and CBMX LIST
# combine_sheets(fileName)
#{{{        
def combine_sheets():
    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    fileName = args[0]

    # Open an existing excel file
    wb = openpyxl.load_workbook(fileName)
    sheet1 = wb['Globalore List']
    sheet2 = wb['CBMX LIST']

    #################
    # DO STUFF HERE #
    #################
    start = 2
    end = sheet1.max_row + 1
    begin = sheet2.max_row + 1
    for row in range(start, begin):
        # Format the company names
        sheet2[CCOMPANY_CHINA + str(row)].value = format_company(sheet2[CCOMPANY_CHINA + str(row)].value)
        sheet2[CCOMPANY_INTL  + str(row)].value = format_company(sheet2[CCOMPANY_INTL  + str(row)].value)
        # Remove all question marks from the row
        sheet2[CCOUNTRY       + str(row)].value = remove_question_mark(sheet2[CCOUNTRY + str(row)].value)
        sheet2[CCONTACT       + str(row)].value = remove_question_mark(sheet2[CCONTACT + str(row)].value)
        sheet2[CMOBILE        + str(row)].value = remove_question_mark(sheet2[CMOBILE  + str(row)].value)
        sheet2[CEMAIL         + str(row)].value = remove_question_mark(sheet2[CEMAIL   + str(row)].value)
    for row in range(start, end):
        sheet2[CCOMPANY_CHINA + str(row + begin)].value = format_company(sheet1[ICOMPANY + str(row)].value)
        sheet2[CSOURCE        + str(row + begin)].value = sheet1[ISOURCE  + str(row)].value
        sheet2[CTYPE          + str(row + begin)].value = sheet1[ITYPE    + str(row)].value
        sheet2[CDATE          + str(row + begin)].value = sheet1[IDATE    + str(row)].value

    wb.save("mergedFile.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# strip_leading_numbers()
combine_sheets()
# split_companies()

# Open an existing excel file
# wb = openpyxl.load_workbook(fileName)
# sheet = wb['Globalore List']
