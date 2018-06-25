import openpyxl
import os
import pygame
import string
import sys
import time

printing = True
saving = True

COMPANY      = "A"
COMMODITIES  = "B"
ROLE         = "C"
LAST_UPDATED = "D"

min_word_len = 5

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        word = str(word)
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# replace_punct(word)
#{{{
def replace_punct(word):
    return word.replace('.', '').replace(',', '').replace('(', ' (').replace(')', ') ').replace('  ', ' ').strip()
#}}}

# company object with all relevant fields
# Company
#{{{
class Company(object):
    company            = ''
    commodity          = ''
    role               = ''
    last_updated       = ''

    def __init__(self):
        company      = "-"
#}}}

# Create a new company from sheet information
# new_company_from_sheet(sheet, row)
#{{{
def new_company_from_sheet(sheet, row):
    company = Company()

    if sheet[COMPANY          + str(row)].value != None:
        company.company       = sheet[COMPANY               + str(row)].value.title()
    if sheet[COMMODITIES      + str(row)].value != None:
        company.commodity     = sheet[COMMODITIES           + str(row)].value
    if sheet[ROLE             + str(row)].value != None:
        company.role          = sheet[ROLE                  + str(row)].value
    if sheet[LAST_UPDATED     + str(row)].value != None:
        company.last_updated  = sheet[LAST_UPDATED          + str(row)].value

    return company
#}}}

# Edit Distance Function
# edit_distance(word1, word2, low_threshold, high_threshold)
#{{{
# modified edit distance algorithm:
# - see if string one is a substring of the next string
#   - if it is there is a match
#   - it not look at edit distance for substring and substring of equal length
#     - if it is above a certain percentage or below a certain edit number we can safely assume they are a match. keep that string and try with next one

def edit_distance(word1, word2, low_threshold, high_threshold):
    if printing:
        print("Looking at '" + word1 + "' and '" + word2 + "'")
    word1 = word1.lower()
    word2 = word2.lower()
    len_1 = len(word1)
    len_2 = len(word2)
    edit_distance, percent_match = 0, 0

    # make sure shorter word is first
    if len_1 > len_2:
        word1, word2 = word2, word1
        len_1 = len_2

    if (len_1 >= min_word_len):
        if (word1 in word2):
            if printing:
                print ("Edit distance: 0")
                print ("Percent Match: 100")
            edit_distance, percent_match = 0, 100
        else:
            # shorten longer word to length of first word
            word2 = word2[0:len_1]
            # the matrix whose last element -> edit distance
            x = [[0] * (len_1 + 1) for _ in range(len_1 + 1)]

            # initialization of base case values
            for i in range(0, len_1 + 1): 
                x[i][0] = i
            for j in range(0, len_1 + 1):
                x[0][j] = j
            for i in range (1, len_1 + 1):
                for j in range(1, len_1 + 1):
                    if word1[i - 1] == word2[j - 1]:
                        x[i][j] = x[i - 1][j - 1] 
                    else:
                        x[i][j]= min(x[i][j - 1], x[i - 1][j], x[i - 1][j - 1]) + 1
            edit_distance = x[i][j]
            percent_match = ((len_1 - edit_distance) / len_1) * 100
            if printing:
                print ("Edit distance " + str(x[i][j]))
                print ("Percent match: " + "%.2f" % percent_match)
    if percent_match > low_threshold and percent_match <= high_threshold:
        if printing:
            print("MATCH!")
        return True
    else:
        return False
#}}}

def process_iron(*args):
    # turn the arguments into variable companys
    args = args[0]
    filecompany = args[1]
    low_threshold = int(args[2])
    cols = args[3:]
    if printing:
        print("Opening...")
    # Open the file for editing
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("companies")
    wb = openpyxl.load_workbook(filecompany)
    sheet = wb.worksheets[0]

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    # Create a new file to store duplicate companys
    dupe = openpyxl.Workbook()
    # Open the worksheet we want to edit
    dupesheet = dupe.create_sheet("companies")
    # if 'sheet' appears randomly we can delete it
    rm = dupe.get_sheet_by_name('Sheet')
    dupe.remove_sheet(rm)

    # - create an object for a new primary company and account company pair
    #   - store previous object in a new sheet
    #   - store all information here
    #   - look at next company and see if it matches (edit distance)
    #     - if it is a match combine and keep going, otherwise repeat
    compare = ""
    current = ""
    last = sheet.max_row
    count = 1
    dupes = 2
    first = 2
    high_threshold = 100
    company = Company()

    # Create Headers
    #{{{
    if saving:
        outsheet[COMPANY       + '1'].value = "Company"
        outsheet[COMMODITIES   + '1'].value = "Commodity"
        outsheet[ROLE          + '1'].value = "Role"
        outsheet[LAST_UPDATED  + '1'].value = "LastUpdated"

        dupesheet[COMPANY      + '1'].value = "Company"
        dupesheet[COMMODITIES  + '1'].value = "Commodity"
        dupesheet[ROLE         + '1'].value = "Role"
        dupesheet[LAST_UPDATED + '1'].value = "LastUpdated"
    #}}}

    for row in range (first, last + 1):
        # if the previous value is blank we create the new object and store information in it
        '''
        first_company = str(sheet[first_company_col + str(row)].value)
        if first_company != "":
            standardize_str(first_company)
        '''
        compareCriteria = ""
        for col in cols:
            print(col, row)
            compareCriteria = compareCriteria + standardize_str(sheet[col + str(row)].value) + " "
        if row == first:
            # compare = first_company + " " + last_company
            compare = compareCriteria
            company = new_company_from_sheet(sheet, row)
            count = count + 1
        else:
            sheet[COMPANY + str(row)].value = replace_punct(sheet[col + str(row)].value)
            # current = first_company + " " + last_company
            current = compareCriteria
            match = edit_distance(compare, current, low_threshold, high_threshold)
            # matchingSuffixes = standardize_str(sheet[SUFFIX + str(row)].value) == standardize_str(sheet[SUFFIX + str(row - 1)].value)
            # combine information and move on
            if match:
                company = new_company_from_sheet(sheet, row)
                '''
                - if there is a duplicate
                - store previous item in list in duplicate list
                - store the duplicate in next spot in duplicate list
                '''
                # combine information
                # keep original
                dupes = dupes + 1

                dupeCompany = new_company_from_sheet(sheet, row - 1)

                #{{{
                dupesheet[COMPANY + str(dupes)].value            = dupeCompany.company 
                dupesheet[COMMODITIES + str(dupes)].value        = dupeCompany.commodity 
                dupesheet[ROLE + str(dupes)].value               = dupeCompany.role 
                dupesheet[LAST_UPDATED + str(dupes)].value       = dupeCompany.last_updated 
                dupesheet['E' + str(dupes)].value = row - 1
                #}}}

                # keep duplicate
                dupes = dupes + 1

                dupeCompany = new_company_from_sheet(sheet, row)
                #{{{
                dupesheet[COMPANY + str(dupes)].value            = dupeCompany.company 
                dupesheet[COMMODITIES + str(dupes)].value        = dupeCompany.commodity 
                dupesheet[ROLE + str(dupes)].value               = dupeCompany.role 
                dupesheet[LAST_UPDATED + str(dupes)].value       = dupeCompany.last_updated 
                dupesheet['E' + str(dupes)].value = row
                #}}}

                dupes = dupes + 1

                # save the combined company
                #{{{
                dupesheet[COMPANY + str(dupes)].value            = company.company 
                dupesheet[COMMODITIES + str(dupes)].value        = company.commodity 
                dupesheet[ROLE + str(dupes)].value               = company.role 
                dupesheet[LAST_UPDATED + str(dupes)].value       = company.last_updated 
                #}}}

                # create a blank space
                dupes = dupes + 1
            # store the information and create a new company
            else:
                # store information
                #{{{
                outsheet[COMPANY + str(count)].value            = company.company 
                outsheet[COMMODITIES + str(count)].value        = company.commodity 
                outsheet[ROLE + str(count)].value               = company.role 
                outsheet[LAST_UPDATED + str(count)].value       = company.last_updated 
                #}}}
                # reset compare value
                compare = compareCriteria
                # create a new company
                company = new_company_from_sheet(sheet, row)
                count = count + 1

    if printing:
        print()
        print("Out of " + str(1 + last - first) + " companies " + str(count) + " were unique companies")
        print("Saving...")

    out.save("purged" + str(cols) + ".xlsx")
    dupe.save("duplicates" + str(cols) + ".xlsx")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

process_iron(sys.argv)
