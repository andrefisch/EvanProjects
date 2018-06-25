from datetime import datetime
from splitNames import determine_names
import openpyxl
import pygame
import re
import string
import sys
import time

#############
# VARIABLES #
#############

#{{{
saving = True
printing = True
testing = True
low_threshold = 85
high_threshold = 100
min_word_len = 5

'''
FIRST_NAME           = 'A'
MIDDLE_NAME          = 'B'
LAST_NAME            = 'C'
TITLE                = 'D'
SUFFIX               = 'E'
WEB_PAGE             = 'F'
NOTES                = 'G'
EMAIL_ADDRESS        = 'H'
EMAIL_ADDRESS2       = 'I'
EMAIL_ADDRESS3       = 'J'
HOME_PHONE           = 'K'
MOBILE_PHONE         = 'L'
HOME_ADDRESS         = 'M'
HOME_STREET          = 'N'
HOME_CITY            = 'O'
HOME_STATE           = 'P'
HOME_POSTAL_CODE     = 'Q'
HOME_COUNTRY         = 'R'
CONTACT_MAIN_PHONE   = 'S'
BUSINESS_PHONE       = 'T'
BUSINESS_PHONE2      = 'U'
BUSINESS_FAX         = 'V'
COMPANY              = 'W'
JOB_TITLE            = 'X'
DEPARTMENT           = 'Y'
OFFICE_LOCATION      = 'Z'
BUSINESS_ADDRESS     = 'AA'
BUSINESS_STREET      = 'AB'
BUSINESS_CITY        = 'AC'
BUSINESS_STATE       = 'AD'
BUSINESS_POSTAL_CODE = 'AE'
BUSINESS_COUNTRY     = 'AF'
CATEGORIES           = 'AG'
CONNECTED_ON         = 'AH'
LATITUDE             = 'AI'
LONGITUDE            = 'AJ'
'''

NAME                          = "A"
GIVEN_NAME                    = "B"
ADDITIONAL_NAME               = "C"
FAMILY_NAME                   = "D"
YOMI_NAME                     = "E"
GIVEN_NAME_YOMI               = "F"
ADDITIONAL_NAME_YOMI          = "G"
FAMILY_NAME_YOMI              = "H"
NAME_PREFIX                   = "I"
NAME_SUFFIX                   = "J"
INITIALS                      = "K"
NICKNAME                      = "L"
SHORT_NAME                    = "M"
MAIDEN_NAME                   = "N"
BIRTHDAY                      = "O"
GENDER                        = "P"
LOCATION                      = "Q"
BILLING_INFORMATION           = "R"
DIRECTORY_SERVER              = "S"
MILEAGE                       = "T"
OCCUPATION                    = "U"
HOBBY                         = "V"
SENSITIVITY                   = "W"
PRIORITY                      = "X"
SUBJECT                       = "Y"
NOTES                         = "Z"
GROUP_MEMBERSHIP              = "AA"
EMAIL1_TYPE                   = "AB"
EMAIL1_VALUE                  = "AC"
EMAIL2_TYPE                   = "AD"
EMAIL2_VALUE                  = "AE"
EMAIL3_TYPE                   = "AF"
EMAIL3_VALUE                  = "AG"
EMAIL4_TYPE                   = "AH"
EMAIL4_VALUE                  = "AI"
EMAIL5_TYPE                   = "AJ"
EMAIL5_VALUE                  = "AK"
IM1_TYPE                      = "AL"
IM1_SERVICE                   = "AM"
IM1_VALUE                     = "AN"
PHONE1_TYPE                   = "AO"
PHONE1_VALUE                  = "AP"
PHONE2_TYPE                   = "AQ"
PHONE2_VALUE                  = "AR"
PHONE3_TYPE                   = "AS"
PHONE3_VALUE                  = "AT"
PHONE4_TYPE                   = "AU"
PHONE4_VALUE                  = "AV"
PHONE5_TYPE                   = "AW"
PHONE5_VALUE                  = "AX"
ADDRESS1_TYPE                 = "AY"
ADDRESS1_FORMATED             = "AZ"
ADDRESS1_STREET               = "BA"
ADDRESS1_CITY                 = "BB"
ADDRESS1_POBOX                = "BC"
ADDRESS1_REGION               = "BD"
ADDRESS1_POSTAL_CODE          = "BE"
ADDRESS1_COUNTRY              = "BF"
ADDRESS1_EXTENDED_ADDRESS     = "BG"
ADDRESS2_TYPE                 = "BH"
ADDRESS2_FORMATED             = "BI"
ADDRESS2_STREET               = "BJ"
ADDRESS2_CITY                 = "BK"
ADDRESS2_POBOX                = "BL"
ADDRESS2_REGION               = "BM"
ADDRESS2_POSTAL_CODE          = "BN"
ADDRESS2_COUNTRY              = "BO"
ADDRESS2_EXTENDED_ADDRESS     = "BP"
ADDRESS3_TYPE                 = "BQ"
ADDRESS3_FORMATED             = "BR"
ADDRESS3_STREET               = "BS"
ADDRESS3_CITY                 = "BT"
ADDRESS3_POBOX                = "BU"
ADDRESS3_REGION               = "BV"
ADDRESS3_POSTAL_CODE          = "BW"
ADDRESS3_COUNTRY              = "BX"
ADDRESS3_EXTENDED_ADDRESS     = "BY"
ORGANIZATION1_TYPE            = "BZ"
ORGANIZATION1_NAME            = "CA"
ORGANIZATION1_YOMI_NAME       = "CB"
ORGANIZATION1_TITLE           = "CC"
ORGANIZATION1_DEPARTMENT      = "CD"
ORGANIZATION1_SYMBOL          = "CE"
ORGANIZATION1_LOCATION        = "CF"
ORGANIZATION1_JOB_DESCRIPTION = "CG"
RELATION1_TYPE                = "CH"
RELATION1_VALUE               = "CI"
EXTERNAL_ID1_TYPE             = "CJ"
EXTERNAl_ID1_VALUE            = "CK"
WEBSITE1_TYPE                 = "CL"
WEBSITE1_VALUE                = "CM"
CALENDAR_LINK1_TYPE           = "CN"
CALENDAR_LINK1_VALUE          = "CO"
JOT1_TYPE                     = "CP"
JOT1_VALUE                    = "CQ"
#}}}


##################
# HELPER METHODS #
##################

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        word = str(word)
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# Combine with slash
# special_combine(word1, word2)
# {{{
def special_combine(word1, word2):
    if word1 == "" or word1 == None:
        return word2
    elif word2 == "" or word2 == None:
        return word1
    elif word1 in word2:
        return word2
    elif word2 in word1:
        return word1
    else:
        return word1 + "/" + word2
#}}}

# Add to notes
# add_to_notes(contact, category, inf, oldInfo)
#{{{
def add_to_notes(contact, category, info, oldInfo):
    if info != "" and info != None and oldInfo != "" and oldInfo != None and info != oldInfo:
        return str(contact.notes) + "; " + str(category) + " used to be " + str(info)
    else:
        return str(contact.notes)
#}}}

# special_save_email(contact, oldEmail, row)
#{{{
def special_save_email(sheet, contact, oldEmail, row):
    if sheet[EMAIL1_VALUE + str(row)].value == None:
        sheet[EMAIL1_VALUE + str(row)].value = oldEmail
    elif sheet[EMAIL2_VALUE + str(row)].value == None:
        sheet[EMAIL2_VALUE + str(row)].value = oldEmail
    elif sheet[EMAIL3_VALUE + str(row)].value == None:
        sheet[EMAIL3_VALUE + str(row)].value = oldEmail
    elif sheet[EMAIL4_VALUE + str(row)].value == None:
        sheet[EMAIL4_VALUE + str(row)].value = oldEmail
    elif sheet[EMAIL5_VALUE + str(row)].value == None:
        sheet[EMAIL5_VALUE + str(row)].value = oldEmail
#}}}

# Keep the non-none value
# keep_non_none(var1, var2)
#{{{
def keep_non_none(var1, var2):
    if var1 == None or var1 == "":
        return (var2, var1)
    elif var2 == None or var2 == "":
        return (var1, var2)
    else:
        return (var2, var1)
#}}}

# Edit Distance Function
# edit_distance(word1, word2, low_threshold, high_threshold)
#{{{
# modified edit distance algorithm:
# - see if string one is a substring of the next string
#   - if it is there is a match
#   - it not look at edit distance for substring and substring of equal length
#     - if it is above a certain percentage or below a certain edit number we can safely assume they are a match. keep that string and try with next one

def edit_distance(word1, word2, low_threshold, high_threshold):
    if printing:
        print("Looking at '" + word1 + "' and '" + word2 + "'")
    word1 = word1.lower()
    word2 = word2.lower()
    len_1 = len(word1)
    len_2 = len(word2)
    edit_distance, percent_match = 0, 0

    # make sure shorter word is first
    if len_1 > len_2:
        word1, word2 = word2, word1
        len_1 = len_2

    if (len_1 >= min_word_len):
        if (word1 in word2):
            if printing:
                print ("Edit distance: 0")
                print ("Percent Match: 100")
            edit_distance, percent_match = 0, 100
        else:
            # shorten longer word to length of first word
            word2 = word2[0:len_1]
            # the matrix whose last element -> edit distance
            x = [[0] * (len_1 + 1) for _ in range(len_1 + 1)]

            # initialization of base case values
            for i in range(0, len_1 + 1): 
                x[i][0] = i
            for j in range(0, len_1 + 1):
                x[0][j] = j
            for i in range (1, len_1 + 1):
                for j in range(1, len_1 + 1):
                    if word1[i - 1] == word2[j - 1]:
                        x[i][j] = x[i - 1][j - 1] 
                    else:
                        x[i][j]= min(x[i][j - 1], x[i - 1][j], x[i - 1][j - 1]) + 1
            edit_distance = x[i][j]
            percent_match = ((len_1 - edit_distance) / len_1) * 100
            if printing:
                print ("Edit distance " + str(x[i][j]))
                print ("Percent match: " + "%.2f" % percent_match)
    if percent_match > low_threshold and percent_match <= high_threshold:
        if printing:
            print("MATCH!")
        return True
    else:
        return False
#}}}

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# Remove all spaces from the last n nonspace characters of the given string
# remove_end_space(string, num)
#{{{
def remove_end_space(string, num):
    string2 = string[::-1]
    count = 0
    index = 0
    for i in range(0, len(string2)):
        if string2[i] != " ":
            count += 1
        if count >= num:
            index = i
            break
    first = string2[index:]
    second = string2[:index]
    second = second.replace(" ", "")
    return (second + first)[::-1]
#}}}

# Format a single phone number
# formatting_phone_number(number)
#{{{
def formatting_phone_number(number):
    original = number
    # replace certain punctuations with a space
    regexPunct = '([\.:;|/@]| {2,})'
    number = re.sub(regexPunct, " ", number)
    # replace all +'s with nothing
    number = re.sub('[\+\[\]\'=]', "", number)
    # remove all letters
    regexLetters = '[a-zA-Z]'
    number = re.sub(regexLetters, "", number)
    # remove all spaces before the first number
    regexBegin = '^ *'
    number = re.sub(regexBegin, "", number)
    # remove everything after the last number
    regexEnd = '\D*$'
    number = re.sub(regexEnd, "", number)
    # 077, 078 -> +4477 or +4478
    regex077 = '^\D*077(.*)'
    regex078 = '^\D*078(.*)'
    match077 = re.search(regex077, number)
    match078 = re.search(regex078, number)
    # remove leading 001 and 011
    regex01 = '^(001 ?-?|011 ?-?)'
    number = re.sub(regex01, "", number)
    # remove all leading 0's
    regex0 = '^0+(.*)'
    match0 = re.search(regex0, number)

    # if the number starts with 00 replace 00 with +
    # WORKS
    if match0:
        number = "+" + match0.group(1)
    # if the number starts with 077 replace 077 with +4477
    # WORKS
    if match077:
        number = "+4477" + match077.group(1)
    # if the number starts with 078 replace 077 with +4478
    # WORKS
    if match078:
        number = "+4478" + match078.group(1)
    # if number contains a useless (0), remove it
    regexParen = ' ?\([+ 0]?\) ?'
    number = re.sub(regexParen, '', number)
    # if number contains a weird (), remove it
    regex1      = '^\D*1'
    regexUSA    = '^\D*(\+?1?)?\D*(\d{3})\D*(\d{3})\D*(\d{4})$'
    regexUSAno1 = '^(\(\d{3}\))\D*(\d{3})\D*(\d{4})'
    # regexEurope = '\D*(\+?\d{2,3})\D*'
    matchUSA    = re.search(regexUSA,    number)
    matchUSAno1 = re.search(regexUSAno1, number)
    # if the phone number is a USA number with a 1 in front
    if matchUSA:
        number = "+1 (" + matchUSA.group(2) + ") " + matchUSA.group(3) + "-" + matchUSA.group(4)
    elif matchUSAno1:
    # if the phone number is a USA number without a 1 in front
        number = "+1 " + matchUSAno1.group(1) + " " + matchUSAno1.group(2) + "-" + matchUSAno1.group(3)
    # if the number is a european number
    elif len(number) > 0 and number[0] != '+':
        regexParens = '[\(\)]'
        number = re.sub(regexParens, "", number)
        regexTopCountries = '^(30|31|33|41|44|49|55|60|61|65|82|86|90|380|852|886|966|971) ?(.*)'
        matchTopCountries = re.search(regexTopCountries, number)
        if matchTopCountries:
            number = matchTopCountries.group(1) + " " + matchTopCountries.group(2)
        # remove all leading 0's again
        regex0 = '^0+(.*)'
        match0 = re.search(regex0, number)
        regexDash = '-'
        number = re.sub(regexDash, " ", number)
        regexSpaces = ' {2,}'
        number = re.sub(regexSpaces, " ", number)
        number = "+" + number


    # POSTPROCESSING
    # remove the 0 from 440
    regex440 = '\+440'
    number = re.sub(regex440, "+44", number)
    regex0xx = '(\+[2-9][^ ].*\()0(.*)'
    match0xx = re.search(regex0xx, number)
    if match0xx:
        number = match0xx.group(1) + match0xx.group(2)
    # remove spaces from last 4 non-space characters
    number = remove_end_space(number, 4)
    # print out the original number and its formatted version
    if printing:
        if original != "None":
            print(original + " -> " + number)

    return number
#}}}

# Contact object with all relevant fields
# Contact
#{{{
class Contact(object):
    name                          = ''
    given_name                    = ''
    additional_name               = ''
    family_name                   = ''
    yomi_name                     = ''
    given_name_yomi               = ''
    additional_name_yomi          = ''
    family_name_yomi              = ''
    name_prefix                   = ''
    name_suffix                   = ''
    initials                      = ''
    nickname                      = ''
    short_name                    = ''
    maiden_name                   = ''
    birthday                      = ''
    gender                        = ''
    location                      = ''
    billing_information           = ''
    directory_server              = ''
    mileage                       = ''
    occupation                    = ''
    hobby                         = ''
    sensitivity                   = ''
    priority                      = ''
    subject                       = ''
    notes                         = ''
    group_membership              = ''
    email1_type                   = ''
    email1_value                  = ''
    email2_type                   = ''
    email2_value                  = ''
    email3_type                   = ''
    email3_value                  = ''
    email4_type                   = ''
    email4_value                  = ''
    email5_type                   = ''
    email5_value                  = ''
    im1_type                      = ''
    im1_service                   = ''
    im1_value                     = ''
    phone1_type                   = ''
    phone1_value                  = ''
    phone2_type                   = ''
    phone2_value                  = ''
    phone3_type                   = ''
    phone3_value                  = ''
    phone4_type                   = ''
    phone4_value                  = ''
    phone5_type                   = ''
    phone5_value                  = ''
    address1_type                 = ''
    address1_formated             = ''
    address1_street               = ''
    address1_city                 = ''
    address1_pobox                = ''
    address1_region               = ''
    address1_postal_code          = ''
    address1_country              = ''
    address1_extended_address     = ''
    address2_type                 = ''
    address2_formated             = ''
    address2_street               = ''
    address2_city                 = ''
    address2_pobox                = ''
    address2_region               = ''
    address2_postal_code          = ''
    address2_country              = ''
    address2_extended_address     = ''
    address3_type                 = ''
    address3_formated             = ''
    address3_street               = ''
    address3_city                 = ''
    address3_pobox                = ''
    address3_region               = ''
    address3_postal_code          = ''
    address3_country              = ''
    address3_extended_address     = ''
    organization1_type            = ''
    organization1_name            = ''
    organization1_yomi_name       = ''
    organization1_title           = ''
    organization1_department      = ''
    organization1_symbol          = ''
    organization1_location        = ''
    organization1_job_description = ''
    relation1_type                = ''
    relation1_value               = ''
    external_id1_type             = ''
    external_id1_value            = ''
    website1_type                 = ''
    website1_value                = ''
    calendar_link1_type           = ''
    calendar_link1_value          = ''
    jot1_type                     = ''
    jot1_value                    = ''

    def __init__(self):
        name      = "-"
        notes     = ""

#}}}

# Create a new contact from sheet information
# new_contact_from_sheet(sheet, row)
#{{{
def new_contact_from_sheet(sheet, row):
    contact = Contact()

    if sheet[NAME                             + str(row)].value != None:
        contact.name                          = sheet[NAME                          + str(row)].value.title()
    if sheet[GIVEN_NAME                       + str(row)].value != None:
        contact.given_name                    = sheet[GIVEN_NAME                    + str(row)].value.title()
    if sheet[ADDITIONAL_NAME                  + str(row)].value != None:
        contact.additional_name               = sheet[ADDITIONAL_NAME               + str(row)].value.title()
    if sheet[FAMILY_NAME                      + str(row)].value != None:
        contact.family_name                   = sheet[FAMILY_NAME                   + str(row)].value.title()
    if sheet[YOMI_NAME                        + str(row)].value != None:
        contact.yomi_name                     = sheet[YOMI_NAME                     + str(row)].value.title()
    if sheet[GIVEN_NAME_YOMI                  + str(row)].value != None:
        contact.given_name_yomi               = sheet[GIVEN_NAME_YOMI               + str(row)].value.title()
    if sheet[ADDITIONAL_NAME_YOMI             + str(row)].value != None:
        contact.additional_name_yomi          = sheet[ADDITIONAL_NAME_YOMI          + str(row)].value.title()
    if sheet[FAMILY_NAME_YOMI                 + str(row)].value != None:
        contact.family_name_yomi              = sheet[FAMILY_NAME_YOMI              + str(row)].value.title()
    if sheet[NAME_PREFIX                      + str(row)].value != None:
        contact.name_prefix                   = sheet[NAME_PREFIX                   + str(row)].value.title()
    if sheet[NAME_SUFFIX                      + str(row)].value != None:
        contact.name_suffix                   = sheet[NAME_SUFFIX                   + str(row)].value.title()
    if sheet[INITIALS                         + str(row)].value != None:
        contact.initials                      = sheet[INITIALS                      + str(row)].value.title()
    if sheet[NICKNAME                         + str(row)].value != None:
        contact.nickname                      = sheet[NICKNAME                      + str(row)].value.title()
    if sheet[SHORT_NAME                       + str(row)].value != None:
        contact.short_name                    = sheet[SHORT_NAME                    + str(row)].value.title()
    if sheet[MAIDEN_NAME                      + str(row)].value != None:
        contact.maiden_name                   = sheet[MAIDEN_NAME                   + str(row)].value.title()
    if sheet[BIRTHDAY                         + str(row)].value != None:
        contact.birthday                      = sheet[BIRTHDAY                      + str(row)].value
    if sheet[GENDER                           + str(row)].value != None:
        contact.gender                        = sheet[GENDER                        + str(row)].value
    if sheet[LOCATION                         + str(row)].value != None:
        contact.location                      = sheet[LOCATION                      + str(row)].value
    if sheet[BILLING_INFORMATION              + str(row)].value != None:
        contact.billing_information           = sheet[BILLING_INFORMATION           + str(row)].value
    if sheet[DIRECTORY_SERVER                 + str(row)].value != None:
        contact.directory_server              = sheet[DIRECTORY_SERVER              + str(row)].value
    if sheet[MILEAGE                          + str(row)].value != None:
        contact.mileage                       = sheet[MILEAGE                       + str(row)].value
    if sheet[OCCUPATION                       + str(row)].value != None:
        contact.occupation                    = sheet[OCCUPATION                    + str(row)].value
    if sheet[HOBBY                            + str(row)].value != None:
        contact.hobby                         = sheet[HOBBY                         + str(row)].value
    if sheet[SENSITIVITY                      + str(row)].value != None:
        contact.sensitivity                   = sheet[SENSITIVITY                   + str(row)].value
    if sheet[PRIORITY                         + str(row)].value != None:
        contact.priority                      = sheet[PRIORITY                      + str(row)].value
    if sheet[SUBJECT                          + str(row)].value != None:
        contact.subject                       = sheet[SUBJECT                       + str(row)].value
    if sheet[NOTES                            + str(row)].value != None:
        contact.notes                         = sheet[NOTES                         + str(row)].value
    if sheet[GROUP_MEMBERSHIP                 + str(row)].value != None:
        contact.group_membership              = sheet[GROUP_MEMBERSHIP              + str(row)].value
    if sheet[EMAIL1_TYPE                      + str(row)].value != None:
        contact.email1_type                   = sheet[EMAIL1_TYPE                   + str(row)].value
    if sheet[EMAIL1_VALUE                     + str(row)].value != None:
        contact.email1_value                  = sheet[EMAIL1_VALUE                  + str(row)].value
    if sheet[EMAIL2_TYPE                      + str(row)].value != None:
        contact.email2_type                   = sheet[EMAIL2_TYPE                   + str(row)].value
    if sheet[EMAIL2_VALUE                     + str(row)].value != None:
        contact.email2_value                  = sheet[EMAIL2_VALUE                  + str(row)].value
    if sheet[EMAIL3_TYPE                      + str(row)].value != None:
        contact.email3_type                   = sheet[EMAIL3_TYPE                   + str(row)].value
    if sheet[EMAIL3_VALUE                     + str(row)].value != None:
        contact.email3_value                  = sheet[EMAIL3_VALUE                  + str(row)].value
    if sheet[EMAIL4_TYPE                      + str(row)].value != None:
        contact.email4_type                   = sheet[EMAIL4_TYPE                   + str(row)].value
    if sheet[EMAIL4_VALUE                     + str(row)].value != None:
        contact.email4_value                  = sheet[EMAIL4_VALUE                  + str(row)].value
    if sheet[EMAIL5_TYPE                      + str(row)].value != None:
        contact.email5_type                   = sheet[EMAIL5_TYPE                   + str(row)].value
    if sheet[EMAIL5_VALUE                     + str(row)].value != None:
        contact.email5_value                  = sheet[EMAIL5_VALUE                  + str(row)].value
    if sheet[IM1_TYPE                         + str(row)].value != None:
        contact.im1_type                      = sheet[IM1_TYPE                      + str(row)].value
    if sheet[IM1_SERVICE                      + str(row)].value != None:
        contact.im1_service                   = sheet[IM1_SERVICE                   + str(row)].value
    if sheet[IM1_VALUE                        + str(row)].value != None:
        contact.im1_value                     = sheet[IM1_VALUE                     + str(row)].value
    if sheet[PHONE1_TYPE                      + str(row)].value != None:
        contact.phone1_type                   = sheet[PHONE1_TYPE                   + str(row)].value
    if sheet[PHONE1_VALUE                     + str(row)].value != None:
        contact.phone1_value                  = sheet[PHONE1_VALUE                  + str(row)].value
    if sheet[PHONE2_TYPE                      + str(row)].value != None:
        contact.phone2_type                   = sheet[PHONE2_TYPE                   + str(row)].value
    if sheet[PHONE2_VALUE                     + str(row)].value != None:
        contact.phone2_value                  = sheet[PHONE2_VALUE                  + str(row)].value
    if sheet[PHONE3_TYPE                      + str(row)].value != None:
        contact.phone3_type                   = sheet[PHONE3_TYPE                   + str(row)].value
    if sheet[PHONE3_VALUE                     + str(row)].value != None:
        contact.phone3_value                  = sheet[PHONE3_VALUE                  + str(row)].value
    if sheet[PHONE4_TYPE                      + str(row)].value != None:
        contact.phone4_type                   = sheet[PHONE4_TYPE                   + str(row)].value
    if sheet[PHONE4_VALUE                     + str(row)].value != None:
        contact.phone4_value                  = sheet[PHONE4_VALUE                  + str(row)].value
    if sheet[PHONE5_TYPE                      + str(row)].value != None:
        contact.phone5_type                   = sheet[PHONE5_TYPE                   + str(row)].value
    if sheet[PHONE5_VALUE                     + str(row)].value != None:
        contact.phone5_value                  = sheet[PHONE5_VALUE                  + str(row)].value
    if sheet[ADDRESS1_TYPE                    + str(row)].value != None:
        contact.address1_type                 = sheet[ADDRESS1_TYPE                 + str(row)].value.title()
    if sheet[ADDRESS1_FORMATED                + str(row)].value != None:
        contact.address1_formated             = sheet[ADDRESS1_FORMATED             + str(row)].value.title()
    if sheet[ADDRESS1_STREET                  + str(row)].value != None:
        contact.address1_street               = sheet[ADDRESS1_STREET               + str(row)].value.title()
    if sheet[ADDRESS1_CITY                    + str(row)].value != None:
        contact.address1_city                 = sheet[ADDRESS1_CITY                 + str(row)].value.title()
    if sheet[ADDRESS1_POBOX                   + str(row)].value != None:
        contact.address1_pobox                = sheet[ADDRESS1_POBOX                + str(row)].value
    if sheet[ADDRESS1_REGION                  + str(row)].value != None:
        contact.address1_region               = sheet[ADDRESS1_REGION               + str(row)].value.title()
    if sheet[ADDRESS1_POSTAL_CODE             + str(row)].value != None:
        contact.address1_postal_code          = sheet[ADDRESS1_POSTAL_CODE          + str(row)].value
    if sheet[ADDRESS1_COUNTRY                 + str(row)].value != None:
        contact.address1_country              = sheet[ADDRESS1_COUNTRY              + str(row)].value.title()
    if sheet[ADDRESS1_EXTENDED_ADDRESS        + str(row)].value != None:
        contact.address1_extended_address     = sheet[ADDRESS1_EXTENDED_ADDRESS     + str(row)].value.title()
    if sheet[ADDRESS2_TYPE                    + str(row)].value != None:
        contact.address2_type                 = sheet[ADDRESS2_TYPE                 + str(row)].value.title()
    if sheet[ADDRESS2_FORMATED                + str(row)].value != None:
        contact.address2_formated             = sheet[ADDRESS2_FORMATED             + str(row)].value.title()
    if sheet[ADDRESS2_STREET                  + str(row)].value != None:
        contact.address2_street               = sheet[ADDRESS2_STREET               + str(row)].value.title()
    if sheet[ADDRESS2_CITY                    + str(row)].value != None:
        contact.address2_city                 = sheet[ADDRESS2_CITY                 + str(row)].value.title()
    if sheet[ADDRESS2_POBOX                   + str(row)].value != None:
        contact.address2_pobox                = sheet[ADDRESS2_POBOX                + str(row)].value
    if sheet[ADDRESS2_REGION                  + str(row)].value != None:
        contact.address2_region               = sheet[ADDRESS2_REGION               + str(row)].value.title()
    if sheet[ADDRESS2_POSTAL_CODE             + str(row)].value != None:
        contact.address2_postal_code          = sheet[ADDRESS2_POSTAL_CODE          + str(row)].value
    if sheet[ADDRESS2_COUNTRY                 + str(row)].value != None:
        contact.address2_country              = sheet[ADDRESS2_COUNTRY              + str(row)].value.title()
    if sheet[ADDRESS2_EXTENDED_ADDRESS        + str(row)].value != None:
        contact.address2_extended_address     = sheet[ADDRESS2_EXTENDED_ADDRESS     + str(row)].value.title()
    if sheet[ADDRESS3_TYPE                    + str(row)].value != None:
        contact.address3_type                 = sheet[ADDRESS3_TYPE                 + str(row)].value.title()
    if sheet[ADDRESS3_FORMATED                + str(row)].value != None:
        contact.address3_formated             = sheet[ADDRESS3_FORMATED             + str(row)].value.title()
    if sheet[ADDRESS3_STREET                  + str(row)].value != None:
        contact.address3_street               = sheet[ADDRESS3_STREET               + str(row)].value.title()
    if sheet[ADDRESS3_CITY                    + str(row)].value != None:
        contact.address3_city                 = sheet[ADDRESS3_CITY                 + str(row)].value.title()
    if sheet[ADDRESS3_POBOX                   + str(row)].value != None:
        contact.address3_pobox                = sheet[ADDRESS3_POBOX                + str(row)].value
    if sheet[ADDRESS3_REGION                  + str(row)].value != None:
        contact.address3_region               = sheet[ADDRESS3_REGION               + str(row)].value.title()
    if sheet[ADDRESS3_POSTAL_CODE             + str(row)].value != None:
        contact.address3_postal_code          = sheet[ADDRESS3_POSTAL_CODE          + str(row)].value
    if sheet[ADDRESS3_COUNTRY                 + str(row)].value != None:
        contact.address3_country              = sheet[ADDRESS3_COUNTRY              + str(row)].value.title()
    if sheet[ADDRESS3_EXTENDED_ADDRESS        + str(row)].value != None:
        contact.address3_extended_address     = sheet[ADDRESS3_EXTENDED_ADDRESS     + str(row)].value.title()
    if sheet[ORGANIZATION1_TYPE               + str(row)].value != None:
        contact.organization1_type            = sheet[ORGANIZATION1_TYPE            + str(row)].value
    if sheet[ORGANIZATION1_NAME               + str(row)].value != None:
        contact.organization1_name            = sheet[ORGANIZATION1_NAME            + str(row)].value
    if sheet[ORGANIZATION1_YOMI_NAME          + str(row)].value != None:
        contact.organization1_yomi_name       = sheet[ORGANIZATION1_YOMI_NAME       + str(row)].value
    if sheet[ORGANIZATION1_TITLE              + str(row)].value != None:
        contact.organization1_title           = sheet[ORGANIZATION1_TITLE           + str(row)].value
    if sheet[ORGANIZATION1_DEPARTMENT         + str(row)].value != None:
        contact.organization1_department      = sheet[ORGANIZATION1_DEPARTMENT      + str(row)].value
    if sheet[ORGANIZATION1_SYMBOL             + str(row)].value != None:
        contact.organization1_symbol          = sheet[ORGANIZATION1_SYMBOL          + str(row)].value
    if sheet[ORGANIZATION1_LOCATION           + str(row)].value != None:
        contact.organization1_location        = sheet[ORGANIZATION1_LOCATION        + str(row)].value
    if sheet[ORGANIZATION1_JOB_DESCRIPTION    + str(row)].value != None:
        contact.organization1_job_description = sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value
    if sheet[RELATION1_TYPE                   + str(row)].value != None:
        contact.relation1_type                = sheet[RELATION1_TYPE                + str(row)].value
    if sheet[RELATION1_VALUE                  + str(row)].value != None:
        contact.relation1_value               = sheet[RELATION1_VALUE               + str(row)].value
    if sheet[EXTERNAL_ID1_TYPE                + str(row)].value != None:
        contact.external_id1_type             = sheet[EXTERNAL_ID1_TYPE             + str(row)].value
    if sheet[EXTERNAl_ID1_VALUE               + str(row)].value != None:
        contact.external_id1_value            = sheet[EXTERNAl_ID1_VALUE            + str(row)].value
    if sheet[WEBSITE1_TYPE                    + str(row)].value != None:
        contact.website1_type                 = sheet[WEBSITE1_TYPE                 + str(row)].value
    if sheet[WEBSITE1_VALUE                   + str(row)].value != None:
        contact.website1_value                = sheet[WEBSITE1_VALUE                + str(row)].value
    if sheet[CALENDAR_LINK1_TYPE              + str(row)].value != None:
        contact.calendar_link1_type           = sheet[CALENDAR_LINK1_TYPE           + str(row)].value
    if sheet[CALENDAR_LINK1_VALUE             + str(row)].value != None:
        contact.calendar_link1_value          = sheet[CALENDAR_LINK1_VALUE          + str(row)].value
    if sheet[JOT1_TYPE                        + str(row)].value != None:
        contact.jot1_type                     = sheet[JOT1_TYPE                     + str(row)].value
    if sheet[JOT1_VALUE                       + str(row)].value != None:
        contact.jot1_value                    = sheet[JOT1_VALUE                    + str(row)].value

    return contact
#}}}

# Update Contact from sheet information
# update_contact_from_sheet(sheet, row, contact)
#{{{
def update_contact_from_sheet(sheet, row, contact):
    # update these fields by adding a slash
    # INCLUDE CHANGE IN COMPANY AND CHANGE IN POSITION IN THE NOTES
    contact.notes       = special_combine(contact.notes,       sheet[NOTES       + str(row)].value)
    # update all 
    if sheet[NAME + str(row)].value != None or contact.name != None:
        info = keep_non_none(contact.name, sheet[NAME + str(row)].value)
        contact.name = info[0]
        contact.notes = add_to_notes(contact, "Name", info[1], info[0])
    if sheet[GIVEN_NAME + str(row)].value != None or contact.given_name != None:
        info = keep_non_none(contact.given_name, sheet[GIVEN_NAME + str(row)].value)
        contact.given_name = info[0]
        contact.notes = add_to_notes(contact, "Given Name", info[1], info[0])
    if sheet[ADDITIONAL_NAME + str(row)].value != None or contact.additional_name != None:
        info = keep_non_none(contact.additional_name, sheet[ADDITIONAL_NAME + str(row)].value)
        contact.additional_name = info[0]
        contact.notes = add_to_notes(contact, "Additional Name", info[1], info[0])
    if sheet[FAMILY_NAME + str(row)].value != None or contact.family_name != None:
        info = keep_non_none(contact.family_name, sheet[FAMILY_NAME + str(row)].value)
        contact.family_name = info[0]
        contact.notes = add_to_notes(contact, "Family Name", info[1], info[0])
    if sheet[YOMI_NAME + str(row)].value != None or contact.yomi_name != None:
        info = keep_non_none(contact.yomi_name, sheet[YOMI_NAME + str(row)].value)
        contact.yomi_name = info[0]
        contact.notes = add_to_notes(contact, "Yomi Name", info[1], info[0])
    if sheet[GIVEN_NAME_YOMI + str(row)].value != None or contact.given_name_yomi  != None:
        info = keep_non_none(contact.given_name_yomi, sheet[GIVEN_NAME_YOMI + str(row)].value)
        contact.given_name_yomi = info[0]
        contact.notes = add_to_notes(contact, "GIVEN_NAME_YOMI", info[1], info[0])
    if sheet[ADDITIONAL_NAME_YOMI + str(row)].value != None or contact.additional_name_yomi  != None:
        info = keep_non_none(contact.additional_name_yomi, sheet[ADDITIONAL_NAME_YOMI + str(row)].value)
        contact.additional_name_yomi = info[0]
        contact.notes = add_to_notes(contact, "ADDITIONAL_NAME_YOMI", info[1], info[0])
    if sheet[FAMILY_NAME_YOMI + str(row)].value != None or contact.family_name_yomi  != None:
        info = keep_non_none(contact.family_name_yomi, sheet[FAMILY_NAME_YOMI + str(row)].value)
        contact.family_name_yomi = info[0]
        contact.notes = add_to_notes(contact, "FAMILY_NAME_YOMI", info[1], info[0])
    if sheet[NAME_PREFIX + str(row)].value != None or contact.name_prefix  != None:
        info = keep_non_none(contact.name_prefix, sheet[NAME_PREFIX + str(row)].value)
        contact.name_prefix = info[0]
        contact.notes = add_to_notes(contact, "NAME_PREFIX", info[1], info[0])
    if sheet[NAME_SUFFIX + str(row)].value != None or contact.name_suffix  != None:
        info = keep_non_none(contact.name_suffix, sheet[NAME_SUFFIX + str(row)].value)
        contact.name_suffix = info[0]
        contact.notes = add_to_notes(contact, "NAME_SUFFIX", info[1], info[0])
    if sheet[INITIALS + str(row)].value != None or contact.initials  != None:
        info = keep_non_none(contact.initials, sheet[INITIALS + str(row)].value)
        contact.initials = info[0]
        contact.notes = add_to_notes(contact, "INITIALS", info[1], info[0])
    if sheet[NICKNAME + str(row)].value != None or contact.nickname  != None:
        info = keep_non_none(contact.nickname, sheet[NICKNAME + str(row)].value)
        contact.nickname = info[0]
        contact.notes = add_to_notes(contact, "NICKNAME", info[1], info[0])
    if sheet[SHORT_NAME + str(row)].value != None or contact.short_name  != None:
        info = keep_non_none(contact.short_name, sheet[SHORT_NAME + str(row)].value)
        contact.short_name = info[0]
        contact.notes = add_to_notes(contact, "SHORT_NAME", info[1], info[0])
    if sheet[MAIDEN_NAME + str(row)].value != None or contact.maiden_name  != None:
        info = keep_non_none(contact.maiden_name, sheet[MAIDEN_NAME + str(row)].value)
        contact.maiden_name = info[0]
        contact.notes = add_to_notes(contact, "MAIDEN_NAME", info[1], info[0])
    if sheet[BIRTHDAY + str(row)].value != None or contact.birthday  != None:
        info = keep_non_none(contact.birthday, sheet[BIRTHDAY + str(row)].value)
        contact.birthday = info[0]
        contact.notes = add_to_notes(contact, "BIRTHDAY", info[1], info[0])
    if sheet[GENDER + str(row)].value != None or contact.gender  != None:
        info = keep_non_none(contact.gender, sheet[GENDER + str(row)].value)
        contact.gender = info[0]
        contact.notes = add_to_notes(contact, "GENDER", info[1], info[0])
    if sheet[LOCATION + str(row)].value != None or contact.location  != None:
        info = keep_non_none(contact.location, sheet[LOCATION + str(row)].value)
        contact.location = info[0]
        contact.notes = add_to_notes(contact, "LOCATION", info[1], info[0])
    if sheet[BILLING_INFORMATION + str(row)].value != None or contact.billing_information  != None:
        info = keep_non_none(contact.billing_information, sheet[BILLING_INFORMATION + str(row)].value)
        contact.billing_information = info[0]
        contact.notes = add_to_notes(contact, "BILLING_INFORMATION", info[1], info[0])
    if sheet[DIRECTORY_SERVER + str(row)].value != None or contact.directory_server  != None:
        info = keep_non_none(contact.directory_server, sheet[DIRECTORY_SERVER + str(row)].value)
        contact.directory_server = info[0]
        contact.notes = add_to_notes(contact, "DIRECTORY_SERVER", info[1], info[0])
    if sheet[MILEAGE + str(row)].value != None or contact.mileage  != None:
        info = keep_non_none(contact.mileage, sheet[MILEAGE + str(row)].value)
        contact.mileage = info[0]
        contact.notes = add_to_notes(contact, "MILEAGE", info[1], info[0])
    if sheet[OCCUPATION + str(row)].value != None or contact.occupation  != None:
        info = keep_non_none(contact.occupation, sheet[OCCUPATION + str(row)].value)
        contact.occupation = info[0]
        contact.notes = add_to_notes(contact, "OCCUPATION", info[1], info[0])
    if sheet[HOBBY + str(row)].value != None or contact.hobby  != None:
        info = keep_non_none(contact.hobby, sheet[HOBBY + str(row)].value)
        contact.hobby = info[0]
        contact.notes = add_to_notes(contact, "HOBBY", info[1], info[0])
    if sheet[SENSITIVITY + str(row)].value != None or contact.sensitivity  != None:
        info = keep_non_none(contact.sensitivity, sheet[SENSITIVITY + str(row)].value)
        contact.sensitivity = info[0]
        contact.notes = add_to_notes(contact, "SENSITIVITY", info[1], info[0])
    if sheet[PRIORITY + str(row)].value != None or contact.priority  != None:
        info = keep_non_none(contact.priority, sheet[PRIORITY + str(row)].value)
        contact.priority = info[0]
        contact.notes = add_to_notes(contact, "PRIORITY", info[1], info[0])
    if sheet[SUBJECT + str(row)].value != None or contact.subject  != None:
        info = keep_non_none(contact.subject, sheet[SUBJECT + str(row)].value)
        contact.subject = info[0]
        contact.notes = add_to_notes(contact, "SUBJECT", info[1], info[0])
    if sheet[GROUP_MEMBERSHIP + str(row)].value != None or contact.group_membership  != None:
        info = keep_non_none(contact.group_membership, sheet[GROUP_MEMBERSHIP + str(row)].value)
        contact.group_membership = info[0]
        contact.notes = add_to_notes(contact, "GROUP_MEMBERSHIP", info[1], info[0])
    if sheet[EMAIL1_TYPE + str(row)].value != None or contact.email1_type  != None:
        info = keep_non_none(contact.email1_type, sheet[EMAIL1_TYPE + str(row)].value)
        contact.email1_type = info[0]
        contact.notes = add_to_notes(contact, "EMAIL1_TYPE", info[1], info[0])
    if sheet[EMAIL1_VALUE + str(row)].value != None or contact.email1_value  != None:
        info = keep_non_none(contact.email1_value, sheet[EMAIL1_VALUE + str(row)].value)
        contact.email1_value = info[0]
        if (info[0] != info[1]):
            contact.notes = special_save_email(sheet, contact, info[1], row)
    if sheet[EMAIL2_TYPE + str(row)].value != None or contact.email2_type  != None:
        info = keep_non_none(contact.email2_type, sheet[EMAIL2_TYPE + str(row)].value)
        contact.email2_type = info[0]
        contact.notes = add_to_notes(contact, "EMAIL2_TYPE", info[1], info[0])
    if sheet[EMAIL2_VALUE + str(row)].value != None or contact.email2_value  != None:
        info = keep_non_none(contact.email2_value, sheet[EMAIL2_VALUE + str(row)].value)
        contact.email2_value = info[0]
        if (info[0] != info[1]):
            contact.notes = special_save_email(sheet, contact, info[1], row)
    if sheet[EMAIL3_TYPE + str(row)].value != None or contact.email3_type  != None:
        info = keep_non_none(contact.email3_type, sheet[EMAIL3_TYPE + str(row)].value)
        contact.email3_type = info[0]
        contact.notes = add_to_notes(contact, "EMAIL3_TYPE", info[1], info[0])
    if sheet[EMAIL3_VALUE + str(row)].value != None or contact.email3_value  != None:
        info = keep_non_none(contact.email3_value, sheet[EMAIL3_VALUE + str(row)].value)
        contact.email3_value = info[0]
        if (info[0] != info[1]):
            contact.notes = special_save_email(sheet, contact, info[1], row)
    if sheet[EMAIL4_TYPE + str(row)].value != None or contact.email4_type  != None:
        info = keep_non_none(contact.email4_type, sheet[EMAIL4_TYPE + str(row)].value)
        contact.email4_type = info[0]
        contact.notes = add_to_notes(contact, "EMAIL4_TYPE", info[1], info[0])
    if sheet[EMAIL4_VALUE + str(row)].value != None or contact.email4_value  != None:
        info = keep_non_none(contact.email4_value, sheet[EMAIL4_VALUE + str(row)].value)
        contact.email4_value = info[0]
        if (info[0] != info[1]):
            contact.notes = special_save_email(sheet, contact, info[1], row)
    if sheet[EMAIL5_TYPE + str(row)].value != None or contact.email5_type  != None:
        info = keep_non_none(contact.email5_type, sheet[EMAIL5_TYPE + str(row)].value)
        contact.email5_type = info[0]
        contact.notes = add_to_notes(contact, "EMAIL5_TYPE", info[1], info[0])
    if sheet[EMAIL5_VALUE + str(row)].value != None or contact.email5_value  != None:
        info = keep_non_none(contact.email5_value, sheet[EMAIL5_VALUE + str(row)].value)
        contact.email5_value = info[0]
        if (info[0] != info[1]):
            contact.notes = special_save_email(sheet, contact, info[1], row)
    if sheet[IM1_TYPE + str(row)].value != None or contact.im1_type  != None:
        info = keep_non_none(contact.im1_type, sheet[IM1_TYPE + str(row)].value)
        contact.im1_type = info[0]
        contact.notes = add_to_notes(contact, "IM1_TYPE", info[1], info[0])
    if sheet[IM1_SERVICE + str(row)].value != None or contact.im1_service  != None:
        info = keep_non_none(contact.im1_service, sheet[IM1_SERVICE + str(row)].value)
        contact.im1_service = info[0]
        contact.notes = add_to_notes(contact, "IM1_SERVICE", info[1], info[0])
    if sheet[IM1_VALUE + str(row)].value != None or contact.im1_value  != None:
        info = keep_non_none(contact.im1_value, sheet[IM1_VALUE + str(row)].value)
        contact.im1_value = info[0]
        contact.notes = add_to_notes(contact, "IM1_VALUE", info[1], info[0])
    if sheet[PHONE1_TYPE + str(row)].value != None or contact.phone1_type  != None:
        info = keep_non_none(contact.phone1_type, sheet[PHONE1_TYPE + str(row)].value)
        contact.phone1_type = info[0]
        contact.notes = add_to_notes(contact, "PHONE1_TYPE", info[1], info[0])
    if sheet[PHONE1_VALUE + str(row)].value != None or contact.phone1_value  != None:
        info = keep_non_none(contact.phone1_value, sheet[PHONE1_VALUE + str(row)].value)
        contact.phone1_value = info[0]
        contact.notes = add_to_notes(contact, "PHONE1_VALUE", info[1], info[0])
    if sheet[PHONE2_TYPE + str(row)].value != None or contact.phone2_type  != None:
        info = keep_non_none(contact.phone2_type, sheet[PHONE2_TYPE + str(row)].value)
        contact.phone2_type = info[0]
        contact.notes = add_to_notes(contact, "PHONE2_TYPE", info[1], info[0])
    if sheet[PHONE2_VALUE + str(row)].value != None or contact.phone2_value  != None:
        info = keep_non_none(contact.phone2_value, sheet[PHONE2_VALUE + str(row)].value)
        contact.phone2_value = info[0]
        contact.notes = add_to_notes(contact, "PHONE2_VALUE", info[1], info[0])
    if sheet[PHONE3_TYPE + str(row)].value != None or contact.phone3_type  != None:
        info = keep_non_none(contact.phone3_type, sheet[PHONE3_TYPE + str(row)].value)
        contact.phone3_type = info[0]
        contact.notes = add_to_notes(contact, "PHONE3_TYPE", info[1], info[0])
    if sheet[PHONE3_VALUE + str(row)].value != None or contact.phone3_value  != None:
        info = keep_non_none(contact.phone3_value, sheet[PHONE3_VALUE + str(row)].value)
        contact.phone3_value = info[0]
        contact.notes = add_to_notes(contact, "PHONE3_VALUE", info[1], info[0])
    if sheet[PHONE4_TYPE + str(row)].value != None or contact.phone4_type  != None:
        info = keep_non_none(contact.phone4_type, sheet[PHONE4_TYPE + str(row)].value)
        contact.phone4_type = info[0]
        contact.notes = add_to_notes(contact, "PHONE4_TYPE", info[1], info[0])
    if sheet[PHONE4_VALUE + str(row)].value != None or contact.phone4_value  != None:
        info = keep_non_none(contact.phone4_value, sheet[PHONE4_VALUE + str(row)].value)
        contact.phone4_value = info[0]
        contact.notes = add_to_notes(contact, "PHONE4_VALUE", info[1], info[0])
    if sheet[PHONE5_TYPE + str(row)].value != None or contact.phone5_type  != None:
        info = keep_non_none(contact.phone5_type, sheet[PHONE5_TYPE + str(row)].value)
        contact.phone5_type = info[0]
        contact.notes = add_to_notes(contact, "PHONE5_TYPE", info[1], info[0])
    if sheet[PHONE5_VALUE + str(row)].value != None or contact.phone5_value  != None:
        info = keep_non_none(contact.phone5_value, sheet[PHONE5_VALUE + str(row)].value)
        contact.phone5_value = info[0]
        contact.notes = add_to_notes(contact, "PHONE5_VALUE", info[1], info[0])
    if sheet[ADDRESS1_TYPE + str(row)].value != None or contact.address1_type  != None:
        info = keep_non_none(contact.address1_type, sheet[ADDRESS1_TYPE + str(row)].value)
        contact.address1_type = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_TYPE", info[1], info[0])
    if sheet[ADDRESS1_FORMATED + str(row)].value != None or contact.address1_formated  != None:
        info = keep_non_none(contact.address1_formated, sheet[ADDRESS1_FORMATED + str(row)].value)
        contact.address1_formated = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_FORMATED", info[1], info[0])
    if sheet[ADDRESS1_STREET + str(row)].value != None or contact.address1_street  != None:
        info = keep_non_none(contact.address1_street, sheet[ADDRESS1_STREET + str(row)].value)
        contact.address1_street = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_STREET", info[1], info[0])
    if sheet[ADDRESS1_CITY + str(row)].value != None or contact.address1_city  != None:
        info = keep_non_none(contact.address1_city, sheet[ADDRESS1_CITY + str(row)].value)
        contact.address1_city = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_CITY", info[1], info[0])
    if sheet[ADDRESS1_POBOX + str(row)].value != None or contact.address1_pobox  != None:
        info = keep_non_none(contact.address1_pobox, sheet[ADDRESS1_POBOX + str(row)].value)
        contact.address1_pobox = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_POBOX", info[1], info[0])
    if sheet[ADDRESS1_REGION + str(row)].value != None or contact.address1_region  != None:
        info = keep_non_none(contact.address1_region, sheet[ADDRESS1_REGION + str(row)].value)
        contact.address1_region = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_REGION", info[1], info[0])
    if sheet[ADDRESS1_POSTAL_CODE + str(row)].value != None or contact.address1_postal_code  != None:
        info = keep_non_none(contact.address1_postal_code, sheet[ADDRESS1_POSTAL_CODE + str(row)].value)
        contact.address1_postal_code = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_POSTAL_CODE", info[1], info[0])
    if sheet[ADDRESS1_COUNTRY + str(row)].value != None or contact.address1_country  != None:
        info = keep_non_none(contact.address1_country, sheet[ADDRESS1_COUNTRY + str(row)].value)
        contact.address1_country = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_COUNTRY", info[1], info[0])
    if sheet[ADDRESS1_EXTENDED_ADDRESS + str(row)].value != None or contact.address1_extended_address  != None:
        info = keep_non_none(contact.address1_extended_address, sheet[ADDRESS1_EXTENDED_ADDRESS + str(row)].value)
        contact.address1_extended_address = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS1_EXTENDED_ADDRESS", info[1], info[0])
    if sheet[ADDRESS2_TYPE + str(row)].value != None or contact.address2_type  != None:
        info = keep_non_none(contact.address2_type, sheet[ADDRESS2_TYPE + str(row)].value)
        contact.address2_type = info[0]
        contact.notes = add_to_notes(contact, "Address2 Type", info[1], info[0])
    if sheet[ADDRESS2_FORMATED + str(row)].value != None or contact.address2_formated  != None:
        info = keep_non_none(contact.address2_formated, sheet[ADDRESS2_FORMATED + str(row)].value)
        contact.address2_formated = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_FORMATED", info[1], info[0])
    if sheet[ADDRESS2_STREET + str(row)].value != None or contact.address2_street  != None:
        info = keep_non_none(contact.address2_street, sheet[ADDRESS2_STREET + str(row)].value)
        contact.address2_street = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_STREET", info[1], info[0])
    if sheet[ADDRESS2_CITY + str(row)].value != None or contact.address2_city  != None:
        info = keep_non_none(contact.address2_city, sheet[ADDRESS2_CITY + str(row)].value)
        contact.address2_city = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_CITY", info[1], info[0])
    if sheet[ADDRESS2_POBOX + str(row)].value != None or contact.address2_pobox  != None:
        info = keep_non_none(contact.address2_pobox, sheet[ADDRESS2_POBOX + str(row)].value)
        contact.address2_pobox = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_POBOX", info[1], info[0])
    if sheet[ADDRESS2_REGION + str(row)].value != None or contact.address2_region  != None:
        info = keep_non_none(contact.address2_region, sheet[ADDRESS2_REGION + str(row)].value)
        contact.address2_region = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_REGION", info[1], info[0])
    if sheet[ADDRESS2_POSTAL_CODE + str(row)].value != None or contact.address2_postal_code  != None:
        info = keep_non_none(contact.address2_postal_code, sheet[ADDRESS2_POSTAL_CODE + str(row)].value)
        contact.address2_postal_code = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_POSTAL_CODE", info[1], info[0])
    if sheet[ADDRESS2_COUNTRY + str(row)].value != None or contact.address2_country  != None:
        info = keep_non_none(contact.address2_country, sheet[ADDRESS2_COUNTRY + str(row)].value)
        contact.address2_country = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS2_COUNTRY", info[1], info[0])
    if sheet[ADDRESS2_EXTENDED_ADDRESS + str(row)].value != None or contact.address2_extended_address  != None:
        info = keep_non_none(contact.address2_extended_address , sheet[ADDRESS2_EXTENDED_ADDRESS + str(row)].value)
        contact.address2_extended_address = info[0]
        contact.notes = add_to_notes(contact, "Address 2 Extended Address", info[1], info[0])
    if sheet[ADDRESS3_TYPE + str(row)].value != None or contact.address3_type  != None:
        info = keep_non_none(contact.address3_type, sheet[ADDRESS3_TYPE + str(row)].value)
        contact.address3_type = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_TYPE", info[1], info[0])
    if sheet[ADDRESS3_FORMATED + str(row)].value != None or contact.address3_formated  != None:
        info = keep_non_none(contact.address3_formated, sheet[ADDRESS3_FORMATED + str(row)].value)
        contact.address3_formated = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_FORMATED", info[1], info[0])
    if sheet[ADDRESS3_STREET + str(row)].value != None or contact.address3_street  != None:
        info = keep_non_none(contact.address3_street, sheet[ADDRESS3_STREET + str(row)].value)
        contact.address3_street = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_STREET", info[1], info[0])
    if sheet[ADDRESS3_CITY + str(row)].value != None or contact.address3_city  != None:
        info = keep_non_none(contact.address3_city, sheet[ADDRESS3_CITY + str(row)].value)
        contact.address3_city = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_CITY", info[1], info[0])
    if sheet[ADDRESS3_POBOX + str(row)].value != None or contact.address3_pobox  != None:
        info = keep_non_none(contact.address3_pobox, sheet[ADDRESS3_POBOX + str(row)].value)
        contact.address3_pobox = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_POBOX", info[1], info[0])
    if sheet[ADDRESS3_REGION + str(row)].value != None or contact.address3_region  != None:
        info = keep_non_none(contact.address3_region, sheet[ADDRESS3_REGION + str(row)].value)
        contact.address3_region = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_REGION", info[1], info[0])
    if sheet[ADDRESS3_POSTAL_CODE + str(row)].value != None or contact.address3_postal_code  != None:
        info = keep_non_none(contact.address3_postal_code, sheet[ADDRESS3_POSTAL_CODE + str(row)].value)
        contact.address3_postal_code = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_POSTAL_CODE", info[1], info[0])
    if sheet[ADDRESS3_COUNTRY + str(row)].value != None or contact.address3_country  != None:
        info = keep_non_none(contact.address3_country, sheet[ADDRESS3_COUNTRY + str(row)].value)
        contact.address3_country = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_COUNTRY", info[1], info[0])
    if sheet[ADDRESS3_EXTENDED_ADDRESS + str(row)].value != None or contact.address3_extended_address  != None:
        info = keep_non_none(contact.address3_extended_address, sheet[ADDRESS3_EXTENDED_ADDRESS + str(row)].value)
        contact.address3_extended_address = info[0]
        contact.notes = add_to_notes(contact, "ADDRESS3_EXTENDED_ADDRESS", info[1], info[0])
    if sheet[ORGANIZATION1_TYPE + str(row)].value != None or contact.organization1_type  != None:
        info = keep_non_none(contact.organization1_type, sheet[ORGANIZATION1_TYPE + str(row)].value)
        contact.organization1_type = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_TYPE", info[1], info[0])
    if sheet[ORGANIZATION1_NAME + str(row)].value != None or contact.organization1_name  != None:
        info = keep_non_none(contact.organization1_name, sheet[ORGANIZATION1_NAME + str(row)].value)
        contact.organization1_name = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_NAME", info[1], info[0])
    if sheet[ORGANIZATION1_YOMI_NAME + str(row)].value != None or contact.organization1_yomi_name  != None:
        info = keep_non_none(contact.organization1_yomi_name, sheet[ORGANIZATION1_YOMI_NAME + str(row)].value)
        contact.organization1_yomi_name = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_YOMI_NAME", info[1], info[0])
    if sheet[ORGANIZATION1_TITLE + str(row)].value != None or contact.organization1_title  != None:
        info = keep_non_none(contact.organization1_title, sheet[ORGANIZATION1_TITLE + str(row)].value)
        contact.organization1_title = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_TITLE", info[1], info[0])
    if sheet[ORGANIZATION1_DEPARTMENT + str(row)].value != None or contact.organization1_department  != None:
        info = keep_non_none(contact.organization1_department, sheet[ORGANIZATION1_DEPARTMENT + str(row)].value)
        contact.organization1_department = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_DEPARTMENT", info[1], info[0])
    if sheet[ORGANIZATION1_SYMBOL + str(row)].value != None or contact.organization1_symbol  != None:
        info = keep_non_none(contact.organization1_symbol, sheet[ORGANIZATION1_SYMBOL + str(row)].value)
        contact.organization1_symbol = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_SYMBOL", info[1], info[0])
    if sheet[ORGANIZATION1_LOCATION + str(row)].value != None or contact.organization1_location  != None:
        info = keep_non_none(contact.organization1_location, sheet[ORGANIZATION1_LOCATION + str(row)].value)
        contact.organization1_location = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_LOCATION", info[1], info[0])
    if sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value != None or contact.organization1_job_description  != None:
        info = keep_non_none(contact.organization1_job_description, sheet[ORGANIZATION1_JOB_DESCRIPTION + str(row)].value)
        contact.organization1_job_description = info[0]
        contact.notes = add_to_notes(contact, "ORGANIZATION1_JOB_DESCRIPTION", info[1], info[0])
    if sheet[RELATION1_TYPE + str(row)].value != None or contact.relation1_type  != None:
        info = keep_non_none(contact.relation1_type, sheet[RELATION1_TYPE + str(row)].value)
        contact.relation1_type = info[0]
        contact.notes = add_to_notes(contact, "RELATION1_TYPE", info[1], info[0])
    if sheet[RELATION1_VALUE + str(row)].value != None or contact.relation1_value  != None:
        info = keep_non_none(contact.relation1_value, sheet[RELATION1_VALUE + str(row)].value)
        contact.relation1_value = info[0]
        contact.notes = add_to_notes(contact, "RELATION1_VALUE", info[1], info[0])
    if sheet[EXTERNAL_ID1_TYPE + str(row)].value != None or contact.external_id1_type  != None:
        info = keep_non_none(contact.external_id1_type, sheet[EXTERNAL_ID1_TYPE + str(row)].value)
        contact.external_id1_type = info[0]
        contact.notes = add_to_notes(contact, "EXTERNAL_ID1_TYPE", info[1], info[0])
    if sheet[EXTERNAl_ID1_VALUE + str(row)].value != None or contact.external_id1_value  != None:
        info = keep_non_none(contact.external_id1_value, sheet[EXTERNAl_ID1_VALUE + str(row)].value)
        contact.external_id1_value = info[0]
        contact.notes = add_to_notes(contact, "EXTERNAl_ID1_VALUE", info[1], info[0])
    if sheet[WEBSITE1_TYPE + str(row)].value != None or contact.website1_type  != None:
        info = keep_non_none(contact.website1_type, sheet[WEBSITE1_TYPE + str(row)].value)
        contact.website1_type = info[0]
        contact.notes = add_to_notes(contact, "WEBSITE1_TYPE", info[1], info[0])
    if sheet[WEBSITE1_VALUE + str(row)].value != None or contact.website1_value  != None:
        info = keep_non_none(contact.website1_value, sheet[WEBSITE1_VALUE + str(row)].value)
        contact.website1_value = info[0]
        contact.notes = add_to_notes(contact, "WEBSITE1_VALUE", info[1], info[0])
    if sheet[CALENDAR_LINK1_TYPE + str(row)].value != None or contact.calendar_link1_type  != None:
        info = keep_non_none(contact.calendar_link1_type, sheet[CALENDAR_LINK1_TYPE + str(row)].value)
        contact.calendar_link1_type = info[0]
        contact.notes = add_to_notes(contact, "CALENDAR_LINK1_TYPE", info[1], info[0])
    if sheet[CALENDAR_LINK1_VALUE + str(row)].value != None or contact.calendar_link1_value  != None:
        info = keep_non_none(contact.calendar_link1_value, sheet[CALENDAR_LINK1_VALUE + str(row)].value)
        contact.calendar_link1_value = info[0]
        contact.notes = add_to_notes(contact, "CALENDAR_LINK1_VALUE", info[1], info[0])
    if sheet[JOT1_TYPE + str(row)].value != None or contact.jot1_type  != None:
        info = keep_non_none(contact.jot1_type, sheet[JOT1_TYPE + str(row)].value)
        contact.jot1_type = info[0]
        contact.notes = add_to_notes(contact, "JOT1_TYPE", info[1], info[0])
    if sheet[JOT1_VALUE + str(row)].value != None or contact.jot1_value  != None:
        info = keep_non_none(contact.jot1_value, sheet[JOT1_VALUE + str(row)].value)
        contact.jot1_value = info[0]
        contact.notes = add_to_notes(contact, "JOT1_VALUE", info[1], info[0])

    return contact
#}}}

################
# MAIN METHODS #
################

# Format all numbers in a column
# format_all_numbers(fileName, sheetName, col)
# Call using the command line
# format_all_numbers(filename, startRow, *cols)
#{{{
def format_all_numbers(*args):
    # turn the arguments into variable names
    args = args[0]
    fileName = args[1]
    startRow = int(args[2])
    cols = args[3:]
    if printing:
        print("Opening...")
    wb = openpyxl.load_workbook(args[1])
    # sheet = wb[sheetName]
    sheet = wb.worksheets[0]

    # look through all columns and change the phone numbers
    first = startRow
    last = sheet.max_row
    for col in cols:
        for row in range (first, last + 1):
            number = str(sheet[col + str(row)].value)
            formatted = formatting_phone_number(number)
            if saving:
                sheet[col + str(row)].value = formatted

    if printing:
        print("Processing " + str((last + 1) - first) + " rows...")

    # add the word 'formatted' and save the new file where the original is
    newName = 'formatted'
    index = fileName[::-1].find('/')
    if printing and saving:
        end = fileName[-index - 1:]
        fileName = fileName[:-index - 1] + newName + end[0].capitalize() + end[1:]
        print("Saving " + fileName)
        wb.save(fileName)

    # note for when the script is over
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Change 'US' and 'United States of America' to 'United States'
# standardize_USA(fileName, start, *cols)
#{{{
# def standardize_USA(fileName, start, column):
def standardize_USA(*args):
    # turn the arguments into variable names
    args = args[0]
    fileName = args[1]
    start = int(args[2])
    cols = args[3:]
    # Open an existing excel file
    if printing:
        print("Opening...")
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    #################
    # DO STUFF HERE #
    #################
    for col in cols:
        for row in range (start, sheet.max_row + 1):
            country = str(sheet[col+ str(row)].value)
            regexUSA = '(U\.?S\.?A?\.?|United ?States ?(of ?America)?)'
            matchUSA = re.search(regexUSA, country, re.IGNORECASE)
            if matchUSA:
                sheet[col+ str(row)].value = "United States"
            regexUK = '(U\.?K\.?)'
            matchUK = re.search(regexUK, country, re.IGNORECASE)
            if matchUK:
                sheet[col+ str(row)].value = "United Kingdom"


    if printing:
        print("Saving...")

    wb.save("betterFile.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

    if printing:
        print()
        print("Done!")
#}}}

# Change 'Ma' to 'MA' etc
# fix_states(fileName, start, *cols)
#{{{
def fix_states(*args):
    # turn the arguments into variable names
    args = args[0]
    fileName = args[1]
    start = int(args[2])
    cols = args[3:]
    # Open an existing excel file
    if printing:
        print("Opening...")
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    #################
    # DO STUFF HERE #
    #################
    for col in cols:
        if printing:
            print("Working on column " + col)
        for row in range (start, sheet.max_row + 1):
            state = sheet[col + str(row)].value
            if state:
                sheet[col + str(row)].value = state.upper()

    if printing:
        print("Saving...")

    wb.save("states.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

    if printing:
        print()
        print("Done!")
#}}}

# Combine LinkedIn and Google Contacts
# Copies all contacts from first list to second
# combine_contacts(fileName1, sheetName1, fileName2, sheetName2)
#{{{
def combine_contacts(fileName1, sheetName1, fileName2, sheetName2):
    if printing:
        print("Opening...")
    wb1 = openpyxl.load_workbook(fileName1 + ".xlsx")
    sheet1 = wb1[sheetName1]
    wb2 = openpyxl.load_workbook(fileName2 + ".xlsx")
    sheet2 = wb2[sheetName2]

    # if 'sheet' appears randomly we can delete it
    # rm = out.get_sheet_by_name('Sheet')
    # out.remove_sheet(rm)

    first = 2
    last = sheet2.max_row
    start = sheet1.max_row

    for row in range (first, start + 1):
        if sheet1[NAME + str(row)].value != None:
            sheet1[NAME + str(row)].value = str(sheet1[NAME + str(row)].value).title()
        if sheet1[GIVEN_NAME + str(row)].value != None:
            sheet1[GIVEN_NAME + str(row)].value = str(sheet1[GIVEN_NAME + str(row)].value).title()
        if sheet1[ADDITIONAL_NAME + str(row)].value != None:
            sheet1[ADDITIONAL_NAME + str(row)].value = str(sheet1[ADDITIONAL_NAME + str(row)].value).title()
        if sheet1[FAMILY_NAME + str(row)].value != None:
            sheet1[FAMILY_NAME + str(row)].value = str(sheet1[FAMILY_NAME + str(row)].value).title()
        if sheet1[YOMI_NAME + str(row)].value != None:
            sheet1[YOMI_NAME + str(row)].value = str(sheet1[YOMI_NAME + str(row)].value).title()
        if sheet1[GIVEN_NAME_YOMI + str(row)].value != None:
            sheet1[GIVEN_NAME_YOMI + str(row)].value = str(sheet1[GIVEN_NAME_YOMI + str(row)].value).title()
        if sheet1[ADDITIONAL_NAME_YOMI + str(row)].value != None:
            sheet1[ADDITIONAL_NAME_YOMI + str(row)].value = str(sheet1[ADDITIONAL_NAME_YOMI + str(row)].value).title()
        if sheet1[FAMILY_NAME_YOMI + str(row)].value != None:
            sheet1[FAMILY_NAME_YOMI + str(row)].value = str(sheet1[FAMILY_NAME_YOMI + str(row)].value).title()
        if sheet1[NAME_PREFIX + str(row)].value != None:
            sheet1[NAME_PREFIX + str(row)].value = str(sheet1[NAME_PREFIX + str(row)].value).title()
        if sheet1[NAME_SUFFIX + str(row)].value != None:
            sheet1[NAME_SUFFIX + str(row)].value = str(sheet1[NAME_SUFFIX + str(row)].value).title()
        if sheet1[INITIALS + str(row)].value != None:
            sheet1[INITIALS + str(row)].value = str(sheet1[INITIALS + str(row)].value).title()
        if sheet1[NICKNAME + str(row)].value != None:
            sheet1[NICKNAME + str(row)].value = str(sheet1[NICKNAME + str(row)].value).title()
        if sheet1[SHORT_NAME + str(row)].value != None:
            sheet1[SHORT_NAME + str(row)].value = str(sheet1[SHORT_NAME + str(row)].value).title()
        if sheet1[MAIDEN_NAME + str(row)].value != None:
            sheet1[MAIDEN_NAME + str(row)].value = str(sheet1[MAIDEN_NAME + str(row)].value).title()
        if sheet1[ADDRESS1_TYPE + str(row)].value != None:
            sheet1[ADDRESS1_TYPE + str(row)].value = str(sheet1[ADDRESS1_TYPE + str(row)].value).title()
        if sheet1[ADDRESS1_FORMATED + str(row)].value != None:
            sheet1[ADDRESS1_FORMATED + str(row)].value = str(sheet1[ADDRESS1_FORMATED + str(row)].value).title()
        if sheet1[ADDRESS1_STREET + str(row)].value != None:
            sheet1[ADDRESS1_STREET + str(row)].value = str(sheet1[ADDRESS1_STREET + str(row)].value).title()
        if sheet1[ADDRESS1_CITY + str(row)].value != None:
            sheet1[ADDRESS1_CITY + str(row)].value = str(sheet1[ADDRESS1_CITY + str(row)].value).title()
        if sheet1[ADDRESS1_REGION + str(row)].value != None and len(sheet1[ADDRESS1_REGION + str(row)].value) != 2:
            sheet1[ADDRESS1_REGION + str(row)].value = str(sheet1[ADDRESS1_REGION + str(row)].value).title()
        if sheet1[ADDRESS1_COUNTRY + str(row)].value != None:
            sheet1[ADDRESS1_COUNTRY + str(row)].value = str(sheet1[ADDRESS1_COUNTRY + str(row)].value).title()
        if sheet1[ADDRESS1_EXTENDED_ADDRESS + str(row)].value != None:
            sheet1[ADDRESS1_EXTENDED_ADDRESS + str(row)].value = str(sheet1[ADDRESS1_EXTENDED_ADDRESS + str(row)].value).title()
        if sheet1[ADDRESS2_TYPE + str(row)].value != None:
            sheet1[ADDRESS2_TYPE + str(row)].value = str(sheet1[ADDRESS2_TYPE + str(row)].value).title()
        if sheet1[ADDRESS2_FORMATED + str(row)].value != None:
            sheet1[ADDRESS2_FORMATED + str(row)].value = str(sheet1[ADDRESS2_FORMATED + str(row)].value).title()
        if sheet1[ADDRESS2_STREET + str(row)].value != None:
            sheet1[ADDRESS2_STREET + str(row)].value = str(sheet1[ADDRESS2_STREET + str(row)].value).title()
        if sheet1[ADDRESS2_CITY + str(row)].value != None:
            sheet1[ADDRESS2_CITY + str(row)].value = str(sheet1[ADDRESS2_CITY + str(row)].value).title()
        if sheet1[ADDRESS2_REGION + str(row)].value != None and len(sheet1[ADDRESS2_REGION + str(row)].value) != 2:
            sheet1[ADDRESS2_REGION + str(row)].value = str(sheet1[ADDRESS2_REGION + str(row)].value).title()
        if sheet1[ADDRESS2_COUNTRY + str(row)].value != None:
            sheet1[ADDRESS2_COUNTRY + str(row)].value = str(sheet1[ADDRESS2_COUNTRY + str(row)].value).title()
        if sheet1[ADDRESS2_EXTENDED_ADDRESS + str(row)].value != None:
            sheet1[ADDRESS2_EXTENDED_ADDRESS + str(row)].value = str(sheet1[ADDRESS2_EXTENDED_ADDRESS + str(row)].value).title()
        if sheet1[ADDRESS3_TYPE + str(row)].value != None:
            sheet1[ADDRESS3_TYPE + str(row)].value = str(sheet1[ADDRESS3_TYPE + str(row)].value).title()
        if sheet1[ADDRESS3_FORMATED + str(row)].value != None:
            sheet1[ADDRESS3_FORMATED + str(row)].value = str(sheet1[ADDRESS3_FORMATED + str(row)].value).title()
        if sheet1[ADDRESS3_STREET + str(row)].value != None:
            sheet1[ADDRESS3_STREET + str(row)].value = str(sheet1[ADDRESS3_STREET + str(row)].value).title()
        if sheet1[ADDRESS3_CITY + str(row)].value != None:
            sheet1[ADDRESS3_CITY + str(row)].value = str(sheet1[ADDRESS3_CITY + str(row)].value).title()
        if sheet1[ADDRESS3_REGION + str(row)].value != None and len(sheet1[ADDRESS3_REGION + str(row)].value) != 2:
            sheet1[ADDRESS3_REGION + str(row)].value = str(sheet1[ADDRESS3_REGION + str(row)].value).title()
        if sheet1[ADDRESS3_COUNTRY + str(row)].value != None:
            sheet1[ADDRESS3_COUNTRY + str(row)].value = str(sheet1[ADDRESS3_COUNTRY + str(row)].value).title()
        if sheet1[ADDRESS3_EXTENDED_ADDRESS + str(row)].value != None:
            sheet1[ADDRESS3_EXTENDED_ADDRESS + str(row)].value = str(sheet1[ADDRESS3_EXTENDED_ADDRESS + str(row)].value).title()

    for row in range (first, last + 1):
        sheet1[NAME + str(row + start)].value = sheet2["A" + str(row)].value + " " + sheet2["B" + str(row)].value
        sheet1[GIVEN_NAME + str(row + start)].value = sheet2["A" + str(row)].value
        sheet1[FAMILY_NAME + str(row + start)].value = sheet2["B" + str(row)].value
        sheet1[EMAIL1_VALUE + str(row + start)].value = sheet2["C" + str(row)].value
        sheet1[ORGANIZATION1_NAME + str(row + start)].value = sheet2["D" + str(row)].value
        sheet1[ORGANIZATION1_TITLE + str(row + start)].value = sheet2["E" + str(row)].value
        sheet1[GROUP_MEMBERSHIP + str(row + start)].value = sheet2["F" + str(row)].value

    wb1.save("combined.xlsx")

    if printing:
        print()
        print("Done!")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Combine combined list with a CMAShipping List
# Copies all contacts from sheet 1 to sheet 2
# combine_with_CMA(fileName1, sheetName1, fileName2, sheetName2)
#{{{
'''
- loop through list 1 putting in list 2
  - FORMULA: row + (start + 1) - first
'''
def combine_with_CMA(fileName1, sheetName1, fileName2, sheetName2):
    if printing:
        print("Opening...")
    wb1 = openpyxl.load_workbook(fileName1 + ".xlsx")
    sheet = wb1[sheetName1]
    wb2 = openpyxl.load_workbook(fileName2 + ".xlsx")
    outsheet = wb2[sheetName2]

    first = 3
    last = sheet.max_row
    start = outsheet.max_row

    for row in range (first, last + 1):
        index = row + (start + 1) - first
        if printing:
            print(str(row) + " " + sheet['A' + str(row)].value)
        names = determine_names(sheet['A' + str(row)].value)

        # split name here
        outsheet[FIRST_NAME           + str(index)].value = names['first_name']
        outsheet[MIDDLE_NAME          + str(index)].value = names['middle_name']
        outsheet[LAST_NAME            + str(index)].value = names['last_name']

        # position and company
        outsheet[COMPANY              + str(index)].value = sheet['C' + str(row)].value
        outsheet[JOB_TITLE            + str(index)].value = sheet['B' + str(row)].value

        # put address stuff here
        outsheet[BUSINESS_STREET      + str(index)].value = sheet['D' + str(row)].value
        outsheet[BUSINESS_CITY        + str(index)].value = sheet['G' + str(row)].value
        outsheet[BUSINESS_STATE       + str(index)].value = sheet['H' + str(row)].value
        outsheet[BUSINESS_POSTAL_CODE + str(index)].value = sheet['I' + str(row)].value
        outsheet[BUSINESS_COUNTRY     + str(index)].value = sheet['J' + str(row)].value
        
        # contact info
        outsheet[BUSINESS_PHONE       + str(index)].value = sheet['K' + str(row)].value
        outsheet[EMAIL_ADDRESS        + str(index)].value = sheet['L' + str(row)].value

        # location info
        outsheet[LATITUDE             + str(index)].value = sheet['M' + str(row)].value
        outsheet[LONGITUDE            + str(index)].value = sheet['N' + str(row)].value

        # put CMA SHIPPING
        outsheet[CATEGORIES           + str(index)].value = 'CMA Shipping conference'
        # put March 22, 2018
        outsheet[CONNECTED_ON         + str(index)].value = '03/22/2018'


    wb2.save("combined.xlsx")

    if printing:
        print()
        print("Done!")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# Remove Duplicate Contacts
# remove_duplicate_contacts(fileName, sheetName, first, last)
#{{{
# all rows algorithm
# - the first row is automatically its own entity
# - for all other rows take combination of primary contact and current contact we are looking at and compare edit distance to it
#   - if it is a match combine and keep looking through list until we find a bad match
#   - if distance is way off save old contact info. store new contact and keep going
# def remove_duplicate_contacts(fileName, first, match_threshold, *cols):
def remove_duplicate_contacts(*args):
    # turn the arguments into variable names
    args = args[0]
    fileName = args[1]
    low_threshold = int(args[2])
    cols = args[3:]
    if printing:
        print("Opening...")
    # Open the file for editing
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("contacts")
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    # Create a new file to store duplicate contacts
    dupe = openpyxl.Workbook()
    # Open the worksheet we want to edit
    dupesheet = dupe.create_sheet("contacts")
    # if 'sheet' appears randomly we can delete it
    rm = dupe.get_sheet_by_name('Sheet')
    dupe.remove_sheet(rm)

    # - create an object for a new primary contact and account name pair
    #   - store previous object in a new sheet
    #   - store all information here
    #   - look at next contact and see if it matches (edit distance)
    #     - if it is a match combine and keep going, otherwise repeat
    compare = ""
    current = ""
    last = sheet.max_row
    count = 1
    dupes = 2
    first = 2
    contact = Contact()

    # Create Headers
    #{{{
    if saving:
        outsheet[NAME + '1'].value                           = "Name"
        outsheet[GIVEN_NAME + '1'].value                     = "Given Name"
        outsheet[ADDITIONAL_NAME + '1'].value                = "Additional Name"
        outsheet[FAMILY_NAME + '1'].value                    = "Family Name"
        outsheet[YOMI_NAME + '1'].value                      = "Yomi name"
        outsheet[GIVEN_NAME_YOMI + '1'].value                = "Given Name Yomi"
        outsheet[ADDITIONAL_NAME_YOMI + '1'].value           = "Additional Name Yomi"
        outsheet[FAMILY_NAME_YOMI + '1'].value               = "Family Name Yomi"
        outsheet[NAME_PREFIX + '1'].value                    = "Name Prefix"
        outsheet[NAME_SUFFIX + '1'].value                    = "Name Suffix"
        outsheet[INITIALS + '1'].value                       = "Initials"
        outsheet[NICKNAME + '1'].value                       = "Nickname"
        outsheet[SHORT_NAME + '1'].value                     = "Short Name"
        outsheet[MAIDEN_NAME + '1'].value                    = "Maiden Name"
        outsheet[BIRTHDAY + '1'].value                       = "Birthday"
        outsheet[GENDER + '1'].value                         = "Gender"
        outsheet[LOCATION + '1'].value                       = "Location"
        outsheet[BILLING_INFORMATION + '1'].value            = "Billing Information"
        outsheet[DIRECTORY_SERVER + '1'].value               = "Directory Server"
        outsheet[MILEAGE + '1'].value                        = "Mileage"
        outsheet[OCCUPATION + '1'].value                     = "Occupation"
        outsheet[HOBBY + '1'].value                          = "Hobby"
        outsheet[SENSITIVITY + '1'].value                    = "Sensitivity"
        outsheet[PRIORITY + '1'].value                       = "Priority"
        outsheet[SUBJECT + '1'].value                        = "Subject"
        outsheet[NOTES + '1'].value                          = "Notes"
        outsheet[GROUP_MEMBERSHIP + '1'].value               = "Group Membership"
        outsheet[EMAIL1_TYPE + '1'].value                    = "E-mail 1 - Type"
        outsheet[EMAIL1_VALUE + '1'].value                   = "E-mail 1 - Value"
        outsheet[EMAIL2_TYPE + '1'].value                    = "E-mail 2 - Type"
        outsheet[EMAIL2_VALUE + '1'].value                   = "E-mail 2 - Value"
        outsheet[EMAIL3_TYPE + '1'].value                    = "E-mail 3 - Type"
        outsheet[EMAIL3_VALUE + '1'].value                   = "E-mail 3 - Value"
        outsheet[EMAIL4_TYPE + '1'].value                    = "E-mail 4 - Type"
        outsheet[EMAIL4_VALUE + '1'].value                   = "E-mail 4 - Value"
        outsheet[EMAIL5_TYPE + '1'].value                    = "E-mail 5 - Type"
        outsheet[EMAIL5_VALUE + '1'].value                   = "E-mail 5 - Value"
        outsheet[IM1_TYPE + '1'].value                       = "IM 1 - Type"
        outsheet[IM1_SERVICE + '1'].value                    = "IM 1 - Service"
        outsheet[IM1_VALUE + '1'].value                      = "IM 1 - Value"
        outsheet[PHONE1_TYPE + '1'].value                    = "Phone 1 - Type"
        outsheet[PHONE1_VALUE + '1'].value                   = "Phone 1 – Value"
        outsheet[PHONE2_TYPE + '1'].value                    = "Phone 2 – Type"
        outsheet[PHONE2_VALUE + '1'].value                   = "Phone 2 – Value"
        outsheet[PHONE3_TYPE + '1'].value                    = "Phone 3 – Type"
        outsheet[PHONE3_VALUE + '1'].value                   = "Phone 3 – Value"
        outsheet[PHONE4_TYPE + '1'].value                    = "Phone 4 – Type"
        outsheet[PHONE4_VALUE + '1'].value                   = "Phone 4 – Value"
        outsheet[PHONE5_TYPE + '1'].value                    = "Phone 5 – Type"
        outsheet[PHONE5_VALUE + '1'].value                   = "Phone 5 – Value"
        outsheet[ADDRESS1_TYPE + '1'].value                  = "Address 1 - Type"
        outsheet[ADDRESS1_FORMATED + '1'].value              = "Address 1 - Formatted"
        outsheet[ADDRESS1_STREET + '1'].value                = "Address 1 - Street"
        outsheet[ADDRESS1_CITY + '1'].value                  = "Address 1 - City"
        outsheet[ADDRESS1_POBOX + '1'].value                 = "Address 1 - PO Box"
        outsheet[ADDRESS1_REGION + '1'].value                = "Address 1 - Region"
        outsheet[ADDRESS1_POSTAL_CODE + '1'].value           = "Address 1 - Postal Code"
        outsheet[ADDRESS1_COUNTRY + '1'].value               = "Address 1 - Country"
        outsheet[ADDRESS1_EXTENDED_ADDRESS + '1'].value      = "Address 1 - Extended Address"
        outsheet[ADDRESS2_TYPE + '1'].value                  = "Address 2 - Type"
        outsheet[ADDRESS2_FORMATED + '1'].value              = "Address 2 - Formatted"
        outsheet[ADDRESS2_STREET + '1'].value                = "Address 2 - Street"
        outsheet[ADDRESS2_CITY + '1'].value                  = "Address 2 - City"
        outsheet[ADDRESS2_POBOX + '1'].value                 = "Address 2 - PO Box"
        outsheet[ADDRESS2_REGION + '1'].value                = "Address 2 - Region"
        outsheet[ADDRESS2_POSTAL_CODE + '1'].value           = "Address 2 - Postal Code"
        outsheet[ADDRESS2_COUNTRY + '1'].value               = "Address 2 - Country"
        outsheet[ADDRESS2_EXTENDED_ADDRESS + '1'].value      = "Address 2 - Extended Address"
        outsheet[ADDRESS3_TYPE + '1'].value                  = "Address 3 - Type"
        outsheet[ADDRESS3_FORMATED + '1'].value              = "Address 3 - Formatted"
        outsheet[ADDRESS3_STREET + '1'].value                = "Address 3 - Street"
        outsheet[ADDRESS3_CITY + '1'].value                  = "Address 3 - City"
        outsheet[ADDRESS3_POBOX + '1'].value                 = "Address 3 - PO Box"
        outsheet[ADDRESS3_REGION + '1'].value                = "Address 3 - Region"
        outsheet[ADDRESS3_POSTAL_CODE + '1'].value           = "Address 3 - Postal Code"
        outsheet[ADDRESS3_COUNTRY + '1'].value               = "Address 3 - Country"
        outsheet[ADDRESS3_EXTENDED_ADDRESS + '1'].value      = "Address 3 - Extended Address"
        outsheet[ORGANIZATION1_TYPE + '1'].value             = "Organization 1 - Type"
        outsheet[ORGANIZATION1_NAME + '1'].value             = "Organization 1 - Name"
        outsheet[ORGANIZATION1_YOMI_NAME + '1'].value        = "Organization 1 - Yomi Name"
        outsheet[ORGANIZATION1_TITLE + '1'].value            = "Organization 1 - Title"
        outsheet[ORGANIZATION1_DEPARTMENT + '1'].value       = "Organization 1 - Department"
        outsheet[ORGANIZATION1_SYMBOL + '1'].value           = "Organization 1 - Symbol"
        outsheet[ORGANIZATION1_LOCATION + '1'].value         = "Organization 1 - Location"
        outsheet[ORGANIZATION1_JOB_DESCRIPTION + '1'].value  = "Organization 1 - Job Description"
        outsheet[RELATION1_TYPE + '1'].value                 = "Relation 1 - Type"
        outsheet[RELATION1_VALUE + '1'].value                = "Relation 1 - Value"
        outsheet[EXTERNAL_ID1_TYPE + '1'].value              = "External ID 1 - Type"
        outsheet[EXTERNAl_ID1_VALUE + '1'].value             = "External ID 1 - Value"
        outsheet[WEBSITE1_TYPE + '1'].value                  = "Website 1 - Type"
        outsheet[WEBSITE1_VALUE + '1'].value                 = "Website 1 - Value"
        outsheet[CALENDAR_LINK1_TYPE + '1'].value            = "Calendar Link 1 - Type"
        outsheet[CALENDAR_LINK1_VALUE + '1'].value           = "Calendar Link 1 - Value"
        outsheet[JOT1_TYPE + '1'].value                      = "Jot 1 - Type"
        outsheet[JOT1_VALUE + '1'].value                     = "Jot 1 - Value"

        dupesheet[NAME + '1'].value                          = "Name"
        dupesheet[GIVEN_NAME + '1'].value                    = "Given Name"
        dupesheet[ADDITIONAL_NAME + '1'].value               = "Additional Name"
        dupesheet[FAMILY_NAME + '1'].value                   = "Family Name"
        dupesheet[YOMI_NAME + '1'].value                     = "Yomi name"
        dupesheet[GIVEN_NAME_YOMI + '1'].value               = "Given Name Yomi"
        dupesheet[ADDITIONAL_NAME_YOMI + '1'].value          = "Additional Name Yomi"
        dupesheet[FAMILY_NAME_YOMI + '1'].value              = "Family Name Yomi"
        dupesheet[NAME_PREFIX + '1'].value                   = "Name Prefix"
        dupesheet[NAME_SUFFIX + '1'].value                   = "Name Suffix"
        dupesheet[INITIALS + '1'].value                      = "Initials"
        dupesheet[NICKNAME + '1'].value                      = "Nickname"
        dupesheet[SHORT_NAME + '1'].value                    = "Short Name"
        dupesheet[MAIDEN_NAME + '1'].value                   = "Maiden Name"
        dupesheet[BIRTHDAY + '1'].value                      = "Birthday"
        dupesheet[GENDER + '1'].value                        = "Gender"
        dupesheet[LOCATION + '1'].value                      = "Location"
        dupesheet[BILLING_INFORMATION + '1'].value           = "Billing Information"
        dupesheet[DIRECTORY_SERVER + '1'].value              = "Directory Server"
        dupesheet[MILEAGE + '1'].value                       = "Mileage"
        dupesheet[OCCUPATION + '1'].value                    = "Occupation"
        dupesheet[HOBBY + '1'].value                         = "Hobby"
        dupesheet[SENSITIVITY + '1'].value                   = "Sensitivity"
        dupesheet[PRIORITY + '1'].value                      = "Priority"
        dupesheet[SUBJECT + '1'].value                       = "Subject"
        dupesheet[NOTES + '1'].value                         = "Notes"
        dupesheet[GROUP_MEMBERSHIP + '1'].value              = "Group Membership"
        dupesheet[EMAIL1_TYPE + '1'].value                   = "E-mail 1 - Type"
        dupesheet[EMAIL1_VALUE + '1'].value                  = "E-mail 1 - Value"
        dupesheet[EMAIL2_TYPE + '1'].value                   = "E-mail 2 - Type"
        dupesheet[EMAIL2_VALUE + '1'].value                  = "E-mail 2 - Value"
        dupesheet[EMAIL3_TYPE + '1'].value                   = "E-mail 3 - Type"
        dupesheet[EMAIL3_VALUE + '1'].value                  = "E-mail 3 - Value"
        dupesheet[EMAIL4_TYPE + '1'].value                   = "E-mail 4 - Type"
        dupesheet[EMAIL4_VALUE + '1'].value                  = "E-mail 4 - Value"
        dupesheet[EMAIL5_TYPE + '1'].value                   = "E-mail 5 - Type"
        dupesheet[EMAIL5_VALUE + '1'].value                  = "E-mail 5 - Value"
        dupesheet[IM1_TYPE + '1'].value                      = "IM 1 - Type"
        dupesheet[IM1_SERVICE + '1'].value                   = "IM 1 - Service"
        dupesheet[IM1_VALUE + '1'].value                     = "IM 1 - Value"
        dupesheet[PHONE1_TYPE + '1'].value                   = "Phone 1 - Type"
        dupesheet[PHONE1_VALUE + '1'].value                  = "Phone 1 – Value"
        dupesheet[PHONE2_TYPE + '1'].value                   = "Phone 2 – Type"
        dupesheet[PHONE2_VALUE + '1'].value                  = "Phone 2 – Value"
        dupesheet[PHONE3_TYPE + '1'].value                   = "Phone 3 – Type"
        dupesheet[PHONE3_VALUE + '1'].value                  = "Phone 3 – Value"
        dupesheet[PHONE4_TYPE + '1'].value                   = "Phone 4 – Type"
        dupesheet[PHONE4_VALUE + '1'].value                  = "Phone 4 – Value"
        dupesheet[PHONE5_TYPE + '1'].value                   = "Phone 5 – Type"
        dupesheet[PHONE5_VALUE + '1'].value                  = "Phone 5 – Value"
        dupesheet[ADDRESS1_TYPE + '1'].value                 = "Address 1 - Type"
        dupesheet[ADDRESS1_FORMATED + '1'].value             = "Address 1 - Formatted"
        dupesheet[ADDRESS1_STREET + '1'].value               = "Address 1 - Street"
        dupesheet[ADDRESS1_CITY + '1'].value                 = "Address 1 - City"
        dupesheet[ADDRESS1_POBOX + '1'].value                = "Address 1 - PO Box"
        dupesheet[ADDRESS1_REGION + '1'].value               = "Address 1 - Region"
        dupesheet[ADDRESS1_POSTAL_CODE + '1'].value          = "Address 1 - Postal Code"
        dupesheet[ADDRESS1_COUNTRY + '1'].value              = "Address 1 - Country"
        dupesheet[ADDRESS1_EXTENDED_ADDRESS + '1'].value     = "Address 1 - Extended Address"
        dupesheet[ADDRESS2_TYPE + '1'].value                 = "Address 2 - Type"
        dupesheet[ADDRESS2_FORMATED + '1'].value             = "Address 2 - Formatted"
        dupesheet[ADDRESS2_STREET + '1'].value               = "Address 2 - Street"
        dupesheet[ADDRESS2_CITY + '1'].value                 = "Address 2 - City"
        dupesheet[ADDRESS2_POBOX + '1'].value                = "Address 2 - PO Box"
        dupesheet[ADDRESS2_REGION + '1'].value               = "Address 2 - Region"
        dupesheet[ADDRESS2_POSTAL_CODE + '1'].value          = "Address 2 - Postal Code"
        dupesheet[ADDRESS2_COUNTRY + '1'].value              = "Address 2 - Country"
        dupesheet[ADDRESS2_EXTENDED_ADDRESS + '1'].value     = "Address 2 - Extended Address"
        dupesheet[ADDRESS3_TYPE + '1'].value                 = "Address 3 - Type"
        dupesheet[ADDRESS3_FORMATED + '1'].value             = "Address 3 - Formatted"
        dupesheet[ADDRESS3_STREET + '1'].value               = "Address 3 - Street"
        dupesheet[ADDRESS3_CITY + '1'].value                 = "Address 3 - City"
        dupesheet[ADDRESS3_POBOX + '1'].value                = "Address 3 - PO Box"
        dupesheet[ADDRESS3_REGION + '1'].value               = "Address 3 - Region"
        dupesheet[ADDRESS3_POSTAL_CODE + '1'].value          = "Address 3 - Postal Code"
        dupesheet[ADDRESS3_COUNTRY + '1'].value              = "Address 3 - Country"
        dupesheet[ADDRESS3_EXTENDED_ADDRESS + '1'].value     = "Address 3 - Extended Address"
        dupesheet[ORGANIZATION1_TYPE + '1'].value            = "Organization 1 - Type"
        dupesheet[ORGANIZATION1_NAME + '1'].value            = "Organization 1 - Name"
        dupesheet[ORGANIZATION1_YOMI_NAME + '1'].value       = "Organization 1 - Yomi Name"
        dupesheet[ORGANIZATION1_TITLE + '1'].value           = "Organization 1 - Title"
        dupesheet[ORGANIZATION1_DEPARTMENT + '1'].value      = "Organization 1 - Department"
        dupesheet[ORGANIZATION1_SYMBOL + '1'].value          = "Organization 1 - Symbol"
        dupesheet[ORGANIZATION1_LOCATION + '1'].value        = "Organization 1 - Location"
        dupesheet[ORGANIZATION1_JOB_DESCRIPTION + '1'].value = "Organization 1 - Job Description"
        dupesheet[RELATION1_TYPE + '1'].value                = "Relation 1 - Type"
        dupesheet[RELATION1_VALUE + '1'].value               = "Relation 1 - Value"
        dupesheet[EXTERNAL_ID1_TYPE + '1'].value             = "External ID 1 - Type"
        dupesheet[EXTERNAl_ID1_VALUE + '1'].value            = "External ID 1 - Value"
        dupesheet[WEBSITE1_TYPE + '1'].value                 = "Website 1 - Type"
        dupesheet[WEBSITE1_VALUE + '1'].value                = "Website 1 - Value"
        dupesheet[CALENDAR_LINK1_TYPE + '1'].value           = "Calendar Link 1 - Type"
        dupesheet[CALENDAR_LINK1_VALUE + '1'].value          = "Calendar Link 1 - Value"
        dupesheet[JOT1_TYPE + '1'].value                     = "Jot 1 - Type"
        dupesheet[JOT1_VALUE + '1'].value                    = "Jot 1 - Value"
    #}}}

    for row in range (first, last + 1):
        # if the previous value is blank we create the new object and store information in it
        '''
        first_name = str(sheet[first_name_col + str(row)].value)
        if first_name != "":
            standardize_str(first_name)
        '''
        compareCriteria = ""
        for col in cols:
            compareCriteria = compareCriteria + standardize_str(sheet[col + str(row)].value) + " "
        if row == first:
            # compare = first_name + " " + last_name
            compare = compareCriteria
            contact = new_contact_from_sheet(sheet, row)
            count = count + 1
        else:
            # current = first_name + " " + last_name
            current = compareCriteria
            match = edit_distance(compare, current, low_threshold, high_threshold)
            # matchingSuffixes = standardize_str(sheet[SUFFIX + str(row)].value) == standardize_str(sheet[SUFFIX + str(row - 1)].value)
            # combine information and move on
            if match:
                contact = update_contact_from_sheet(sheet, row, contact)
                '''
                - if there is a duplicate
                - store previous item in list in duplicate list
                - store the duplicate in next spot in duplicate list
                '''
                # combine information
                # keep original
                dupes = dupes + 1

                dupeContact = new_contact_from_sheet(sheet, row - 1)

                #{{{
                dupesheet[NAME + str(dupes)].value                          = dupeContact.name 
                dupesheet[GIVEN_NAME + str(dupes)].value                    = dupeContact.given_name 
                dupesheet[ADDITIONAL_NAME + str(dupes)].value               = dupeContact.additional_name 
                dupesheet[FAMILY_NAME + str(dupes)].value                   = dupeContact.family_name 
                dupesheet[YOMI_NAME + str(dupes)].value                     = dupeContact.yomi_name 
                dupesheet[GIVEN_NAME_YOMI + str(dupes)].value               = dupeContact.given_name_yomi 
                dupesheet[ADDITIONAL_NAME_YOMI + str(dupes)].value          = dupeContact.additional_name_yomi 
                dupesheet[FAMILY_NAME_YOMI + str(dupes)].value              = dupeContact.family_name_yomi 
                dupesheet[NAME_PREFIX + str(dupes)].value                   = dupeContact.name_prefix 
                dupesheet[NAME_SUFFIX + str(dupes)].value                   = dupeContact.name_suffix 
                dupesheet[INITIALS + str(dupes)].value                      = dupeContact.initials 
                dupesheet[NICKNAME + str(dupes)].value                      = dupeContact.nickname 
                dupesheet[SHORT_NAME + str(dupes)].value                    = dupeContact.short_name 
                dupesheet[MAIDEN_NAME + str(dupes)].value                   = dupeContact.maiden_name 
                dupesheet[BIRTHDAY + str(dupes)].value                      = dupeContact.birthday 
                dupesheet[GENDER + str(dupes)].value                        = dupeContact.gender 
                dupesheet[LOCATION + str(dupes)].value                      = dupeContact.location 
                dupesheet[BILLING_INFORMATION + str(dupes)].value           = dupeContact.billing_information 
                dupesheet[DIRECTORY_SERVER + str(dupes)].value              = dupeContact.directory_server 
                dupesheet[MILEAGE + str(dupes)].value                       = dupeContact.mileage 
                dupesheet[OCCUPATION + str(dupes)].value                    = dupeContact.occupation 
                dupesheet[HOBBY + str(dupes)].value                         = dupeContact.hobby 
                dupesheet[SENSITIVITY + str(dupes)].value                   = dupeContact.sensitivity 
                dupesheet[PRIORITY + str(dupes)].value                      = dupeContact.priority 
                dupesheet[SUBJECT + str(dupes)].value                       = dupeContact.subject 
                dupesheet[NOTES + str(dupes)].value                         = dupeContact.notes 
                dupesheet[GROUP_MEMBERSHIP + str(dupes)].value              = dupeContact.group_membership 
                dupesheet[EMAIL1_TYPE + str(dupes)].value                   = dupeContact.email1_type 
                dupesheet[EMAIL1_VALUE + str(dupes)].value                  = dupeContact.email1_value 
                dupesheet[EMAIL2_TYPE + str(dupes)].value                   = dupeContact.email2_type 
                dupesheet[EMAIL2_VALUE + str(dupes)].value                  = dupeContact.email2_value 
                dupesheet[EMAIL3_TYPE + str(dupes)].value                   = dupeContact.email3_type 
                dupesheet[EMAIL3_VALUE + str(dupes)].value                  = dupeContact.email3_value 
                dupesheet[EMAIL4_TYPE + str(dupes)].value                   = dupeContact.email4_type 
                dupesheet[EMAIL4_VALUE + str(dupes)].value                  = dupeContact.email4_value 
                dupesheet[EMAIL5_TYPE + str(dupes)].value                   = dupeContact.email5_type 
                dupesheet[EMAIL5_VALUE + str(dupes)].value                  = dupeContact.email5_value 
                dupesheet[IM1_TYPE + str(dupes)].value                      = dupeContact.im1_type 
                dupesheet[IM1_SERVICE + str(dupes)].value                   = dupeContact.im1_service 
                dupesheet[IM1_VALUE + str(dupes)].value                     = dupeContact.im1_value 
                dupesheet[PHONE1_TYPE + str(dupes)].value                   = dupeContact.phone1_type 
                dupesheet[PHONE1_VALUE + str(dupes)].value                  = dupeContact.phone1_value 
                dupesheet[PHONE2_TYPE + str(dupes)].value                   = dupeContact.phone2_type 
                dupesheet[PHONE2_VALUE + str(dupes)].value                  = dupeContact.phone2_value 
                dupesheet[PHONE3_TYPE + str(dupes)].value                   = dupeContact.phone3_type 
                dupesheet[PHONE3_VALUE + str(dupes)].value                  = dupeContact.phone3_value 
                dupesheet[PHONE4_TYPE + str(dupes)].value                   = dupeContact.phone4_type 
                dupesheet[PHONE4_VALUE + str(dupes)].value                  = dupeContact.phone4_value 
                dupesheet[PHONE5_TYPE + str(dupes)].value                   = dupeContact.phone5_type 
                dupesheet[PHONE5_VALUE + str(dupes)].value                  = dupeContact.phone5_value 
                dupesheet[ADDRESS1_TYPE + str(dupes)].value                 = dupeContact.address1_type 
                dupesheet[ADDRESS1_FORMATED + str(dupes)].value             = dupeContact.address1_formated 
                dupesheet[ADDRESS1_STREET + str(dupes)].value               = dupeContact.address1_street 
                dupesheet[ADDRESS1_CITY + str(dupes)].value                 = dupeContact.address1_city 
                dupesheet[ADDRESS1_POBOX + str(dupes)].value                = dupeContact.address1_pobox 
                dupesheet[ADDRESS1_REGION + str(dupes)].value               = dupeContact.address1_region 
                dupesheet[ADDRESS1_POSTAL_CODE + str(dupes)].value          = dupeContact.address1_postal_code 
                dupesheet[ADDRESS1_COUNTRY + str(dupes)].value              = dupeContact.address1_country 
                dupesheet[ADDRESS1_EXTENDED_ADDRESS + str(dupes)].value     = dupeContact.address1_extended_address 
                dupesheet[ADDRESS2_TYPE + str(dupes)].value                 = dupeContact.address2_type 
                dupesheet[ADDRESS2_FORMATED + str(dupes)].value             = dupeContact.address2_formated 
                dupesheet[ADDRESS2_STREET + str(dupes)].value               = dupeContact.address2_street 
                dupesheet[ADDRESS2_CITY + str(dupes)].value                 = dupeContact.address2_city 
                dupesheet[ADDRESS2_POBOX + str(dupes)].value                = dupeContact.address2_pobox 
                dupesheet[ADDRESS2_REGION + str(dupes)].value               = dupeContact.address2_region 
                dupesheet[ADDRESS2_POSTAL_CODE + str(dupes)].value          = dupeContact.address2_postal_code 
                dupesheet[ADDRESS2_COUNTRY + str(dupes)].value              = dupeContact.address2_country 
                dupesheet[ADDRESS2_EXTENDED_ADDRESS + str(dupes)].value     = dupeContact.address2_extended_address 
                dupesheet[ADDRESS3_TYPE + str(dupes)].value                 = dupeContact.address3_type 
                dupesheet[ADDRESS3_FORMATED + str(dupes)].value             = dupeContact.address3_formated 
                dupesheet[ADDRESS3_STREET + str(dupes)].value               = dupeContact.address3_street 
                dupesheet[ADDRESS3_CITY + str(dupes)].value                 = dupeContact.address3_city 
                dupesheet[ADDRESS3_POBOX + str(dupes)].value                = dupeContact.address3_pobox 
                dupesheet[ADDRESS3_REGION + str(dupes)].value               = dupeContact.address3_region 
                dupesheet[ADDRESS3_POSTAL_CODE + str(dupes)].value          = dupeContact.address3_postal_code 
                dupesheet[ADDRESS3_COUNTRY + str(dupes)].value              = dupeContact.address3_country 
                dupesheet[ADDRESS3_EXTENDED_ADDRESS + str(dupes)].value     = dupeContact.address3_extended_address 
                dupesheet[ORGANIZATION1_TYPE + str(dupes)].value            = dupeContact.organization1_type 
                dupesheet[ORGANIZATION1_NAME + str(dupes)].value            = dupeContact.organization1_name 
                dupesheet[ORGANIZATION1_YOMI_NAME + str(dupes)].value       = dupeContact.organization1_yomi_name 
                dupesheet[ORGANIZATION1_TITLE + str(dupes)].value           = dupeContact.organization1_title 
                dupesheet[ORGANIZATION1_DEPARTMENT + str(dupes)].value      = dupeContact.organization1_department 
                dupesheet[ORGANIZATION1_SYMBOL + str(dupes)].value          = dupeContact.organization1_symbol 
                dupesheet[ORGANIZATION1_LOCATION + str(dupes)].value        = dupeContact.organization1_location 
                dupesheet[ORGANIZATION1_JOB_DESCRIPTION + str(dupes)].value = dupeContact.organization1_job_description 
                dupesheet[RELATION1_TYPE + str(dupes)].value                = dupeContact.relation1_type 
                dupesheet[RELATION1_VALUE + str(dupes)].value               = dupeContact.relation1_value 
                dupesheet[EXTERNAL_ID1_TYPE + str(dupes)].value             = dupeContact.external_id1_type 
                dupesheet[EXTERNAl_ID1_VALUE + str(dupes)].value            = dupeContact.external_id1_value 
                dupesheet[WEBSITE1_TYPE + str(dupes)].value                 = dupeContact.website1_type 
                dupesheet[WEBSITE1_VALUE + str(dupes)].value                = dupeContact.website1_value 
                dupesheet[CALENDAR_LINK1_TYPE + str(dupes)].value           = dupeContact.calendar_link1_type 
                dupesheet[CALENDAR_LINK1_VALUE + str(dupes)].value          = dupeContact.calendar_link1_value 
                dupesheet[JOT1_TYPE + str(dupes)].value                     = dupeContact.jot1_type 
                dupesheet[JOT1_VALUE + str(dupes)].value                    = dupeContact.jot1_value 
                dupesheet['C' + str(dupes)].value = row - 1
                #}}}

                # keep duplicate
                dupes = dupes + 1

                dupeContact = new_contact_from_sheet(sheet, row)
                #{{{
                dupesheet[NAME + str(dupes)].value                          = dupeContact.name 
                dupesheet[GIVEN_NAME + str(dupes)].value                    = dupeContact.given_name 
                dupesheet[ADDITIONAL_NAME + str(dupes)].value               = dupeContact.additional_name 
                dupesheet[FAMILY_NAME + str(dupes)].value                   = dupeContact.family_name 
                dupesheet[YOMI_NAME + str(dupes)].value                     = dupeContact.yomi_name 
                dupesheet[GIVEN_NAME_YOMI + str(dupes)].value               = dupeContact.given_name_yomi 
                dupesheet[ADDITIONAL_NAME_YOMI + str(dupes)].value          = dupeContact.additional_name_yomi 
                dupesheet[FAMILY_NAME_YOMI + str(dupes)].value              = dupeContact.family_name_yomi 
                dupesheet[NAME_PREFIX + str(dupes)].value                   = dupeContact.name_prefix 
                dupesheet[NAME_SUFFIX + str(dupes)].value                   = dupeContact.name_suffix 
                dupesheet[INITIALS + str(dupes)].value                      = dupeContact.initials 
                dupesheet[NICKNAME + str(dupes)].value                      = dupeContact.nickname 
                dupesheet[SHORT_NAME + str(dupes)].value                    = dupeContact.short_name 
                dupesheet[MAIDEN_NAME + str(dupes)].value                   = dupeContact.maiden_name 
                dupesheet[BIRTHDAY + str(dupes)].value                      = dupeContact.birthday 
                dupesheet[GENDER + str(dupes)].value                        = dupeContact.gender 
                dupesheet[LOCATION + str(dupes)].value                      = dupeContact.location 
                dupesheet[BILLING_INFORMATION + str(dupes)].value           = dupeContact.billing_information 
                dupesheet[DIRECTORY_SERVER + str(dupes)].value              = dupeContact.directory_server 
                dupesheet[MILEAGE + str(dupes)].value                       = dupeContact.mileage 
                dupesheet[OCCUPATION + str(dupes)].value                    = dupeContact.occupation 
                dupesheet[HOBBY + str(dupes)].value                         = dupeContact.hobby 
                dupesheet[SENSITIVITY + str(dupes)].value                   = dupeContact.sensitivity 
                dupesheet[PRIORITY + str(dupes)].value                      = dupeContact.priority 
                dupesheet[SUBJECT + str(dupes)].value                       = dupeContact.subject 
                dupesheet[NOTES + str(dupes)].value                         = dupeContact.notes 
                dupesheet[GROUP_MEMBERSHIP + str(dupes)].value              = dupeContact.group_membership 
                dupesheet[EMAIL1_TYPE + str(dupes)].value                   = dupeContact.email1_type 
                dupesheet[EMAIL1_VALUE + str(dupes)].value                  = dupeContact.email1_value 
                dupesheet[EMAIL2_TYPE + str(dupes)].value                   = dupeContact.email2_type 
                dupesheet[EMAIL2_VALUE + str(dupes)].value                  = dupeContact.email2_value 
                dupesheet[EMAIL3_TYPE + str(dupes)].value                   = dupeContact.email3_type 
                dupesheet[EMAIL3_VALUE + str(dupes)].value                  = dupeContact.email3_value 
                dupesheet[EMAIL4_TYPE + str(dupes)].value                   = dupeContact.email4_type 
                dupesheet[EMAIL4_VALUE + str(dupes)].value                  = dupeContact.email4_value 
                dupesheet[EMAIL5_TYPE + str(dupes)].value                   = dupeContact.email5_type 
                dupesheet[EMAIL5_VALUE + str(dupes)].value                  = dupeContact.email5_value 
                dupesheet[IM1_TYPE + str(dupes)].value                      = dupeContact.im1_type 
                dupesheet[IM1_SERVICE + str(dupes)].value                   = dupeContact.im1_service 
                dupesheet[IM1_VALUE + str(dupes)].value                     = dupeContact.im1_value 
                dupesheet[PHONE1_TYPE + str(dupes)].value                   = dupeContact.phone1_type 
                dupesheet[PHONE1_VALUE + str(dupes)].value                  = dupeContact.phone1_value 
                dupesheet[PHONE2_TYPE + str(dupes)].value                   = dupeContact.phone2_type 
                dupesheet[PHONE2_VALUE + str(dupes)].value                  = dupeContact.phone2_value 
                dupesheet[PHONE3_TYPE + str(dupes)].value                   = dupeContact.phone3_type 
                dupesheet[PHONE3_VALUE + str(dupes)].value                  = dupeContact.phone3_value 
                dupesheet[PHONE4_TYPE + str(dupes)].value                   = dupeContact.phone4_type 
                dupesheet[PHONE4_VALUE + str(dupes)].value                  = dupeContact.phone4_value 
                dupesheet[PHONE5_TYPE + str(dupes)].value                   = dupeContact.phone5_type 
                dupesheet[PHONE5_VALUE + str(dupes)].value                  = dupeContact.phone5_value 
                dupesheet[ADDRESS1_TYPE + str(dupes)].value                 = dupeContact.address1_type 
                dupesheet[ADDRESS1_FORMATED + str(dupes)].value             = dupeContact.address1_formated 
                dupesheet[ADDRESS1_STREET + str(dupes)].value               = dupeContact.address1_street 
                dupesheet[ADDRESS1_CITY + str(dupes)].value                 = dupeContact.address1_city 
                dupesheet[ADDRESS1_POBOX + str(dupes)].value                = dupeContact.address1_pobox 
                dupesheet[ADDRESS1_REGION + str(dupes)].value               = dupeContact.address1_region 
                dupesheet[ADDRESS1_POSTAL_CODE + str(dupes)].value          = dupeContact.address1_postal_code 
                dupesheet[ADDRESS1_COUNTRY + str(dupes)].value              = dupeContact.address1_country 
                dupesheet[ADDRESS1_EXTENDED_ADDRESS + str(dupes)].value     = dupeContact.address1_extended_address 
                dupesheet[ADDRESS2_TYPE + str(dupes)].value                 = dupeContact.address2_type 
                dupesheet[ADDRESS2_FORMATED + str(dupes)].value             = dupeContact.address2_formated 
                dupesheet[ADDRESS2_STREET + str(dupes)].value               = dupeContact.address2_street 
                dupesheet[ADDRESS2_CITY + str(dupes)].value                 = dupeContact.address2_city 
                dupesheet[ADDRESS2_POBOX + str(dupes)].value                = dupeContact.address2_pobox 
                dupesheet[ADDRESS2_REGION + str(dupes)].value               = dupeContact.address2_region 
                dupesheet[ADDRESS2_POSTAL_CODE + str(dupes)].value          = dupeContact.address2_postal_code 
                dupesheet[ADDRESS2_COUNTRY + str(dupes)].value              = dupeContact.address2_country 
                dupesheet[ADDRESS2_EXTENDED_ADDRESS + str(dupes)].value     = dupeContact.address2_extended_address 
                dupesheet[ADDRESS3_TYPE + str(dupes)].value                 = dupeContact.address3_type 
                dupesheet[ADDRESS3_FORMATED + str(dupes)].value             = dupeContact.address3_formated 
                dupesheet[ADDRESS3_STREET + str(dupes)].value               = dupeContact.address3_street 
                dupesheet[ADDRESS3_CITY + str(dupes)].value                 = dupeContact.address3_city 
                dupesheet[ADDRESS3_POBOX + str(dupes)].value                = dupeContact.address3_pobox 
                dupesheet[ADDRESS3_REGION + str(dupes)].value               = dupeContact.address3_region 
                dupesheet[ADDRESS3_POSTAL_CODE + str(dupes)].value          = dupeContact.address3_postal_code 
                dupesheet[ADDRESS3_COUNTRY + str(dupes)].value              = dupeContact.address3_country 
                dupesheet[ADDRESS3_EXTENDED_ADDRESS + str(dupes)].value     = dupeContact.address3_extended_address 
                dupesheet[ORGANIZATION1_TYPE + str(dupes)].value            = dupeContact.organization1_type 
                dupesheet[ORGANIZATION1_NAME + str(dupes)].value            = dupeContact.organization1_name 
                dupesheet[ORGANIZATION1_YOMI_NAME + str(dupes)].value       = dupeContact.organization1_yomi_name 
                dupesheet[ORGANIZATION1_TITLE + str(dupes)].value           = dupeContact.organization1_title 
                dupesheet[ORGANIZATION1_DEPARTMENT + str(dupes)].value      = dupeContact.organization1_department 
                dupesheet[ORGANIZATION1_SYMBOL + str(dupes)].value          = dupeContact.organization1_symbol 
                dupesheet[ORGANIZATION1_LOCATION + str(dupes)].value        = dupeContact.organization1_location 
                dupesheet[ORGANIZATION1_JOB_DESCRIPTION + str(dupes)].value = dupeContact.organization1_job_description 
                dupesheet[RELATION1_TYPE + str(dupes)].value                = dupeContact.relation1_type 
                dupesheet[RELATION1_VALUE + str(dupes)].value               = dupeContact.relation1_value 
                dupesheet[EXTERNAL_ID1_TYPE + str(dupes)].value             = dupeContact.external_id1_type 
                dupesheet[EXTERNAl_ID1_VALUE + str(dupes)].value            = dupeContact.external_id1_value 
                dupesheet[WEBSITE1_TYPE + str(dupes)].value                 = dupeContact.website1_type 
                dupesheet[WEBSITE1_VALUE + str(dupes)].value                = dupeContact.website1_value 
                dupesheet[CALENDAR_LINK1_TYPE + str(dupes)].value           = dupeContact.calendar_link1_type 
                dupesheet[CALENDAR_LINK1_VALUE + str(dupes)].value          = dupeContact.calendar_link1_value 
                dupesheet[JOT1_TYPE + str(dupes)].value                     = dupeContact.jot1_type 
                dupesheet[JOT1_VALUE + str(dupes)].value                    = dupeContact.jot1_value 
                dupesheet['C' + str(dupes)].value = row
                #}}}

                dupes = dupes + 1

                # save the combined contact
                #{{{
                dupesheet[NAME + str(dupes)].value                          = contact.name 
                dupesheet[GIVEN_NAME + str(dupes)].value                    = contact.given_name 
                dupesheet[ADDITIONAL_NAME + str(dupes)].value               = contact.additional_name 
                dupesheet[FAMILY_NAME + str(dupes)].value                   = contact.family_name 
                dupesheet[YOMI_NAME + str(dupes)].value                     = contact.yomi_name 
                dupesheet[GIVEN_NAME_YOMI + str(dupes)].value               = contact.given_name_yomi 
                dupesheet[ADDITIONAL_NAME_YOMI + str(dupes)].value          = contact.additional_name_yomi 
                dupesheet[FAMILY_NAME_YOMI + str(dupes)].value              = contact.family_name_yomi 
                dupesheet[NAME_PREFIX + str(dupes)].value                   = contact.name_prefix 
                dupesheet[NAME_SUFFIX + str(dupes)].value                   = contact.name_suffix 
                dupesheet[INITIALS + str(dupes)].value                      = contact.initials 
                dupesheet[NICKNAME + str(dupes)].value                      = contact.nickname 
                dupesheet[SHORT_NAME + str(dupes)].value                    = contact.short_name 
                dupesheet[MAIDEN_NAME + str(dupes)].value                   = contact.maiden_name 
                dupesheet[BIRTHDAY + str(dupes)].value                      = contact.birthday 
                dupesheet[GENDER + str(dupes)].value                        = contact.gender 
                dupesheet[LOCATION + str(dupes)].value                      = contact.location 
                dupesheet[BILLING_INFORMATION + str(dupes)].value           = contact.billing_information 
                dupesheet[DIRECTORY_SERVER + str(dupes)].value              = contact.directory_server 
                dupesheet[MILEAGE + str(dupes)].value                       = contact.mileage 
                dupesheet[OCCUPATION + str(dupes)].value                    = contact.occupation 
                dupesheet[HOBBY + str(dupes)].value                         = contact.hobby 
                dupesheet[SENSITIVITY + str(dupes)].value                   = contact.sensitivity 
                dupesheet[PRIORITY + str(dupes)].value                      = contact.priority 
                dupesheet[SUBJECT + str(dupes)].value                       = contact.subject 
                dupesheet[NOTES + str(dupes)].value                         = contact.notes 
                dupesheet[GROUP_MEMBERSHIP + str(dupes)].value              = contact.group_membership 
                dupesheet[EMAIL1_TYPE + str(dupes)].value                   = contact.email1_type 
                dupesheet[EMAIL1_VALUE + str(dupes)].value                  = contact.email1_value 
                dupesheet[EMAIL2_TYPE + str(dupes)].value                   = contact.email2_type 
                dupesheet[EMAIL2_VALUE + str(dupes)].value                  = contact.email2_value 
                dupesheet[EMAIL3_TYPE + str(dupes)].value                   = contact.email3_type 
                dupesheet[EMAIL3_VALUE + str(dupes)].value                  = contact.email3_value 
                dupesheet[EMAIL4_TYPE + str(dupes)].value                   = contact.email4_type 
                dupesheet[EMAIL4_VALUE + str(dupes)].value                  = contact.email4_value 
                dupesheet[EMAIL5_TYPE + str(dupes)].value                   = contact.email5_type 
                dupesheet[EMAIL5_VALUE + str(dupes)].value                  = contact.email5_value 
                dupesheet[IM1_TYPE + str(dupes)].value                      = contact.im1_type 
                dupesheet[IM1_SERVICE + str(dupes)].value                   = contact.im1_service 
                dupesheet[IM1_VALUE + str(dupes)].value                     = contact.im1_value 
                dupesheet[PHONE1_TYPE + str(dupes)].value                   = contact.phone1_type 
                dupesheet[PHONE1_VALUE + str(dupes)].value                  = contact.phone1_value 
                dupesheet[PHONE2_TYPE + str(dupes)].value                   = contact.phone2_type 
                dupesheet[PHONE2_VALUE + str(dupes)].value                  = contact.phone2_value 
                dupesheet[PHONE3_TYPE + str(dupes)].value                   = contact.phone3_type 
                dupesheet[PHONE3_VALUE + str(dupes)].value                  = contact.phone3_value 
                dupesheet[PHONE4_TYPE + str(dupes)].value                   = contact.phone4_type 
                dupesheet[PHONE4_VALUE + str(dupes)].value                  = contact.phone4_value 
                dupesheet[PHONE5_TYPE + str(dupes)].value                   = contact.phone5_type 
                dupesheet[PHONE5_VALUE + str(dupes)].value                  = contact.phone5_value 
                dupesheet[ADDRESS1_TYPE + str(dupes)].value                 = contact.address1_type 
                dupesheet[ADDRESS1_FORMATED + str(dupes)].value             = contact.address1_formated 
                dupesheet[ADDRESS1_STREET + str(dupes)].value               = contact.address1_street 
                dupesheet[ADDRESS1_CITY + str(dupes)].value                 = contact.address1_city 
                dupesheet[ADDRESS1_POBOX + str(dupes)].value                = contact.address1_pobox 
                dupesheet[ADDRESS1_REGION + str(dupes)].value               = contact.address1_region 
                dupesheet[ADDRESS1_POSTAL_CODE + str(dupes)].value          = contact.address1_postal_code 
                dupesheet[ADDRESS1_COUNTRY + str(dupes)].value              = contact.address1_country 
                dupesheet[ADDRESS1_EXTENDED_ADDRESS + str(dupes)].value     = contact.address1_extended_address 
                dupesheet[ADDRESS2_TYPE + str(dupes)].value                 = contact.address2_type 
                dupesheet[ADDRESS2_FORMATED + str(dupes)].value             = contact.address2_formated 
                dupesheet[ADDRESS2_STREET + str(dupes)].value               = contact.address2_street 
                dupesheet[ADDRESS2_CITY + str(dupes)].value                 = contact.address2_city 
                dupesheet[ADDRESS2_POBOX + str(dupes)].value                = contact.address2_pobox 
                dupesheet[ADDRESS2_REGION + str(dupes)].value               = contact.address2_region 
                dupesheet[ADDRESS2_POSTAL_CODE + str(dupes)].value          = contact.address2_postal_code 
                dupesheet[ADDRESS2_COUNTRY + str(dupes)].value              = contact.address2_country 
                dupesheet[ADDRESS2_EXTENDED_ADDRESS + str(dupes)].value     = contact.address2_extended_address 
                dupesheet[ADDRESS3_TYPE + str(dupes)].value                 = contact.address3_type 
                dupesheet[ADDRESS3_FORMATED + str(dupes)].value             = contact.address3_formated 
                dupesheet[ADDRESS3_STREET + str(dupes)].value               = contact.address3_street 
                dupesheet[ADDRESS3_CITY + str(dupes)].value                 = contact.address3_city 
                dupesheet[ADDRESS3_POBOX + str(dupes)].value                = contact.address3_pobox 
                dupesheet[ADDRESS3_REGION + str(dupes)].value               = contact.address3_region 
                dupesheet[ADDRESS3_POSTAL_CODE + str(dupes)].value          = contact.address3_postal_code 
                dupesheet[ADDRESS3_COUNTRY + str(dupes)].value              = contact.address3_country 
                dupesheet[ADDRESS3_EXTENDED_ADDRESS + str(dupes)].value     = contact.address3_extended_address 
                dupesheet[ORGANIZATION1_TYPE + str(dupes)].value            = contact.organization1_type 
                dupesheet[ORGANIZATION1_NAME + str(dupes)].value            = contact.organization1_name 
                dupesheet[ORGANIZATION1_YOMI_NAME + str(dupes)].value       = contact.organization1_yomi_name 
                dupesheet[ORGANIZATION1_TITLE + str(dupes)].value           = contact.organization1_title 
                dupesheet[ORGANIZATION1_DEPARTMENT + str(dupes)].value      = contact.organization1_department 
                dupesheet[ORGANIZATION1_SYMBOL + str(dupes)].value          = contact.organization1_symbol 
                dupesheet[ORGANIZATION1_LOCATION + str(dupes)].value        = contact.organization1_location 
                dupesheet[ORGANIZATION1_JOB_DESCRIPTION + str(dupes)].value = contact.organization1_job_description 
                dupesheet[RELATION1_TYPE + str(dupes)].value                = contact.relation1_type 
                dupesheet[RELATION1_VALUE + str(dupes)].value               = contact.relation1_value 
                dupesheet[EXTERNAL_ID1_TYPE + str(dupes)].value             = contact.external_id1_type 
                dupesheet[EXTERNAl_ID1_VALUE + str(dupes)].value            = contact.external_id1_value 
                dupesheet[WEBSITE1_TYPE + str(dupes)].value                 = contact.website1_type 
                dupesheet[WEBSITE1_VALUE + str(dupes)].value                = contact.website1_value 
                dupesheet[CALENDAR_LINK1_TYPE + str(dupes)].value           = contact.calendar_link1_type 
                dupesheet[CALENDAR_LINK1_VALUE + str(dupes)].value          = contact.calendar_link1_value 
                dupesheet[JOT1_TYPE + str(dupes)].value                     = contact.jot1_type 
                dupesheet[JOT1_VALUE + str(dupes)].value                    = contact.jot1_value 
                #}}}

                # create a blank space
                dupes = dupes + 1
            # store the information and create a new contact
            else:
                # store information
                #{{{
                outsheet[NAME + str(count)].value                          = contact.name 
                outsheet[GIVEN_NAME + str(count)].value                    = contact.given_name 
                outsheet[ADDITIONAL_NAME + str(count)].value               = contact.additional_name 
                outsheet[FAMILY_NAME + str(count)].value                   = contact.family_name 
                outsheet[YOMI_NAME + str(count)].value                     = contact.yomi_name 
                outsheet[GIVEN_NAME_YOMI + str(count)].value               = contact.given_name_yomi 
                outsheet[ADDITIONAL_NAME_YOMI + str(count)].value          = contact.additional_name_yomi 
                outsheet[FAMILY_NAME_YOMI + str(count)].value              = contact.family_name_yomi 
                outsheet[NAME_PREFIX + str(count)].value                   = contact.name_prefix 
                outsheet[NAME_SUFFIX + str(count)].value                   = contact.name_suffix 
                outsheet[INITIALS + str(count)].value                      = contact.initials 
                outsheet[NICKNAME + str(count)].value                      = contact.nickname 
                outsheet[SHORT_NAME + str(count)].value                    = contact.short_name 
                outsheet[MAIDEN_NAME + str(count)].value                   = contact.maiden_name 
                outsheet[BIRTHDAY + str(count)].value                      = contact.birthday 
                outsheet[GENDER + str(count)].value                        = contact.gender 
                outsheet[LOCATION + str(count)].value                      = contact.location 
                outsheet[BILLING_INFORMATION + str(count)].value           = contact.billing_information 
                outsheet[DIRECTORY_SERVER + str(count)].value              = contact.directory_server 
                outsheet[MILEAGE + str(count)].value                       = contact.mileage 
                outsheet[OCCUPATION + str(count)].value                    = contact.occupation 
                outsheet[HOBBY + str(count)].value                         = contact.hobby 
                outsheet[SENSITIVITY + str(count)].value                   = contact.sensitivity 
                outsheet[PRIORITY + str(count)].value                      = contact.priority 
                outsheet[SUBJECT + str(count)].value                       = contact.subject 
                outsheet[NOTES + str(count)].value                         = contact.notes 
                outsheet[GROUP_MEMBERSHIP + str(count)].value              = contact.group_membership 
                outsheet[EMAIL1_TYPE + str(count)].value                   = contact.email1_type 
                outsheet[EMAIL1_VALUE + str(count)].value                  = contact.email1_value 
                outsheet[EMAIL2_TYPE + str(count)].value                   = contact.email2_type 
                outsheet[EMAIL2_VALUE + str(count)].value                  = contact.email2_value 
                outsheet[EMAIL3_TYPE + str(count)].value                   = contact.email3_type 
                outsheet[EMAIL3_VALUE + str(count)].value                  = contact.email3_value 
                outsheet[EMAIL4_TYPE + str(count)].value                   = contact.email4_type 
                outsheet[EMAIL4_VALUE + str(count)].value                  = contact.email4_value 
                outsheet[EMAIL5_TYPE + str(count)].value                   = contact.email5_type 
                outsheet[EMAIL5_VALUE + str(count)].value                  = contact.email5_value 
                outsheet[IM1_TYPE + str(count)].value                      = contact.im1_type 
                outsheet[IM1_SERVICE + str(count)].value                   = contact.im1_service 
                outsheet[IM1_VALUE + str(count)].value                     = contact.im1_value 
                outsheet[PHONE1_TYPE + str(count)].value                   = contact.phone1_type 
                outsheet[PHONE1_VALUE + str(count)].value                  = contact.phone1_value 
                outsheet[PHONE2_TYPE + str(count)].value                   = contact.phone2_type 
                outsheet[PHONE2_VALUE + str(count)].value                  = contact.phone2_value 
                outsheet[PHONE3_TYPE + str(count)].value                   = contact.phone3_type 
                outsheet[PHONE3_VALUE + str(count)].value                  = contact.phone3_value 
                outsheet[PHONE4_TYPE + str(count)].value                   = contact.phone4_type 
                outsheet[PHONE4_VALUE + str(count)].value                  = contact.phone4_value 
                outsheet[PHONE5_TYPE + str(count)].value                   = contact.phone5_type 
                outsheet[PHONE5_VALUE + str(count)].value                  = contact.phone5_value 
                outsheet[ADDRESS1_TYPE + str(count)].value                 = contact.address1_type 
                outsheet[ADDRESS1_FORMATED + str(count)].value             = contact.address1_formated 
                outsheet[ADDRESS1_STREET + str(count)].value               = contact.address1_street 
                outsheet[ADDRESS1_CITY + str(count)].value                 = contact.address1_city 
                outsheet[ADDRESS1_POBOX + str(count)].value                = contact.address1_pobox 
                outsheet[ADDRESS1_REGION + str(count)].value               = contact.address1_region 
                outsheet[ADDRESS1_POSTAL_CODE + str(count)].value          = contact.address1_postal_code 
                outsheet[ADDRESS1_COUNTRY + str(count)].value              = contact.address1_country 
                outsheet[ADDRESS1_EXTENDED_ADDRESS + str(count)].value     = contact.address1_extended_address 
                outsheet[ADDRESS2_TYPE + str(count)].value                 = contact.address2_type 
                outsheet[ADDRESS2_FORMATED + str(count)].value             = contact.address2_formated 
                outsheet[ADDRESS2_STREET + str(count)].value               = contact.address2_street 
                outsheet[ADDRESS2_CITY + str(count)].value                 = contact.address2_city 
                outsheet[ADDRESS2_POBOX + str(count)].value                = contact.address2_pobox 
                outsheet[ADDRESS2_REGION + str(count)].value               = contact.address2_region 
                outsheet[ADDRESS2_POSTAL_CODE + str(count)].value          = contact.address2_postal_code 
                outsheet[ADDRESS2_COUNTRY + str(count)].value              = contact.address2_country 
                outsheet[ADDRESS2_EXTENDED_ADDRESS + str(count)].value     = contact.address2_extended_address 
                outsheet[ADDRESS3_TYPE + str(count)].value                 = contact.address3_type 
                outsheet[ADDRESS3_FORMATED + str(count)].value             = contact.address3_formated 
                outsheet[ADDRESS3_STREET + str(count)].value               = contact.address3_street 
                outsheet[ADDRESS3_CITY + str(count)].value                 = contact.address3_city 
                outsheet[ADDRESS3_POBOX + str(count)].value                = contact.address3_pobox 
                outsheet[ADDRESS3_REGION + str(count)].value               = contact.address3_region 
                outsheet[ADDRESS3_POSTAL_CODE + str(count)].value          = contact.address3_postal_code 
                outsheet[ADDRESS3_COUNTRY + str(count)].value              = contact.address3_country 
                outsheet[ADDRESS3_EXTENDED_ADDRESS + str(count)].value     = contact.address3_extended_address 
                outsheet[ORGANIZATION1_TYPE + str(count)].value            = contact.organization1_type 
                outsheet[ORGANIZATION1_NAME + str(count)].value            = contact.organization1_name 
                outsheet[ORGANIZATION1_YOMI_NAME + str(count)].value       = contact.organization1_yomi_name 
                outsheet[ORGANIZATION1_TITLE + str(count)].value           = contact.organization1_title 
                outsheet[ORGANIZATION1_DEPARTMENT + str(count)].value      = contact.organization1_department 
                outsheet[ORGANIZATION1_SYMBOL + str(count)].value          = contact.organization1_symbol 
                outsheet[ORGANIZATION1_LOCATION + str(count)].value        = contact.organization1_location 
                outsheet[ORGANIZATION1_JOB_DESCRIPTION + str(count)].value = contact.organization1_job_description 
                outsheet[RELATION1_TYPE + str(count)].value                = contact.relation1_type 
                outsheet[RELATION1_VALUE + str(count)].value               = contact.relation1_value 
                outsheet[EXTERNAL_ID1_TYPE + str(count)].value             = contact.external_id1_type 
                outsheet[EXTERNAl_ID1_VALUE + str(count)].value            = contact.external_id1_value 
                outsheet[WEBSITE1_TYPE + str(count)].value                 = contact.website1_type 
                outsheet[WEBSITE1_VALUE + str(count)].value                = contact.website1_value 
                outsheet[CALENDAR_LINK1_TYPE + str(count)].value           = contact.calendar_link1_type 
                outsheet[CALENDAR_LINK1_VALUE + str(count)].value          = contact.calendar_link1_value 
                outsheet[JOT1_TYPE + str(count)].value                     = contact.jot1_type 
                outsheet[JOT1_VALUE + str(count)].value                    = contact.jot1_value 
                #}}}
                # reset compare value
                # compare = first_name + " " + last_name
                compare = compareCriteria
                # create a new contact
                contact = new_contact_from_sheet(sheet, row)
                count = count + 1

    if printing:
        print()
        print("Out of " + str(1 + last - first) + " companies " + str(count) + " were unique contacts")
        print("Saving...")

    out.save("purged" + str(cols) + ".xlsx")
    dupe.save("duplicates" + str(cols) + ".xlsx")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

    if printing:
        print()
        print("Done!")
#}}}

# Fix Country Names
# fix_country_names(sheet)
# {{{
def fix_country_names(fileName, sheetName):
    if printing:
        print("Opening...")
    wb = openpyxl.load_workbook(fileName + ".xlsx")
    sheet = wb[sheetName]

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')

    compare = ""
    current = ""
    count = 1
    first = 2
    last = sheet.max_row

    for row in range (first, last + 1):
        # if the previous value is blank we create the new object and store information in it
        country = str(sheet[BUSINESS_COUNTRY + str(row)].value)
        if country == 'None':
            break
        if row == first:
            compare = country
            count = count + 1
        else:
            current = country
            match = edit_distance(compare, current, low_threshold, high_threshold)
            # combine information and move on
            if match:
                if printing:
                    print("Now changing row " + str(row))
                sheet[BUSINESS_COUNTRY + str(row)].value = compare
            else:
                # reset compare value
                compare = country
                count = count + 1

    if printing:
        print("Saving...")
    wb.save("newCombined.xlsx")

    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()
#}}}

# combine_contacts(2, 100)
# combine_contacts("google", "contacts", 'linkedIn', 'contacts')
# remove_duplicate_contacts("combined", 80, 'A', 'AC')
# combine_with_CMA("CMAShipping", "Attendees", "gli", "contacts")
# standardize_USA(sys.argv)
# remove_duplicate_contacts(sys.argv)
# fix_states(sys.argv)
format_all_numbers(sys.argv)
