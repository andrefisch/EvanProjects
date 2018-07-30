from tqdm import tqdm
import openpyxl
import os
import pygame
import sys
import time

# GETS ALL COLUMNS FROM THIS ROW
# print(sheet[1])

# First file: transfer to
# Second file: transfer from
# First file should have as many or more fields as second file
def mergeByColumnName():
    # Uses sys.argv to pass in arguments
    args = sys.argv[1:]
    firstFile = args[0]
    secondFile = args[1]

    # Open an existing excel file
    wb1 = openpyxl.load_workbook(firstFile)
    sheet1 = wb1.worksheets[0]
    wb2 = openpyxl.load_workbook(secondFile)
    sheet2 = wb2.worksheets[0]

    #################
    # DO STUFF HERE #
    #################
    '''
    - find max_row of sheet1
    - loop through sheet2
    - match up columns and add data
      - put columns into dict
    '''
    dicty1 = {}
    dicty2 = {}
    for i in range(0, len(sheet1['1'])):
        if sheet1['1'][i].value == None:
            break
        dicty1[sheet1['1'][i].value.lower()] = sheet1['1'][i].column
    for i in range(0, len(sheet2['1'])):
        if sheet2['1'][i].value == None:
            break
        dicty2[sheet2['1'][i].value.lower()] = sheet2['1'][i].column

    k = {}
    for key in dicty2:
        if key in dicty2 and key in dicty1:
            k[dicty2[key]] = dicty1[key]


    print(len(dicty1), " columns detected in", firstFile)
    print(len(dicty2), " columns detected in", secondFile)
    print(len(k), " common columns being processed")

    '''
    # find fields that are not being processed
    listy = []
    for key in k:
        if key not in dicty2:
            listy.append(k)

    print("Fields that did not match:")
    for item in listy:
        print(item)
    '''

    start = 2
    end    = sheet2.max_row + 1
    row    = sheet1.max_row + 1

    '''
    - load in a row from second sheet
    - match up each column from second spreadsheet to first spreadsheet
    - copy information
    '''
    for i in tqdm(range(start, end)):
        values = sheet2[str(i)]
        for j in range(0, len(sheet2[str(i)])):
            if values[j].column in k:
                sheet1[k[values[j].column] + str(row)] = values[j].value
            else:
                continue
            # print('putting ' + str(values[j].value) + ' in column ' + k[values[j].column])
        row = row + 1


    # sheet['A1'].value = "DATA GOES HERE"

    wb1.save("merge.xlsx")

    # LMK when the script is done
    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

mergeByColumnName()
