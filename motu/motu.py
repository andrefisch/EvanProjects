# which column is the unique identifier (primary key)
# in event of two names having different data in the same column (collision) should I take most recent?
# you have some accounts which only seem to have multiple rows because they are present in more than one industry or segment. do you want to have all of that information represented in the same cell?

# want to eliminate duplicate companies but keep all information
# eliminate duplicate companies and people

# for first part use company as unique identifier
# for second part use name as unique identifier

from datetime import datetime
import openpyxl
import pygame
import string
import time

saving = True
printing = True
testing = True
low_threshold = 75
high_threshold = 100
min_word_len = 5
bad_words = ['korea', 'global', 'shipping', 'golden', 'petra', 'shanghai', 'qingdao', 'marita', 'univer', 'royal', 'prima', 'universal']

APPELLATION = 'A'
PRIMARY_CONTACT = 'B'
COMPANY_NAME = 'C'
INDUSTRY = 'D'
SEGMENT = 'E'
SOURCE = 'F'
LAST_UPDATED = 'G'
TITLE = 'H'
CITY = 'I'
STATE = 'J'
ZIP_CODE = 'K'
REGION = 'L'
MAIN_PHONE = 'M'
MOBILE_PHONE = 'N'
ADDRESS1 = 'O'
COUNTRY = 'P'
A1_STREET = 'Q'
A1_STREET2 = 'R'
A1_STREET3 = 'S'
SECONDARY_ADDRESS = 'T'
EMAIL = 'U'
NOTES_PARENTCO = 'V'
FAX = 'W'
WEBSITE = 'X'
COMMENTS = 'Y'


# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# Combine with slash
# special_combine(word1, word2)
# {{{
def special_combine(word1, word2):
    if word1 == "" or word1 == None:
        return word2
    elif word2 == "" or word2 == None:
        return word1
    elif word1 in word2:
        return word2
    elif word2 in word1:
        return word1
    else:
        return word1 + "/" + word2
#}}}

# Edit Distance Function
# edit_distance(word1, word2, low_threshold, high_threshold)
#{{{
# modified edit distance algorithm:
# - see if string one is a substring of the next string
#   - if it is there is a match
#   - it not look at edit distance for substring and substring of equal length
#     - if it is above a certain percentage or below a certain edit number we can safely assume they are a match. keep that string and try with next one

def edit_distance(word1, word2, low_threshold, high_threshold):
    if printing:
        print("Looking at '" + word1 + "' and '" + word2 + "'")
    word1 = word1.lower()
    word2 = word2.lower()
    len_1 = len(word1)
    len_2 = len(word2)
    edit_distance, percent_match = 0, 0

    # make sure shorter word is first
    if len_1 > len_2:
        word1, word2 = word2, word1
        len_1 = len_2

    if (len_1 >= min_word_len and word1 not in bad_words):
        if (word1 in word2):
            if printing:
                print ("Edit distance: 0")
                print ("Percent Match: 100")
            edit_distance, percent_match = 0, 100
        else:
            # shorten longer word to length of first word
            word2 = word2[0:len_1]
            # the matrix whose last element -> edit distance
            x = [[0] * (len_1 + 1) for _ in range(len_1 + 1)]

            # initialization of base case values
            for i in range(0, len_1 + 1): 
                x[i][0] = i
            for j in range(0, len_1 + 1):
                x[0][j] = j
            for i in range (1, len_1 + 1):
                for j in range(1, len_1 + 1):
                    if word1[i - 1] == word2[j - 1]:
                        x[i][j] = x[i - 1][j - 1] 
                    else:
                        x[i][j]= min(x[i][j - 1], x[i - 1][j], x[i - 1][j - 1]) + 1
            edit_distance = x[i][j]
            percent_match = ((len_1 - edit_distance) / len_1) * 100
            if printing:
                print ("Edit distance " + str(x[i][j]))
                print ("Percent match: " + "%.2f" % percent_match)
    if percent_match > low_threshold and percent_match <= high_threshold:
        print("MATCH!")
        return True
    else:
        return False
#}}}

# Company object with 29 fields
# Company
#{{{
class Company(object):
    appellation =       ''
    primary_contact =   ''
    company_name =      ''
    industry =          ''
    segment =           ''
    source =            ''
    last_updated =      ''
    title =             ''
    city =              ''
    state =             ''
    zip_code =          ''
    region =            ''
    main_phone =        ''
    mobile_phone =      ''
    address1 =          ''
    country =           ''
    a1_street =         ''
    a1_street2 =        ''
    a1_street3 =        ''
    secondary_address = ''
    email =             ''
    notes_parentco =    ''
    fax =               ''
    website =           ''
    comments =          ''

    def __init__(self):
        primary_contact = "-"

#}}}

# Create a new company from sheet information
# new_company_from_sheet(sheet, row)
#{{{
def new_company_from_sheet(sheet, row):
    company = Company()
    if sheet[   APPELLATION        + str(row)].value != None:
        company.appellation        = sheet[APPELLATION        + str(row)].value 
    if sheet[   PRIMARY_CONTACT    + str(row)].value != None:
        company.primary_contact    = sheet[PRIMARY_CONTACT    + str(row)].value
    if sheet[   COMPANY_NAME       + str(row)].value != None:
        company.company_name       = sheet[COMPANY_NAME       + str(row)].value
    if sheet[   INDUSTRY           + str(row)].value != None:
        company.industry           = sheet[INDUSTRY           + str(row)].value  
    if sheet[   SEGMENT            + str(row)].value != None:
        company.segment            = sheet[SEGMENT            + str(row)].value  
    if sheet[   SOURCE             + str(row)].value != None:
        company.source             = sheet[SOURCE             + str(row)].value  
    if sheet[   LAST_UPDATED       + str(row)].value != None:
        company.last_updated       = sheet[LAST_UPDATED       + str(row)].value  
    if sheet[   TITLE              + str(row)].value != None:
        company.title              = sheet[TITLE              + str(row)].value  
    if sheet[   CITY               + str(row)].value != None:
        company.city               = sheet[CITY               + str(row)].value  
    if sheet[   STATE              + str(row)].value != None:
        company.state              = sheet[STATE              + str(row)].value  
    if sheet[   ZIP_CODE           + str(row)].value != None:
        company.zip_code           = sheet[ZIP_CODE           + str(row)].value  
    if sheet[   REGION             + str(row)].value != None:
        company.region             = sheet[REGION             + str(row)].value  
    if sheet[   MAIN_PHONE         + str(row)].value != None:
        company.main_phone         = sheet[MAIN_PHONE         + str(row)].value  
    if sheet[   MOBILE_PHONE       + str(row)].value != None:
        company.mobile_phone       = sheet[MOBILE_PHONE       + str(row)].value  
    if sheet[   ADDRESS1           + str(row)].value != None:
        company.address1           = sheet[ADDRESS1           + str(row)].value  
    if sheet[   COUNTRY            + str(row)].value != None:
        company.country            = sheet[COUNTRY            + str(row)].value  
    if sheet[   A1_STREET          + str(row)].value != None:
        company.a1_street          = sheet[A1_STREET          + str(row)].value  
    if sheet[   A1_STREET2         + str(row)].value != None:
        company.a1_street2         = sheet[A1_STREET2         + str(row)].value  
    if sheet[   A1_STREET3         + str(row)].value != None:
        company.a1_street3         = sheet[A1_STREET3         + str(row)].value  
    if sheet[   SECONDARY_ADDRESS  + str(row)].value != None:
        company.secondary_address  = sheet[SECONDARY_ADDRESS  + str(row)].value  
    if sheet[   EMAIL              + str(row)].value != None:
        company.email              = sheet[EMAIL              + str(row)].value  
    if sheet[   NOTES_PARENTCO     + str(row)].value != None:
        company.notes_parentco     = sheet[NOTES_PARENTCO     + str(row)].value  
    if sheet[   FAX                + str(row)].value != None:
        company.fax                = sheet[FAX                + str(row)].value  
    if sheet[   WEBSITE            + str(row)].value != None:
        company.website            = sheet[WEBSITE            + str(row)].value  
    if sheet[   COMMENTS           + str(row)].value != None:
        company.comments           = sheet[COMMENTS           + str(row)].value  
    
    return company
#}}}

# Update Company from sheet information
# update_company_from_sheet(sheet, row, company)
#{{{
def update_company_from_sheet(sheet, row, company):
    # update these fields by adding a slash
    company.industry       = special_combine(company.industry,       sheet[INDUSTRY       + str(row)].value)
    company.segment        = special_combine(company.segment,        sheet[SEGMENT        + str(row)].value)
    company.notes_parentco = special_combine(company.notes_parentco, sheet[NOTES_PARENTCO + str(row)].value)
    company.comments       = special_combine(company.comments,       sheet[COMMENTS       + str(row)].value)
    # if the new row contains more recent information than what is in the company, update
    # otherwise, for now, don't do anything
    
    print("CURRENT company last updated: ", company.last_updated)
    print("NEXT company last updated:    ", sheet[LAST_UPDATED + str(row)].value)
    if company.last_updated == "" or sheet[LAST_UPDATED + str(row)].value == "" or (sheet[LAST_UPDATED + str(row)].value != None and company.last_updated < sheet[LAST_UPDATED + str(row)].value):
        if sheet[   APPELLATION       + str(row)].value != None:
            company.appellation       = sheet[APPELLATION        + str(row)].value 
        if sheet[   PRIMARY_CONTACT   + str(row)].value != None:
            company.primary_contact   = sheet[PRIMARY_CONTACT    + str(row)].value
        if sheet[   COMPANY_NAME      + str(row)].value != None:
            company.company_name      = sheet[COMPANY_NAME       + str(row)].value
        if sheet[   SOURCE            + str(row)].value != None:
            company.source            = sheet[SOURCE             + str(row)].value  
        if sheet[   LAST_UPDATED      + str(row)].value != None:
            company.last_updated      = sheet[LAST_UPDATED       + str(row)].value  
        if sheet[   TITLE             + str(row)].value != None:
            company.title             = sheet[TITLE              + str(row)].value  
        if sheet[   CITY              + str(row)].value != None:
            company.city              = sheet[CITY               + str(row)].value  
        if sheet[   STATE             + str(row)].value != None:
            company.state             = sheet[STATE              + str(row)].value  
        if sheet[   ZIP_CODE          + str(row)].value != None:
            company.zip_code          = sheet[ZIP_CODE           + str(row)].value  
        if sheet[   REGION            + str(row)].value != None:
            company.region            = sheet[REGION             + str(row)].value  
        if sheet[   MAIN_PHONE        + str(row)].value != None:
            company.main_phone        = sheet[MAIN_PHONE         + str(row)].value  
        if sheet[   MOBILE_PHONE      + str(row)].value != None:
            company.mobile_phone      = sheet[MOBILE_PHONE       + str(row)].value  
        if sheet[   ADDRESS1          + str(row)].value != None:
            company.address1          = sheet[ADDRESS1           + str(row)].value  
        if sheet[   COUNTRY           + str(row)].value != None:
            company.country           = sheet[COUNTRY            + str(row)].value  
        if sheet[   A1_STREET         + str(row)].value != None:
            company.a1_street         = sheet[A1_STREET          + str(row)].value  
        if sheet[   A1_STREET2        + str(row)].value != None:
            company.a1_street2        = sheet[A1_STREET2         + str(row)].value  
        if sheet[   A1_STREET3        + str(row)].value != None:
            company.a1_street3        = sheet[A1_STREET3         + str(row)].value  
        if sheet[   SECONDARY_ADDRESS + str(row)].value != None:
            company.secondary_address = sheet[SECONDARY_ADDRESS  + str(row)].value  
        if sheet[   EMAIL             + str(row)].value != None:
            company.email             = sheet[EMAIL              + str(row)].value  
        if sheet[   FAX               + str(row)].value != None:
            company.fax               = sheet[FAX                + str(row)].value  
        if sheet[   WEBSITE           + str(row)].value != None:
            company.website           = sheet[WEBSITE            + str(row)].value  
    return company
#}}}

# Combine Contacts
# combine_contacts(first, last)
#{{{
# all rows algorithm
# - the first row is automatically its own entity
# - for all other rows take combination of primary contact and current company we are looking at and compare edit distance to it
#   - if it is a match combine and keep looking through list until we find a bad match
#   - if distance is way off save old company info. store new company and keep going

def combine_contacts(fileName, sheetName, first, last):
    if printing:
        print("Opening...")
    # Open the file for editing
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("contacts")
    wb = openpyxl.load_workbook(fileName + ".xlsx")
    sheet = wb[sheetName]

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')

    # if 'sheet' appears randomly we can delete it
    rm = out.get_sheet_by_name('Sheet')
    out.remove_sheet(rm)

    # - create an object for a new primary contact and account name pair
    #   - store previous object in a new sheet
    #   - store all information here
    #   - look at next contact and see if it matches (edit distance)
    #     - if it is a match combine and keep going, otherwise repeat
    compare = ""
    current = ""
    primary_contact = ""
    company_name = ""
    primary_contact_col = "C"
    company_name_col = "D"
    count = 1
    company = Company()

    # Create Headers
    #{{{
    if saving:
        outsheet[APPELLATION       + '1'].value          = "Appellation"
        outsheet[PRIMARY_CONTACT   + '1'].value          = "Primary_contact"
        outsheet[COMPANY_NAME      + '1'].value          = "Company"
        outsheet[INDUSTRY          + '1'].value          = "Industry"
        outsheet[SEGMENT           + '1'].value          = "Segment"
        outsheet[SOURCE            + '1'].value          = "Source"
        outsheet[LAST_UPDATED      + '1'].value          = "Last_updated"
        outsheet[TITLE             + '1'].value          = "Title"
        outsheet[CITY              + '1'].value          = "City"
        outsheet[STATE             + '1'].value          = "State"
        outsheet[ZIP_CODE          + '1'].value          = "Zip_code"
        outsheet[REGION            + '1'].value          = "Region"
        outsheet[MAIN_PHONE        + '1'].value          = "Main_phone"
        outsheet[MOBILE_PHONE      + '1'].value          = "Mobile_phone"
        outsheet[ADDRESS1          + '1'].value          = "Address1"
        outsheet[COUNTRY           + '1'].value          = "Country"
        outsheet[A1_STREET         + '1'].value          = "A1_street"
        outsheet[A1_STREET2        + '1'].value          = "A1_street"
        outsheet[A1_STREET3        + '1'].value          = "A1_street"
        outsheet[SECONDARY_ADDRESS + '1'].value          = "Secondary_address"
        outsheet[EMAIL             + '1'].value          = "Email"
        outsheet[NOTES_PARENTCO    + '1'].value          = "Notes_parentCo"
        outsheet[FAX               + '1'].value          = "Fax"
        outsheet[WEBSITE           + '1'].value          = "Website"
        outsheet[COMMENTS          + '1'].value          = "Comments"
    #}}}

    for row in range (first, last + 1):
        # if the previous value is blank we create the new object and store information in it
        primary_contact = str(sheet[primary_contact_col + str(row)].value)
        if primary_contact != "-":
            standardize_str(primary_contact)
        company_name = standardize_str(sheet[company_name_col + str(row)].value)
        if row == first:
            compare = primary_contact + " " + company_name
            company = new_company_from_sheet(sheet, row)
            count = count + 1
        else:
            current = primary_contact + " " + company_name
            match = edit_distance(compare, current, low_threshold, high_threshold)
            # combine information and move on
            if match:
                # combine information
                company = update_company_from_sheet(sheet, row, company)
            # store the information and create a new company
            else:
                # store information
                outsheet[APPELLATION       + str(count)].value = company.appellation
                outsheet[PRIMARY_CONTACT   + str(count)].value = company.primary_contact
                outsheet[COMPANY_NAME      + str(count)].value = company.company_name
                outsheet[INDUSTRY          + str(count)].value = company.industry
                outsheet[SEGMENT           + str(count)].value = company.segment
                outsheet[SOURCE            + str(count)].value = company.source
                outsheet[LAST_UPDATED      + str(count)].value = company.last_updated
                outsheet[TITLE             + str(count)].value = company.title
                outsheet[CITY              + str(count)].value = company.city
                outsheet[STATE             + str(count)].value = company.state
                outsheet[ZIP_CODE          + str(count)].value = company.zip_code
                outsheet[REGION            + str(count)].value = company.region
                outsheet[MAIN_PHONE        + str(count)].value = company.main_phone
                outsheet[MOBILE_PHONE      + str(count)].value = company.mobile_phone
                outsheet[COUNTRY           + str(count)].value = company.country
                outsheet[A1_STREET         + str(count)].value = company.a1_street
                outsheet[A1_STREET2        + str(count)].value = company.a1_street2
                outsheet[A1_STREET3        + str(count)].value = company.a1_street3
                outsheet[SECONDARY_ADDRESS + str(count)].value = company.secondary_address
                outsheet[EMAIL             + str(count)].value = company.email
                outsheet[NOTES_PARENTCO    + str(count)].value = company.notes_parentco
                outsheet[FAX               + str(count)].value = company.fax
                outsheet[WEBSITE           + str(count)].value = company.website
                outsheet[COMMENTS          + str(count)].value = company.comments
                # reset compare value
                compare = primary_contact + " " + company_name
                # create a new company
                company = new_company_from_sheet(sheet, row)
                count = count + 1


    if printing:
        print()
        print("Out of " + str(1 + last - first) + " companies " + str(count) + " were unique contacts")
        print("Saving...")

    out.save("newMOTU.xlsx")

    if printing:
        print()
        print("Done!")

    pygame.mixer.music.play()
    time.sleep(3)
    pygame.mixer.music.stop()
#}}}

# combine_contacts(2, 100)
combine_contacts("MOTU", "contacts", 2, 12655)
