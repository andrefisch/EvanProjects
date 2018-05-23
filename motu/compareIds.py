import openpyxl
import pygame
import time
import string

printing = True
low_threshold = 75
high_threshold = 100
min_word_len = 5

# Edit Distance Function
#{{{
# modified edit distance algorithm:
# - see if string one is a substring of the next string
#   - if it is there is a match
#   - it not look at edit distance for substring and substring of equal length
#     - if it is above a certain percentage or below a certain edit number we can safely assume they are a match. keep that string and try with next one

def edit_distance(word1, word2):
    if printing:
        print("Looking at '" + word1 + "' and '" + word2 + "'")
    word1 = word1.lower()
    word2 = word2.lower()
    len_1 = len(word1)
    len_2 = len(word2)

    # make sure shorter word is first
    if len_1 > len_2:
        word1, word2 = word2, word1
        len_1 = len_2

    if (len_1 >= min_word_len):
        if (word1 in word2):
            if printing:
                print ("Edit distance: 0")
                print ("Percent Match: 100")
            return (0, 100)
        else:
            # shorten longer word to length of first word
            word2 = word2[0:len_1]
            # the matrix whose last element -> edit distance
            x = [[0] * (len_1 + 1) for _ in range(len_1 + 1)]

            # initialization of base case values
            for i in range(0, len_1 + 1): 
                x[i][0] = i
            for j in range(0, len_1 + 1):
                x[0][j] = j
            for i in range (1, len_1 + 1):
                for j in range(1, len_1 + 1):
                    if word1[i - 1] == word2[j - 1]:
                        x[i][j] = x[i - 1][j - 1] 
                    else:
                        x[i][j]= min(x[i][j - 1], x[i - 1][j], x[i - 1][j - 1]) + 1
            edit_distance = x[i][j]
            percent_match = ((len_1 - edit_distance) / len_1) * 100
            if printing:
                print ("Edit distance " + str(x[i][j]))
                print ("Percent match: " + "%.2f" % percent_match)
            return (edit_distance, percent_match)
    return (0, 0)


#}}}

pygame.init()
pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')

# Scans both notebooks for duplicates and creates a copy of the second notebook with all of the duplicates removed
# bigNotebook = str
# littleNotebook = str
# outputName = str
# DO NOT include .xlsx at the end of any of any file names
def compareIds(bigNotebook, littleNotebook, outputName):
    '''
    A. Open both inputs and look for students sheet
    B. Loop through bigNotebook and put them into a dictionary with companyName as the key. Value doesnt matter
    C. Loop through littleNotebook and check for presence of companyName in dictionary
        1. If companyName exists
            a. clear this row
            b. increment duplicate count
    D. save file as outputName.xlsx
    '''
    
    if printing:
        print("Loading...")
    wb1 = openpyxl.load_workbook(bigNotebook + ".xlsx")
    wb2 = openpyxl.load_workbook(littleNotebook + ".xlsx")
    wb3 = openpyxl.Workbook()
    sheet1 = wb1.worksheets[0]
    sheet2 = wb2.worksheets[0]
    sheet3 = wb3.create_sheet("duplicates")
    start1 = 5
    end1  = sheet1.max_row + 1
    start2 = 2
    end2 = sheet2.max_row + 1

    names = {}
    nameList = []
    punctuationTable = str.maketrans({key: None for key in string.punctuation})

    if printing:
        print("Creating dictionary...")
    for row in range(start1, end1):
        companyName = str(sheet1['C' + str(row)].value).lower().translate(punctuationTable)
        # print (companyName)
        names[companyName] = row - 5
        nameList.append(companyName)

    duplicates = 0

    if printing:
        print("Counting Duplicates...")
    for row in range(start2, end2):
        companyName = str(sheet2['A' + str(row)].value).lower().translate(punctuationTable)
        # print("Checking " + companyName)
        if companyName in names:
            duplicates += 1
            companyName = companyName.title()
            if printing:
                print(str(duplicates) + ": " + companyName)
            sheet3['A' + str(duplicates)].value = companyName
    
    # - loop through sheet we just created
    #   - use dict to find names in list
    #   - loop until we dont have a match
    #   - record number of matches
    for row in range(1, sheet3.max_row + 1):
        count = 0
        curr = sheet3['A' + str(row)].value.lower()
        while True:
            index = names[curr] + count
            edit, match = edit_distance(curr, nameList[index])
            if match > low_threshold and match <= high_threshold:
                count = count + 1
                sheet3['B' + str(row)] = count
            else:
                break
        

    if printing:
        print("Saving...")
    sheet = wb3.get_sheet_by_name('Sheet')
    wb3.remove_sheet(sheet)
    wb3.save(outputName + ".xlsx")
    if printing:
        print("Done!")
    pygame.mixer.music.play()
    time.sleep(3)
    pygame.mixer.music.stop()
    return duplicates

print ("Found " + str(compareIds('netpas', 'bv', 'duplicates')) + " duplicate names")
