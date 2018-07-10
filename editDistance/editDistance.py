import openpyxl
import os
import pygame
import string
import sys
import time

printing = True
saving = True

min_word_len = 5

# Strip punctuation and lowercase a string
# standardize_str(word)
#{{{
punctuationTable = str.maketrans({key: None for key in string.punctuation})

def standardize_str(word):
    if word != None:
        word = str(word)
        return word.lower().translate(punctuationTable)
    else:
        return ""
#}}}

# replace_punct(word)
#{{{
def replace_punct(word):
    return word.replace('.', '').replace(',', '').replace('(', ' (').replace(')', ') ').replace('  ', ' ').strip()
#}}}

# Create a new group from sheet information
# new_group_from_sheet(sheet, row)
#{{{
def new_group_from_sheet(sheet, row):
    output = {}
    for cell in sheet[str(row)]:
        if cell.value != None:
            output[cell.column] = cell.value
        else:
            output[cell.column] = ""

    return output
#}}}

# Combine information in two groups
# combine_groups(sheet, dicta, dictb)
#{{{
def combine_groups(sheet, dicta, dictb):
    output = {}
    lastLetter = ""
    for cell in sheet['1']:
        lastLetter = cell.column
    output[chr(ord(lastLetter) + 1)] = ""
    for cell in sheet['1']:
        # if there is only information in second row
        if   dicta[cell.column] == "" and dictb[cell.column] != "":
            output[cell.column] = dictb[cell.column]
        # if there is only information in first row
        elif dicta[cell.column] != "" and dictb[cell.column] == "":
            output[cell.column] = dicta[cell.column]
        # if there is information in both rows take one and put the other in the notes section
        elif dicta[cell.column] != "" and dictb[cell.column] != "":
            output[cell.column] = dicta[cell.column]
            output[chr(ord(lastLetter) + 1)] = output[chr(ord(lastLetter) + 1)] + " " + str(cell.value) + " used to be " + str(dictb[cell.column]) + ";"
        else:
            output[cell.column] = ""
    return output
#}}}

# Edit Distance Function
# edit_distance(word1, word2, low_threshold, high_threshold)
#{{{
# modified edit distance algorithm:
# - see if string one is a substring of the next string
#   - if it is there is a match
#   - it not look at edit distance for substring and substring of equal length
#     - if it is above a certain percentage or below a certain edit number we can safely assume they are a match. keep that string and try with next one

def edit_distance(word1, word2, low_threshold, high_threshold):
    if printing:
        print("Looking at '" + word1 + "' and '" + word2 + "'")
    word1 = word1.lower()
    word2 = word2.lower()
    len_1 = len(word1)
    len_2 = len(word2)
    edit_distance, percent_match = 0, 0

    # make sure shorter word is first
    if len_1 > len_2:
        word1, word2 = word2, word1
        len_1 = len_2

    if (len_1 >= min_word_len):
        if (word1 in word2):
            if printing:
                print ("Edit distance: 0")
                print ("Percent Match: 100")
            edit_distance, percent_match = 0, 100
        else:
            # shorten longer word to length of first word
            word2 = word2[0:len_1]
            # the matrix whose last element -> edit distance
            x = [[0] * (len_1 + 1) for _ in range(len_1 + 1)]

            # initialization of base case values
            for i in range(0, len_1 + 1): 
                x[i][0] = i
            for j in range(0, len_1 + 1):
                x[0][j] = j
            for i in range (1, len_1 + 1):
                for j in range(1, len_1 + 1):
                    if word1[i - 1] == word2[j - 1]:
                        x[i][j] = x[i - 1][j - 1] 
                    else:
                        x[i][j]= min(x[i][j - 1], x[i - 1][j], x[i - 1][j - 1]) + 1
            edit_distance = x[i][j]
            percent_match = ((len_1 - edit_distance) / len_1) * 100
            if printing:
                print ("Edit distance " + str(x[i][j]))
                print ("Percent match: " + "%.2f" % percent_match)
    if percent_match > low_threshold and percent_match <= high_threshold:
        if printing:
            print("MATCH!")
        return True
    else:
        return False
#}}}

def edit_distance_spreadsheet(*args):
    # turn the arguments into variable groups
    args = args[0]
    fileName = args[1]
    low_threshold = int(args[2])
    cols = args[3:]
    if printing:
        print("Opening...")

    # Open file to read
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.worksheets[0]

    # Open the file for editing
    out = openpyxl.Workbook()
    # Open the worksheet we want to edit
    outsheet = out.create_sheet("groups")
    # if 'sheet' appears randomly we can delete it
    rm = out['Sheet']
    out.remove(rm)

    # Create a new file to store duplicate groups
    dupe = openpyxl.Workbook()
    # Open the worksheet we want to edit
    dupesheet = dupe.create_sheet("groups")
    # if 'sheet' appears randomly we can delete it
    rm = dupe['Sheet']
    dupe.remove(rm)

    # - create an object for a new primary group and account group pair
    #   - store previous object in a new sheet
    #   - store all information here
    #   - look at next group and see if it matches (edit distance)
    #     - if it is a match combine and keep going, otherwise repeat
    compare = ""
    current = ""
    last = sheet.max_row
    count = 1
    dupes = 2
    first = 2
    high_threshold = 100
    lastLetter = ""

    # Create Headers
    #{{{
    if saving:
        lastLetter = ""
        for cell in sheet['1']:
            lastLetter = cell.column
            # put the outsheet columns where they are supposed to be
            outsheet [cell.column + str(cell.row)].value = cell.value
            # move the duplicate columns over one so we can preserve original row number
            dupesheet[chr(ord(lastLetter) + 1) + str(cell.row)].value = cell.value
        # Create a new column called changes to document merged chages
        outsheet [chr(ord(lastLetter) + 1) + str(cell.row)].value = "Changes"
        dupesheet[chr(ord(lastLetter) + 1) + str(cell.row)].value = "Changes"

    #}}}

    for row in range (first, last + 1):
        compareCriteria = ""
        for col in cols:
            compareCriteria = compareCriteria + standardize_str(sheet[col + str(row)].value) + " "
        if row == first:
            compare = compareCriteria
            group = new_group_from_sheet(sheet, row)
            keep = {}
            count = count + 1
        else:
            current = compareCriteria
            match = edit_distance(compare, current, low_threshold, high_threshold)
            # matchingSuffixes = standardize_str(sheet[SUFFIX + str(row)].value) == standardize_str(sheet[SUFFIX + str(row - 1)].value)
            # combine information and move on
            if match:
                group = new_group_from_sheet(sheet, row)
                '''
                - if there is a duplicate
                - store previous item in list in duplicate list
                - store the duplicate in next spot in duplicate list
                '''
                # combine information
                # keep original
                dupes = dupes + 1

                dupegroup = new_group_from_sheet(sheet, row - 1)

                keep = combine_groups(sheet, group, dupegroup)

                #{{{
                dupesheet['A' + str(dupes)].value = row - 1
                for cell in sheet['1']:
                    lastLetter = cell.column
                    dupesheet[chr(ord(lastLetter) + 1) + str(dupes)].value = dupegroup[lastLetter]
                #}}}

                # keep duplicate
                dupes = dupes + 1

                dupegroup = new_group_from_sheet(sheet, row)
                #{{{
                dupesheet['A' + str(dupes)].value = row
                for cell in sheet['1']:
                    lastLetter = cell.column
                    dupesheet[chr(ord(lastLetter) + 1) + str(dupes)].value = dupegroup[lastLetter]
                #}}}

                dupes = dupes + 1

                # save the combined group
                #{{{
                for cell in outsheet['1']:
                    lastLetter = cell.column
                    dupesheet[chr(ord(lastLetter) + 1) + str(dupes)].value = keep[lastLetter]
                #}}}

                # create a blank space
                dupes = dupes + 1
            # store the information and create a new group
            else:
                # store information
                #{{{
                if keep:
                    for cell in outsheet['1']:
                        lastLetter = cell.column
                        outsheet[lastLetter + str(count)].value = keep[lastLetter]
                    keep.clear()
                else:
                    for cell in sheet['1']:
                        lastLetter = cell.column
                        outsheet[lastLetter + str(count)].value = group[lastLetter]
                #}}}
                # reset compare value
                compare = compareCriteria
                # create a new group
                group = new_group_from_sheet(sheet, row)
                count = count + 1

    if printing:
        print()
        print("Out of " + str(1 + last - first) + " companies " + str(count) + " were unique companies")
        print("Saving...")

    out.save("purged" + str(cols) + ".xlsx")
    dupe.save("duplicates" + str(cols) + ".xlsx")

    pygame.init()
    pygame.mixer.music.load('/home/andrefisch/python/evan/note.mp3')
    pygame.mixer.music.play()
    time.sleep(5)
    pygame.mixer.music.stop()

edit_distance_spreadsheet(sys.argv)
